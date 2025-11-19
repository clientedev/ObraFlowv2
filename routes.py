import os
import uuid
import io
import hashlib
import mimetypes
import traceback
from datetime import datetime, date, timedelta
from urllib.parse import urlparse
from flask import render_template, redirect, url_for, flash, request, current_app, send_from_directory, jsonify, make_response, session, Response, abort, send_file
from flask_login import login_user, logout_user, login_required, current_user
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from werkzeug.exceptions import HTTPException
from sqlalchemy.orm import joinedload

from app import app, db, csrf
from models import (
    User, Projeto, Relatorio, LegendaPredefinida, Visita, FotoRelatorio,
    Contato, ContatoProjeto, Reembolso, EnvioRelatorio, ChecklistTemplate,
    ComunicacaoVisita, EmailCliente, ChecklistPadrao,
    ChecklistObra, FuncionarioProjeto, AprovadorPadrao, ProjetoChecklistConfig,
    LogEnvioEmail, ConfiguracaoEmail, RelatorioExpress, FotoRelatorioExpress,
    VisitaParticipante, TipoObra, CategoriaObra, Notificacao
)

# ==========================================================================================
# PERMISSION HELPER - Centraliza verificação de permissões para edição de relatórios
# ==========================================================================================
def can_edit_report(user, relatorio):
    """
    Verifica se o usuário tem permissão para editar um relatório.
    
    Regras:
    - Usuários master: podem editar qualquer relatório
    - Usuários não-master: podem editar APENAS seus próprios relatórios em status editável
    
    Status editáveis: Rascunho, preenchimento, Rejeitado, Em edição, Aguardando Aprovação
    """
    # Master tem acesso total
    if user.is_master:
        return True
    
    # Não-master pode editar apenas seus próprios relatórios
    if relatorio.autor_id != user.id:
        return False
    
    # Verificar se o status permite edição
    status_editaveis = {'Rascunho', 'preenchimento', 'Rejeitado', 'Em edição', 'Aguardando Aprovação'}
    return relatorio.status in status_editaveis
# ==========================================================================================

# Health check endpoint for Railway deployment - LIGHTWEIGHT VERSION
@app.route('/health')
def health_check():
    """Health check robusto - versão definitiva"""
    try:
        # Testar conexão com banco
        legendas_count = LegendaPredefinida.query.filter_by(ativo=True).count()

        return jsonify({
            'message': 'Sistema de Gestão de Construção - ELP',
            'status': 'FUNCIONANDO',
            'mode': 'normal',
            'legendas_count': legendas_count,
            'database': 'connected',
            'timestamp': datetime.utcnow().isoformat(),
            'version': '1.0.1'
        }), 200

    except Exception as e:
        current_app.logger.error(f"Health check error: {e}")
        return jsonify({
            'message': 'Sistema de Gestão de Construção - ELP',
            'status': 'ERROR',
            'error': str(e),
            'timestamp': datetime.utcnow().isoformat(),
            'version': '1.0.1'
        }), 500

@app.route('/debug/images-data')
def debug_images_data():
    """Debug para verificar dados das imagens - PÚBLICO PARA TESTE"""
    try:
        # Informações do banco de dados
        db_url = app.config.get('SQLALCHEMY_DATABASE_URI', 'not set')[:100]
        
        # Buscar fotos
        fotos = FotoRelatorio.query.order_by(FotoRelatorio.created_at.desc()).limit(20).all()
        
        debug_data = {
            'database_url': db_url,
            'total_fotos': FotoRelatorio.query.count(),
            'fotos_com_imagem': FotoRelatorio.query.filter(FotoRelatorio.imagem.isnot(None)).count(),
            'fotos_sem_imagem': FotoRelatorio.query.filter(FotoRelatorio.imagem.is_(None)).count(),
            'fotos_recentes': []
        }
        
        for foto in fotos:
            debug_data['fotos_recentes'].append({
                'id': foto.id,
                'relatorio_id': foto.relatorio_id,
                'filename': foto.filename,
                'legenda': foto.legenda,
                'imagem_presente': foto.imagem is not None,
                'imagem_size': len(foto.imagem) if foto.imagem else 0,
                'created_at': foto.created_at.isoformat() if foto.created_at else None
            })
        
        return jsonify(debug_data)
    except Exception as e:
        return jsonify({'error': str(e), 'traceback': traceback.format_exc()}), 500

@app.route('/debug/reports-data')
@login_required
def debug_reports_data():
    """Debug para verificar dados dos relatórios"""
    if not current_user.is_master:
        return jsonify({'error': 'Acesso negado'}), 403
    
    try:
        # Testar conexão básica
        count = Relatorio.query.count()
        
        # Buscar alguns relatórios
        relatorios = Relatorio.query.order_by(Relatorio.created_at.desc()).limit(5).all()
        
        debug_data = {
            'total_relatorios': count,
            'relatorios_amostra': []
        }
        
        for rel in relatorios:
            debug_data['relatorios_amostra'].append({
                'id': rel.id,
                'numero': rel.numero,
                'titulo': rel.titulo,
                'status': rel.status,
                'projeto_id': rel.projeto_id,
                'autor_id': rel.autor_id,
                'created_at': rel.created_at.isoformat() if rel.created_at else None
            })
        
        return jsonify(debug_data)
        
    except Exception as e:
        return jsonify({
            'error': str(e),
            'type': type(e).__name__
        }), 500

    except Exception as e:
        current_app.logger.error(f"Health check error: {e}")
        return jsonify({
            'message': 'Sistema de Gestão de Construção - ELP',
            'status': 'ERROR',
            'error': str(e),
            'timestamp': datetime.utcnow().isoformat(),
            'version': '1.0.1'
        }), 500

@app.route('/debug/reports-status')
@login_required
def debug_reports_status():
    """Debug específico para problemas da rota /reports no Railway"""
    if not current_user.is_master:
        return jsonify({'error': 'Acesso negado'}), 403
    
    try:
        debug_info = {
            'database_status': 'unknown',
            'reports_table_exists': False,
            'reports_count': 0,
            'sample_reports': [],
            'user_info': {
                'id': current_user.id,
                'username': current_user.username,
                'is_master': current_user.is_master
            },
            'environment': 'railway' if os.environ.get('RAILWAY_ENVIRONMENT') else 'local',
            'errors': []
        }
        
        # Teste 1: Conexão com banco
        try:
            from sqlalchemy import text
            with db.engine.connect() as connection:
                result = connection.execute(text("SELECT 1")).scalar()
            debug_info['database_status'] = 'connected'
        except Exception as e:
            debug_info['database_status'] = f'error: {str(e)}'
            debug_info['errors'].append(f"Database connection: {str(e)}")
        
        # Teste 2: Verificar se tabela relatórios existe
        try:
            from sqlalchemy import text
            result = db.session.execute(text("SELECT COUNT(*) FROM information_schema.tables WHERE table_name = 'relatorios'")).scalar()
            debug_info['reports_table_exists'] = result > 0
        except Exception as e:
            debug_info['errors'].append(f"Table check: {str(e)}")
        
        # Teste 3: Contar relatórios
        try:
            debug_info['reports_count'] = Relatorio.query.count()
        except Exception as e:
            debug_info['errors'].append(f"Reports count: {str(e)}")
        
        # Teste 4: Buscar exemplos de relatórios
        try:
            sample_reports = Relatorio.query.limit(3).all()
            for report in sample_reports:
                debug_info['sample_reports'].append({
                    'id': report.id,
                    'numero': report.numero,
                    'titulo': report.titulo,
                    'status': report.status,
                    'created_at': report.created_at.isoformat() if report.created_at else None
                })
        except Exception as e:
            debug_info['errors'].append(f"Sample reports: {str(e)}")
        
        return jsonify(debug_info)
        
    except Exception as e:
        return jsonify({
            'error': 'Debug failed',
            'details': str(e)
        }), 500



@app.route('/health/reports')
def health_reports():
    """Health check específico para funcionalidade de relatórios"""
    try:
        health_data = {
            'status': 'healthy',
            'timestamp': datetime.utcnow().isoformat(),
            'reports_functional': False,
            'database_connected': False,
            'table_exists': False,
            'sample_data': False,
            'details': {}
        }
        
        # Teste 1: Conexão com banco
        try:
            from sqlalchemy import text
            db.session.execute(text('SELECT 1'))
            health_data['database_connected'] = True
            health_data['details']['database'] = 'connected'
        except Exception as e:
            health_data['details']['database'] = f'error: {str(e)}'
        
        # Teste 2: Tabela relatórios existe
        try:
            from sqlalchemy import text
            result = db.session.execute(text("SELECT COUNT(*) FROM information_schema.tables WHERE table_name = 'relatorios'")).scalar()
            health_data['table_exists'] = result > 0
            health_data['details']['table_check'] = f'exists: {result > 0}'
        except Exception as e:
            health_data['details']['table_check'] = f'error: {str(e)}'
        
        # Teste 3: Dados de exemplo
        try:
            count = Relatorio.query.count()
            health_data['sample_data'] = count >= 0
            health_data['details']['reports_count'] = count
        except Exception as e:
            health_data['details']['reports_count'] = f'error: {str(e)}'
        
        # Status geral
        health_data['reports_functional'] = (
            health_data['database_connected'] and 
            health_data['table_exists']
        )
        
        if not health_data['reports_functional']:
            health_data['status'] = 'unhealthy'
        
        status_code = 200 if health_data['reports_functional'] else 503
        return jsonify(health_data), status_code
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'timestamp': datetime.utcnow().isoformat(),
            'error': str(e)
        }), 500




@app.route('/health/full')
def health_check_full():
    """Full health check with database connectivity"""
    try:
        # Basic database connectivity test
        db.session.execute(db.text('SELECT 1'))
        return jsonify({
            'status': 'healthy',
            'timestamp': datetime.utcnow().isoformat(),
            'database': 'connected',
            'service': 'flask-app'
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'timestamp': datetime.utcnow().isoformat(),
            'database': 'disconnected',
            'error': str(e),
            'service': 'flask-app'
        }), 503


@app.route('/debug/reports-test')
@login_required
def debug_reports_test():
    """Debug específico para testar carregamento de relatórios"""
    if not current_user.is_master:
        return jsonify({'error': 'Acesso negado'}), 403
    
    try:
        # Testar diferentes queries
        results = {
            'basic_count': 0,
            'with_relationships': [],
            'raw_sql': [],
            'template_data': [],
            'errors': []
        }
        
        # Teste 1: Contagem básica
        try:
            results['basic_count'] = Relatorio.query.count()
            current_app.logger.info(f"✅ Contagem básica: {results['basic_count']}")
        except Exception as e:
            results['errors'].append(f"Contagem básica falhou: {str(e)}")
        
        # Teste 2: Query com relacionamentos
        try:
            relatorios = Relatorio.query.limit(5).all()
            for rel in relatorios:
                rel_data = {
                    'id': rel.id,
                    'numero': rel.numero,
                    'titulo': rel.titulo,
                    'projeto_id': rel.projeto_id,
                    'autor_id': rel.autor_id,
                    'status': rel.status
                }
                
                # Tentar buscar projeto
                if rel.projeto_id:
                    try:
                        projeto = Projeto.query.get(rel.projeto_id)
                        rel_data['projeto_nome'] = projeto.nome if projeto else 'Projeto não encontrado'
                    except Exception as pe:
                        rel_data['projeto_erro'] = str(pe)
                
                # Tentar buscar autor
                if rel.autor_id:
                    try:
                        autor = User.query.get(rel.autor_id)
                        rel_data['autor_nome'] = autor.nome_completo if autor else 'Autor não encontrado'
                    except Exception as ae:
                        rel_data['autor_erro'] = str(ae)
                
                results['with_relationships'].append(rel_data)
                
        except Exception as e:
            results['errors'].append(f"Query com relacionamentos falhou: {str(e)}")
        
        # Teste 3: SQL raw
        try:
            raw_results = db.session.execute(
                db.text("SELECT id, numero, titulo, projeto_id, autor_id, status FROM relatorios LIMIT 5")
            ).fetchall()
            
            for row in raw_results:
                results['raw_sql'].append({
                    'id': row.id,
                    'numero': row.numero,
                    'titulo': row.titulo,
                    'projeto_id': row.projeto_id,
                    'autor_id': row.autor_id,
                    'status': row.status
                })
                
        except Exception as e:
            results['errors'].append(f"SQL raw falhou: {str(e)}")
        
        # Teste 4: Formato do template
        try:
            relatorios_template = []
            relatorios = Relatorio.query.limit(3).all()
            
            for relatorio in relatorios:
                projeto = Projeto.query.get(relatorio.projeto_id) if relatorio.projeto_id else None
                autor = User.query.get(relatorio.autor_id) if relatorio.autor_id else None
                
                relatorios_template.append({
                    'relatorio': {
                        'id': relatorio.id,
                        'numero': relatorio.numero,
                        'titulo': relatorio.titulo,
                        'status': relatorio.status
                    },
                    'projeto': {
                        'id': projeto.id if projeto else None,
                        'nome': projeto.nome if projeto else None
                    } if projeto else None,
                    'autor': {
                        'id': autor.id if autor else None,
                        'nome_completo': autor.nome_completo if autor else None
                    } if autor else None
                })
            
            results['template_data'] = relatorios_template
            
        except Exception as e:
            results['errors'].append(f"Formato template falhou: {str(e)}")
        
        return jsonify(results)
        
    except Exception as e:
        return jsonify({'error': f'Erro geral: {str(e)}'}), 500

@app.route('/debug/reports-connectivity')
@login_required
def debug_reports_connectivity():
    """Debug específico para conectividade da tabela de relatórios"""
    if not current_user.is_master:
        return jsonify({'error': 'Acesso negado'}), 403
    
    debug_info = {
        'database_status': 'unknown',
        'relatorios_table': 'unknown',
        'relatorios_count': 0,
        'sample_relatorio': None,
        'connection_test': 'unknown',
        'errors': []
    }
    
    try:
        # Teste 1: Conexão básica
        from sqlalchemy import text
        result = db.session.execute(text("SELECT 1")).scalar()
        debug_info['connection_test'] = 'OK' if result == 1 else 'FAILED'
        
        # Teste 2: Tabela relatórios existe
        table_check = db.session.execute(text("SELECT COUNT(*) FROM information_schema.tables WHERE table_name = 'relatorios'")).scalar()
        debug_info['relatorios_table'] = 'EXISTS' if table_check > 0 else 'MISSING'
        
        # Teste 3: Contar relatórios
        if table_check > 0:
            debug_info['relatorios_count'] = Relatorio.query.count()
            
            # Teste 4: Buscar um relatório de exemplo
            primeiro_relatorio = Relatorio.query.first()
            if primeiro_relatorio:
                debug_info['sample_relatorio'] = {
                    'id': primeiro_relatorio.id,
                    'numero': primeiro_relatorio.numero,
                    'titulo': primeiro_relatorio.titulo,
                    'status': primeiro_relatorio.status,
                    'created_at': primeiro_relatorio.created_at.isoformat() if primeiro_relatorio.created_at else None
                }
        
        debug_info['database_status'] = 'HEALTHY'
        
    except Exception as e:
        debug_info['errors'].append(f"Database error: {str(e)}")
        debug_info['database_status'] = 'ERROR'
    
    return jsonify(debug_info)

@app.route('/debug/db-test')
@login_required
def debug_db_test():
    """Debug endpoint para testar conexão com banco"""
    if not current_user.is_master:
        return jsonify({'error': 'Acesso negado'}), 403
    
    try:
        # Teste 1: Conexão básica
        from sqlalchemy import text
        result = db.session.execute(text("SELECT 1")).scalar()
        
        # Teste 2: Contar relatórios
        relatorios_count = Relatorio.query.count()
        
        # Teste 3: Contar projetos
        projetos_count = Projeto.query.count()
        
        # Teste 4: Contar usuários
        usuarios_count = User.query.count()
        
        # Teste 5: Buscar um relatório específico
        primeiro_relatorio = Relatorio.query.first()
        
        return jsonify({
            'status': 'success',
            'database_connection': 'OK',
            'test_query_result': result,
            'relatorios_count': relatorios_count,
            'projetos_count': projetos_count,
            'usuarios_count': usuarios_count,
            'primeiro_relatorio': {
                'id': primeiro_relatorio.id if primeiro_relatorio else None,
                'numero': primeiro_relatorio.numero if primeiro_relatorio else None,
                'titulo': primeiro_relatorio.titulo if primeiro_relatorio else None
            } if primeiro_relatorio else None,
            'database_url': app.config.get('SQLALCHEMY_DATABASE_URI', 'Not set')[:50] + '...'
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'error': str(e),
            'database_connection': 'FAILED'
        }), 500


    """Full health check with database connectivity"""
    try:
        # Basic database connectivity test
        db.session.execute(db.text('SELECT 1'))
        return jsonify({
            'status': 'healthy',
            'timestamp': datetime.utcnow().isoformat(),
            'database': 'connected',
            'service': 'flask-app'
        }), 200
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'timestamp': datetime.utcnow().isoformat(),
            'database': 'disconnected',
            'error': str(e),
            'service': 'flask-app'
        }), 503

from models import User, Projeto, Contato, ContatoProjeto, Visita, Relatorio, FotoRelatorio, Reembolso, EnvioRelatorio, ChecklistTemplate, ChecklistItem, ComunicacaoVisita, EmailCliente, ChecklistPadrao, LogEnvioEmail, ConfiguracaoEmail, RelatorioExpress, FotoRelatorioExpress, LegendaPredefinida, FuncionarioProjeto, AprovadorPadrao, ProjetoChecklistConfig, ChecklistObra, VisitaParticipante, Notificacao
from forms import LoginForm, RegisterForm, UserForm, ProjetoForm, VisitaForm, VisitaRealizadaForm, EmailClienteForm, RelatorioForm, FotoRelatorioForm, ReembolsoForm, ContatoForm, ContatoProjetoForm, LegendaPredefinidaForm, FirstLoginForm
from forms_email import ConfiguracaoEmailForm, EnvioEmailForm
from forms_express import RelatorioExpressForm, FotoExpressForm, EditarFotoExpressForm
from pdf_generator_express import gerar_pdf_relatorio_express, gerar_numero_relatorio_express
from utils import generate_project_number, generate_report_number, generate_visit_number, send_report_email, calculate_reimbursement_total, get_coordinates_from_address
from pdf_generator import generate_visit_report_pdf
from google_drive_backup import backup_to_drive, test_drive_connection
import math
import json

# Função helper para verificar se usuário é aprovador
def current_user_is_aprovador(projeto_id=None):
    """Verifica se o usuário atual é aprovador para um projeto específico ou globalmente"""
    try:
        if not current_user or not current_user.is_authenticated:
            return False

        # Se é master, automaticamente é aprovador
        if current_user.is_master:
            return True

        from models import AprovadorPadrao

        # Primeiro verifica se há configuração específica para o projeto
        if projeto_id:
            try:
                aprovador_especifico = AprovadorPadrao.query.filter_by(
                    projeto_id=projeto_id,
                    aprovador_id=current_user.id,
                    ativo=True
                ).first()
                if aprovador_especifico:
                    return True
            except Exception as e:
                current_app.logger.error(f"Erro ao verificar aprovador específico: {str(e)}")

        # Se não há configuração específica, verifica configuração global
        try:
            aprovador_global = AprovadorPadrao.query.filter_by(
                projeto_id=None,
                aprovador_id=current_user.id,
                ativo=True
            ).first()
            return aprovador_global is not None
        except Exception as e:
            current_app.logger.error(f"Erro ao verificar aprovador global: {str(e)}")
            return False

    except Exception as e:
        current_app.logger.error(f"Erro geral na verificação de aprovador: {str(e)}")
        return False

# Context processor para disponibilizar função nos templates
@app.context_processor
def inject_approval_functions():
    return {
        'current_user_is_aprovador': current_user_is_aprovador
    }

@app.route('/api/checklist-padrao')
def api_checklist_padrao():
    """API para carregar itens do checklist padrão"""
    try:
        # Forçar rollback para evitar transações pendentes
        try:
            db.session.rollback()
        except Exception:
            pass
        
        # Buscar itens ativos ordenados por ordem
        itens = ChecklistPadrao.query.filter_by(ativo=True).order_by(ChecklistPadrao.ordem).all()
        
        checklist_data = []
        for item in itens:
            checklist_data.append({
                'id': item.id,
                'texto': item.texto,
                'ordem': item.ordem,
                'ativo': item.ativo
            })
        
        current_app.logger.info(f"✅ API Checklist Padrão: {len(checklist_data)} itens retornados")
        
        return jsonify({
            'success': True,
            'checklist': checklist_data,
            'total': len(checklist_data)
        })
        
    except Exception as e:
        current_app.logger.error(f"❌ Erro ao buscar checklist padrão: {e}")
        return jsonify({
            'success': False,
            'error': str(e),
            'checklist': []
        }), 500

@app.route('/api/legendas')
def api_legendas():
    """API para carregar legendas pré-definidas do PostgreSQL Railway - VERSÃO CORRIGIDA"""
    try:
        categoria = request.args.get('categoria', 'all')
        current_app.logger.info(f"📋 API LEGENDAS: Buscando categoria='{categoria}'")

        # Forçar rollback para evitar transações pendentes
        try:
            db.session.rollback()
        except Exception:
            pass

        # Query básica sem usar numero_ordem (coluna não existe)
        query = LegendaPredefinida.query.filter_by(ativo=True)

        # Filtrar por categoria se especificado
        if categoria and categoria != 'all':
            query = query.filter_by(categoria=categoria)

        # Buscar legendas usando apenas campos que existem na tabela
        legendas_query = query.order_by(
            LegendaPredefinida.categoria.asc(),
            LegendaPredefinida.id.asc()
        ).all()

        # Converter para JSON usando apenas campos que existem
        legendas_data = []
        for legenda in legendas_query:
            legendas_data.append({
                'id': legenda.id,
                'texto': legenda.texto,
                'categoria': legenda.categoria,
                'ativo': legenda.ativo
            })

        current_app.logger.info(f"✅ API LEGENDAS: {len(legendas_data)} legendas retornadas (categoria={categoria})")

        # Resposta JSON final
        response_data = {
            'success': True,
            'legendas': legendas_data,
            'total': len(legendas_data),
            'fonte': 'railway_postgresql',
            'timestamp': datetime.utcnow().isoformat()
        }

        # Headers anti-cache
        response = jsonify(response_data)
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'

        return response

    except Exception as e:
        current_app.logger.error(f"❌ ERRO CRÍTICO API LEGENDAS: {str(e)}")
        # Forçar rollback em caso de erro
        try:
            db.session.rollback()
        except Exception:
            pass

        return jsonify({
            'success': False,
            'error': f'Erro interno: {str(e)}',
            'legendas': [],
            'total': 0,
            'fonte': 'error_definitivo',
            'timestamp': datetime.utcnow().isoformat()
        }), 500

# Debug routes para identificar diferenças de dados
@app.route('/api/current-user')
@login_required
def api_current_user():
    """Retorna dados do usuário atual para debug"""
    return jsonify({
        'id': current_user.id,
        'username': current_user.username,
        'email': current_user.email,
        'is_master': current_user.is_master,
        'nome_completo': getattr(current_user, 'nome_completo', None),
        'cargo': getattr(current_user, 'cargo', None)
    })

@app.route('/api/user-data-counts')
@login_required
def api_user_data_counts():
    """Retorna contadores de dados para o usuário atual"""
    # Se usuário for master, vê todos os dados
    if current_user.is_master:
        projetos = Projeto.query.count()
        relatorios = Relatorio.query.count()
        visitas = Visita.query.count()
        reembolsos = Reembolso.query.count() if 'Reembolso' in globals() else 0
    else:
        # Usuário normal vê apenas seus dados ou projetos relacionados
        projetos = Projeto.query.count()
        relatorios = Relatorio.query.count()
        visitas = Visita.query.count()
        reembolsos = 0

    return jsonify({
        'projetos': projetos,
        'relatorios': relatorios,
        'visitas': visitas,
        'reembolsos': reembolsos
    })

@app.route('/api/projeto/<int:projeto_id>/funcionarios-emails')
@login_required
def api_projeto_funcionarios_emails(projeto_id):
    """Retorna funcionários e e-mails de um projeto específico para seleção em relatórios"""
    try:
        current_app.logger.info(f"📡 API chamada: /api/projeto/{projeto_id}/funcionarios-emails por usuário {current_user.id}")
        
        projeto = Projeto.query.get_or_404(projeto_id)
        current_app.logger.info(f"✅ Projeto encontrado: {projeto.nome} (ID: {projeto.id})")

        # Verificação de autorização: usuário deve ter acesso ao projeto
        if not current_user.is_master:
            # Verificar se o usuário está associado ao projeto
            user_project_access = FuncionarioProjeto.query.filter_by(
                projeto_id=projeto_id,
                user_id=current_user.id,
                ativo=True
            ).first()

            # Se não for funcionário do projeto e não for responsável, negar acesso
            if not user_project_access and projeto.responsavel_id != current_user.id:
                current_app.logger.warning(f"🚫 Acesso negado para usuário {current_user.id} ao projeto {projeto_id}")
                return jsonify({
                    'success': False,
                    'error': 'Acesso negado ao projeto'
                }), 403

        # Buscar funcionários do projeto (tabela antiga - FuncionarioProjeto)
        funcionarios_antigos = FuncionarioProjeto.query.filter_by(
            projeto_id=projeto_id, 
            ativo=True
        ).all()
        current_app.logger.info(f"📋 Funcionários antigos encontrados: {len(funcionarios_antigos)}")

        # Buscar e-mails do projeto (tabela nova - EmailCliente, onde funcionários também são salvos)
        emails = EmailCliente.query.filter_by(
            projeto_id=projeto_id, 
            ativo=True
        ).all()
        current_app.logger.info(f"📧 Contatos encontrados: {len(emails)}")

        # Unificar funcionários de ambas as tabelas
        funcionarios_data = []
        
        # Adicionar funcionários da tabela antiga (FuncionarioProjeto)
        for func in funcionarios_antigos:
            func_data = {
                'id': f"fp_{func.id}",  # Prefixo para diferenciar origem
                'nome_funcionario': func.nome_funcionario or '',
                'cargo': func.cargo or '',
                'empresa': func.empresa or '',
                'is_responsavel_principal': func.is_responsavel_principal or False
            }
            funcionarios_data.append(func_data)
            current_app.logger.info(f"  📋 Funcionário (antigo): {func_data}")
        
        # Adicionar funcionários da tabela nova (EmailCliente)
        for email in emails:
            func_data = {
                'id': f"ec_{email.id}",  # Prefixo para diferenciar origem
                'nome_funcionario': email.nome_contato or '',
                'cargo': email.cargo or '',
                'empresa': email.empresa or '',
                'is_responsavel_principal': False
            }
            funcionarios_data.append(func_data)
            current_app.logger.info(f"  📋 Funcionário (novo): {func_data}")

        emails_data = []
        for email in emails:
            email_data = {
                'id': email.id,
                'email': email.email or '',
                'nome_contato': email.nome_contato or '',
                'cargo': email.cargo or ''
            }
            emails_data.append(email_data)
            current_app.logger.info(f"  📧 Email: {email_data}")

        response_data = {
            'success': True,
            'funcionarios': funcionarios_data,
            'emails': emails_data,
            'projeto_nome': projeto.nome
        }

        current_app.logger.info(f"✅ API retornando: {len(funcionarios_data)} funcionários, {len(emails_data)} e-mails")
        current_app.logger.info(f"📤 Response completo: {response_data}")

        return jsonify(response_data)

    except HTTPException as e:
        current_app.logger.error(f"❌ HTTPException: {e}")
        raise
    except Exception as e:
        current_app.logger.exception(f"❌ Erro CRÍTICO ao buscar funcionários e e-mails do projeto {projeto_id}")
        return jsonify({
            'success': False,
            'error': f'Erro interno: {str(e)}'
        }), 500

@app.route('/api/projeto/<int:projeto_id>/next-report-number')
@login_required
def api_next_report_number(projeto_id):
    """Retorna o próximo número de relatório disponível para um projeto"""
    try:
        current_app.logger.info(f"📡 API chamada: /api/projeto/{projeto_id}/next-report-number")
        
        # Verificar se o projeto existe
        projeto = Projeto.query.get_or_404(projeto_id)
        
        # Calculate next report number: max of (numeracao_inicial-1, highest existing numero_projeto) + 1
        numeracao_inicial = projeto.numeracao_inicial or 1
        max_numero_existente = db.session.query(
            db.func.max(Relatorio.numero_projeto)
        ).filter_by(projeto_id=projeto_id).scalar()
        
        if max_numero_existente is None:
            # No reports yet, use numeracao_inicial
            proximo_numero_projeto = numeracao_inicial
        else:
            # Ensure we never go below numeracao_inicial and always increment from max
            proximo_numero_projeto = max(numeracao_inicial - 1, max_numero_existente) + 1
        
        next_numero = f"REL-{proximo_numero_projeto:04d}"
        current_app.logger.info(f"✅ Próximo número para projeto {projeto_id}: {next_numero} (numeracao_inicial: {numeracao_inicial}, max_existente: {max_numero_existente})")
        
        return jsonify({
            'success': True,
            'next_numero': next_numero,
            'numero_projeto': proximo_numero_projeto
        })
        
    except HTTPException as e:
        current_app.logger.error(f"❌ HTTPException: {e}")
        raise
    except Exception as e:
        current_app.logger.exception(f"❌ Erro ao buscar próximo número do projeto {projeto_id}")
        return jsonify({
            'success': False,
            'error': f'Erro interno: {str(e)}'
        }), 500

@app.route('/relatorios/ultimo-lembrete')
@login_required
def ultimo_lembrete():
    """Retorna o lembrete do último relatório de uma obra"""
    try:
        obra_id = request.args.get('obra_id', type=int)
        
        if not obra_id:
            return jsonify({'lembrete': None}), 200
        
        # Buscar o relatório mais recente da obra que contenha lembrete_proxima_visita
        ultimo_relatorio = Relatorio.query.filter(
            Relatorio.projeto_id == obra_id,
            Relatorio.lembrete_proxima_visita.isnot(None),
            Relatorio.lembrete_proxima_visita != ''
        ).order_by(Relatorio.data_relatorio.desc()).first()
        
        if not ultimo_relatorio:
            return jsonify({'lembrete': None}), 200
        
        return jsonify({
            'lembrete': ultimo_relatorio.lembrete_proxima_visita,
            'relatorio_origem': ultimo_relatorio.numero,
            'data_relatorio': ultimo_relatorio.data_relatorio.strftime('%d/%m/%Y') if ultimo_relatorio.data_relatorio else None
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"❌ Erro ao buscar último lembrete: {e}")
        return jsonify({'lembrete': None}), 500

@app.route('/api/projetos/<int:projeto_id>/funcionarios')
@login_required
def get_funcionarios_projeto(projeto_id):
    """Retorna funcionários de um projeto específico"""
    try:
        funcionarios = FuncionarioProjeto.query.filter_by(projeto_id=projeto_id, ativo=True).all()
        return jsonify([{
            'id': f.id, 
            'nome': f.nome_funcionario,
            'cargo': f.cargo or '',
            'empresa': f.empresa or ''
        } for f in funcionarios])
    except Exception as e:
        current_app.logger.error(f"❌ Erro ao buscar funcionários do projeto {projeto_id}: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/projetos/<int:projeto_id>/emails')
@login_required
def get_emails_projeto(projeto_id):
    """Retorna e-mails de um projeto específico"""
    try:
        emails = EmailCliente.query.filter_by(projeto_id=projeto_id, ativo=True).all()
        return jsonify([{
            'email': e.email
        } for e in emails])
    except Exception as e:
        current_app.logger.error(f"❌ Erro ao buscar e-mails do projeto {projeto_id}: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/dashboard-stats')
@login_required
def api_dashboard_stats():
    """API para fornecer estatísticas reais do dashboard"""
    try:
        # Buscar dados reais do PostgreSQL
        projetos_ativos = Projeto.query.filter_by(status='Ativo').count()
        visitas_agendadas = Visita.query.filter_by(status='Agendada').count()
        relatorios_pendentes = Relatorio.query.filter(
            Relatorio.status.in_(['Rascunho', 'Aguardando Aprovação'])
        ).count()

        # Reembolsos com verificação de tabela
        try:
            usuarios_ativos = User.query.filter_by(ativo=True).count()
        except:
            usuarios_ativos = 0

        response_data = {
            'success': True,
            'projetos_ativos': projetos_ativos,
            'visitas_agendadas': visitas_agendadas,
            'relatorios_pendentes': relatorios_pendentes,
            'usuarios_ativos': usuarios_ativos,
            'timestamp': datetime.utcnow().isoformat(),
            'user_id': current_user.id,
            'source': 'postgresql'
        }

        # Headers para evitar cache
        response = jsonify(response_data)
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'

        return response

    except Exception as e:
        print(f"ERRO API DASHBOARD: {e}")
        return jsonify({
            'success': False,
            'error': str(e),
            'projetos_ativos': 6,
            'visitas_agendadas': 4,
            'relatorios_pendentes': 0,
            'usuarios_ativos': 2,
            'source': 'fallback'
        }), 500


        projetos = Projeto.query.count()  # Todos os projetos por enquanto
        relatorios = Relatorio.query.count()  # Todos os relatórios por enquanto
        visitas = Visita.query.count()
        reembolsos = 0

    return jsonify({
        'projetos': projetos,
        'relatorios': relatorios,
        'visitas': visitas,
        'reembolsos': reembolsos,
        'user_id': current_user.id,
        'is_master': current_user.is_master
    })

# Proxy route for reverse geocoding to avoid CORS issues
@app.route('/api/reverse-geocoding', methods=['POST'])
@login_required
def reverse_geocoding():
    """Proxy para reverse geocoding usando Nominatim - evita problemas de CORS"""
    try:
        # Verificar se o request tem JSON válido
        if not request.is_json:
            return jsonify({
                'success': False,
                'error': 'Content-Type deve ser application/json'
            }), 400

        data = request.get_json(force=True)
        
        # Validação mais robusta dos dados
        if not data:
            return jsonify({
                'success': False,
                'error': 'JSON vazio ou inválido'
            }), 400
            
        if 'latitude' not in data or 'longitude' not in data:
            return jsonify({
                'success': False,
                'error': 'Latitude e longitude são obrigatórias'
            }), 400

        try:
            lat = float(data['latitude'])
            lon = float(data['longitude'])
        except (ValueError, TypeError):
            return jsonify({
                'success': False,
                'error': 'Latitude e longitude devem ser números válidos'
            }), 400

        # Validar ranges das coordenadas
        if not (-90 <= lat <= 90):
            return jsonify({
                'success': False,
                'error': 'Latitude deve estar entre -90 e 90'
            }), 400
            
        if not (-180 <= lon <= 180):
            return jsonify({
                'success': False,
                'error': 'Longitude deve estar entre -180 e 180'
            }), 400

        # Fazer requisição para Nominatim através do servidor
        import requests
        url = f"https://nominatim.openstreetmap.org/reverse"
        params = {
            'format': 'json',
            'lat': lat,
            'lon': lon,
            'addressdetails': 1,
            'language': 'pt-BR'
        }
        headers = {
            'User-Agent': 'ELP-Sistema-Relatorios/1.0 (https://elpconsultoria.pro)'
        }

        current_app.logger.info(f"🌍 Fazendo reverse geocoding para: {lat}, {lon}")

        response = requests.get(url, params=params, headers=headers, timeout=15)
        
        if response.status_code == 200:
            try:
                geocoding_data = response.json()
                
                # Formatar endereço
                formatted_address = ''
                if geocoding_data and geocoding_data.get('address'):
                    addr = geocoding_data['address']
                    address_parts = []

                    if addr.get('house_number') and addr.get('road'):
                        address_parts.append(f"{addr['road']}, {addr['house_number']}")
                    elif addr.get('road'):
                        address_parts.append(addr['road'])

                    if addr.get('suburb') or addr.get('neighbourhood'):
                        address_parts.append(addr.get('suburb') or addr.get('neighbourhood'))

                    city = addr.get('city') or addr.get('town') or addr.get('village')
                    if city:
                        state = addr.get('state')
                        if state:
                            address_parts.append(f"{city} - {state}")
                        else:
                            address_parts.append(city)

                    formatted_address = ', '.join(filter(None, address_parts))
                    if not formatted_address:
                        formatted_address = geocoding_data.get('display_name', '')

                current_app.logger.info(f"✅ Endereço obtido: {formatted_address}")

                return jsonify({
                    'success': True,
                    'endereco': formatted_address,
                    'raw_data': geocoding_data
                })
                
            except ValueError as json_error:
                current_app.logger.error(f"❌ Erro ao parsear JSON do Nominatim: {json_error}")
                return jsonify({
                    'success': False,
                    'error': 'Resposta inválida do serviço de geocoding'
                }), 500
                
        elif response.status_code == 429:
            return jsonify({
                'success': False,
                'error': 'Muitas requisições. Tente novamente em alguns segundos.'
            }), 429
        else:
            current_app.logger.error(f"❌ Nominatim retornou status: {response.status_code}")
            return jsonify({
                'success': False,
                'error': f'Serviço de geocoding indisponível (código {response.status_code})'
            }), 500

    except requests.exceptions.Timeout:
        current_app.logger.error("❌ Timeout na requisição para Nominatim")
        return jsonify({
            'success': False,
            'error': 'Timeout ao obter endereço. Tente novamente.'
        }), 500
    except requests.exceptions.ConnectionError:
        current_app.logger.error("❌ Erro de conexão com Nominatim")
        return jsonify({
            'success': False,
            'error': 'Erro de conexão com serviço de endereços'
        }), 500
    except requests.exceptions.RequestException as e:
        current_app.logger.error(f"❌ Erro de rede: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Erro de rede: {str(e)}'
        }), 500
    except Exception as e:
        current_app.logger.error(f"❌ Erro crítico no reverse geocoding: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Erro interno do servidor'
        }), 500

def calculate_distance(lat1, lon1, lat2, lon2):
    """Calculate distance between two points using Haversine formula"""
    import math
    
    lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])
    
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
    c = 2 * math.asin(math.sqrt(a))
    
    r = 6371
    
    return r * c

@app.route('/api/nearby-projects')
def get_nearby_projects():
    """Get ALL projects ordered by distance from user location"""
    try:
        lat = request.args.get('lat', type=float)
        lon = request.args.get('lon', type=float)
        
        projects = Projeto.query.all()
        
        all_projects = []
        projects_with_distance = []
        projects_without_distance = []
        
        for project in projects:
            project_data = {
                'id': project.id,
                'nome': project.nome,
                'endereco': project.endereco or 'Endereço não informado',
                'status': project.status,
                'tipo_obra': project.tipo_obra,
                'latitude': project.latitude,
                'longitude': project.longitude,
                'numeracao_inicial': project.numeracao_inicial
            }
            
            if lat and lon and project.latitude and project.longitude:
                distance = calculate_distance(lat, lon, project.latitude, project.longitude)
                project_data['distance'] = round(distance, 2)
                projects_with_distance.append(project_data)
            else:
                project_data['distance'] = 'N/A'
                projects_without_distance.append(project_data)
        
        projects_with_distance.sort(key=lambda x: x['distance'])
        
        projects_without_distance.sort(key=lambda x: x['nome'])
        
        all_projects = projects_with_distance + projects_without_distance
        
        return jsonify({'success': True, 'projects': all_projects, 'total': len(all_projects)})
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/notificacoes')
@login_required
def listar_notificacoes():
    """Listar notificações do usuário autenticado (filtrando expiradas)"""
    try:
        from notification_service import notification_service
        from datetime import datetime
        
        try:
            # Filtrar notificações não expiradas (expires_at > agora OU expires_at é NULL)
            agora = datetime.utcnow()
            notificacoes = Notificacao.query.filter(
                Notificacao.user_id == current_user.id,
                db.or_(
                    Notificacao.expires_at > agora,
                    Notificacao.expires_at == None
                )
            ).order_by(Notificacao.created_at.desc()).all()
        except Exception as db_error:
            current_app.logger.error(f"❌ Erro SQL ao buscar notificações: {db_error}")
            current_app.logger.error(f"❌ Stack trace: {traceback.format_exc()}")
            
            if 'UndefinedColumn' in str(type(db_error).__name__):
                current_app.logger.error(f"❌ Coluna indefinida detectada. Verificar schema do banco de dados.")
                return jsonify({
                    'success': False, 
                    'error': 'Erro de schema no banco de dados. Por favor, contate o administrador.',
                    'details': str(db_error)
                }), 500
            raise
        
        notificacoes_json = []
        nao_lidas = 0
        
        for notif in notificacoes:
            if notif.status == 'nova':
                nao_lidas += 1
            
            notificacoes_json.append({
                'id': notif.id,
                'titulo': notif.titulo,
                'mensagem': notif.mensagem,
                'tipo': notif.tipo,
                'icone': notification_service.get_icone_tipo(notif.tipo),
                'status': notif.status,
                'link_destino': notif.link_destino,
                'created_at': notif.created_at.isoformat() if notif.created_at else None,
                'lida_em': notif.lida_em.isoformat() if notif.lida_em else None
            })
        
        return jsonify({
            'success': True,
            'notificacoes': notificacoes_json,
            'total': len(notificacoes_json),
            'nao_lidas': nao_lidas
        })
    
    except Exception as e:
        current_app.logger.error(f"❌ Erro ao listar notificações: {e}")
        current_app.logger.error(f"❌ Stack trace completo: {traceback.format_exc()}")
        return jsonify({'success': False, 'error': 'Erro ao carregar notificações. Tente novamente.'}), 500

@app.route('/api/notificacoes/marcar-lida', methods=['POST'])
@login_required
def marcar_notificacao_lida():
    """Marcar uma notificação como lida"""
    try:
        from notification_service import notification_service
        
        data = request.get_json()
        notificacao_id = data.get('notificacao_id')
        
        if not notificacao_id:
            return jsonify({'success': False, 'error': 'ID da notificação não fornecido'}), 400
        
        resultado = notification_service.marcar_como_lida(notificacao_id, current_user.id)
        
        if resultado['success']:
            return jsonify({'success': True})
        else:
            return jsonify(resultado), 404
    
    except Exception as e:
        current_app.logger.error(f"❌ Erro ao marcar notificação como lida: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/notificacoes/marcar-todas-lidas', methods=['POST'])
@login_required
def marcar_todas_notificacoes_lidas():
    """Marcar todas as notificações do usuário como lidas"""
    try:
        from notification_service import notification_service
        
        resultado = notification_service.marcar_todas_como_lidas(current_user.id)
        
        return jsonify(resultado)
    
    except Exception as e:
        current_app.logger.error(f"❌ Erro ao marcar todas as notificações como lidas: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/notificacoes/salvar-token', methods=['POST'])
@login_required
def salvar_fcm_token():
    """Salvar o FCM token do usuário para push notifications"""
    try:
        data = request.get_json()
        fcm_token = data.get('fcm_token')
        
        if not fcm_token:
            return jsonify({'success': False, 'error': 'FCM token não fornecido'}), 400
        
        current_user.fcm_token = fcm_token
        db.session.commit()
        
        current_app.logger.info(f"✅ FCM token salvo para usuário {current_user.username}")
        
        return jsonify({'success': True})
    
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"❌ Erro ao salvar FCM token: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/notificacoes/<int:notificacao_id>/ler', methods=['PUT'])
@login_required
def marcar_notificacao_lida_put(notificacao_id):
    """Marcar uma notificação como lida usando PUT"""
    try:
        from notification_service import notification_service
        
        resultado = notification_service.marcar_como_lida(notificacao_id, current_user.id)
        
        if resultado['success']:
            return jsonify({'success': True})
        else:
            return jsonify(resultado), 404
    
    except Exception as e:
        current_app.logger.error(f"❌ Erro ao marcar notificação como lida: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/update_fcm_token', methods=['POST'])
@login_required
def update_fcm_token():
    """
    Endpoint Firebase FCM - Atualizar token de push notification
    Compatível com Firebase Cloud Messaging SDK
    """
    try:
        data = request.get_json()
        token = data.get('fcm_token')
        
        if not token:
            return jsonify({'error': 'Token ausente'}), 400
        
        current_user.fcm_token = token
        db.session.commit()
        
        current_app.logger.info(f"✅ FCM token atualizado para usuário {current_user.username}")
        
        return jsonify({'success': True, 'message': 'Token FCM atualizado com sucesso'}), 200
    
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"❌ Erro ao atualizar FCM token: {e}")
        return jsonify({'error': str(e)}), 500

# Save location route for geolocation tracking
@app.route('/save_location', methods=['POST'])
@login_required
def save_location():
    """Salvar localização do usuário - capturada por GPS ou IP"""
    try:
        # Verificar se o request tem JSON válido
        if not request.is_json:
            return jsonify({
                'status': 'error',
                'message': 'Content-Type deve ser application/json'
            }), 400
        
        data = request.get_json()
        
        # Validar dados
        if not data:
            return jsonify({
                'status': 'error',
                'message': 'Dados vazios ou inválidos'
            }), 400
        
        lat = data.get('lat')
        lng = data.get('lng')
        
        # Validação de latitude e longitude
        if not lat or not lng:
            return jsonify({
                'status': 'error',
                'message': 'Localização inválida: latitude e longitude são obrigatórias'
            }), 400
        
        try:
            lat = float(lat)
            lng = float(lng)
        except (ValueError, TypeError):
            return jsonify({
                'status': 'error',
                'message': 'Latitude e longitude devem ser números válidos'
            }), 400
        
        # Validar ranges
        if not (-90 <= lat <= 90):
            return jsonify({
                'status': 'error',
                'message': 'Latitude deve estar entre -90 e 90'
            }), 400
        
        if not (-180 <= lng <= 180):
            return jsonify({
                'status': 'error',
                'message': 'Longitude deve estar entre -180 e 180'
            }), 400
        
        # Dados adicionais opcionais
        accuracy = data.get('accuracy')
        source = data.get('source', 'gps')  # 'gps' ou 'ip'
        address = data.get('address')
        projeto_id = data.get('projeto_id')
        relatorio_id = data.get('relatorio_id')
        
        # Log da localização
        current_app.logger.info(
            f"📍 Localização recebida de {current_user.username}: "
            f"Lat={lat}, Lng={lng}, Source={source}, "
            f"Accuracy={accuracy}m, Address={address}"
        )
        
        # Retornar sucesso
        return jsonify({
            'status': 'success',
            'message': 'Localização salva com sucesso',
            'lat': lat,
            'lng': lng,
            'source': source,
            'user': current_user.username
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"❌ Erro ao salvar localização: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Erro ao salvar localização: {str(e)}'
        }), 500


# ==================== ALTERNATIVE SIMPLE UPLOAD API ====================
@app.route('/api/upload-photo-simple', methods=['POST'])
@login_required
def api_upload_photo_simple():
    """
    API simplificada para upload de fotos - MÁXIMA COMPATIBILIDADE
    Aceita base64 ou multipart/form-data
    """
    try:
        current_app.logger.info(f"📸 ========== UPLOAD SIMPLE CHAMADO ==========")
        current_app.logger.info(f"📸 Method: {request.method}")
        current_app.logger.info(f"📸 Content-Type: {request.content_type}")
        current_app.logger.info(f"📸 Content-Length: {request.content_length}")
        
        # Tentar obter relatorio_id
        relatorio_id = None
        if request.is_json:
            data = request.get_json()
            relatorio_id = data.get('relatorio_id')
            current_app.logger.info(f"📸 JSON data keys: {list(data.keys())}")
        else:
            relatorio_id = request.form.get('relatorio_id')
            current_app.logger.info(f"📸 Form data keys: {list(request.form.keys())}")
            current_app.logger.info(f"📸 Files keys: {list(request.files.keys())}")
        
        if not relatorio_id:
            return jsonify({
                'success': False,
                'error': 'relatorio_id é obrigatório'
            }), 400
        
        # Verificar se relatório existe
        relatorio = Relatorio.query.get(int(relatorio_id))
        if not relatorio:
            return jsonify({
                'success': False,
                'error': 'Relatório não encontrado'
            }), 404
        
        # Obter dados da imagem
        image_data = None
        filename = 'photo.jpg'
        content_type = 'image/jpeg'
        
        # Método 1: Arquivo multipart
        if 'imagem' in request.files:
            file = request.files['imagem']
            image_data = file.read()
            filename = file.filename or 'photo.jpg'
            content_type = file.mimetype or 'image/jpeg'
            current_app.logger.info(f"📸 Imagem via FILES: {len(image_data)} bytes, tipo={content_type}")
        
        # Método 2: Base64 no JSON
        elif request.is_json and 'imagem_base64' in data:
            import base64
            b64_data = data['imagem_base64']
            if ',' in b64_data:
                b64_data = b64_data.split(',')[1]
            image_data = base64.b64decode(b64_data)
            filename = data.get('filename', 'photo.jpg')
            current_app.logger.info(f"📸 Imagem via BASE64 (JSON): {len(image_data)} bytes")
        
        # Método 3: Base64 no form
        elif 'imagem_base64' in request.form:
            import base64
            b64_data = request.form['imagem_base64']
            if ',' in b64_data:
                b64_data = b64_data.split(',')[1]
            image_data = base64.b64decode(b64_data)
            filename = request.form.get('filename', 'photo.jpg')
            current_app.logger.info(f"📸 Imagem via BASE64 (FORM): {len(image_data)} bytes")
        
        if not image_data:
            current_app.logger.error(f"📸 ERRO: Nenhuma imagem encontrada!")
            return jsonify({
                'success': False,
                'error': 'Nenhuma imagem fornecida'
            }), 400
        
        # Calcular hash
        import hashlib
        imagem_hash = hashlib.sha256(image_data).hexdigest()
        current_app.logger.info(f"📸 Hash calculado: {imagem_hash[:12]}...")
        
        # Criar registro da foto
        foto = FotoRelatorio()
        foto.relatorio_id = int(relatorio_id)
        foto.filename = f"{imagem_hash[:12]}_{filename}"
        foto.filename_original = filename
        foto.legenda = request.form.get('legenda') or request.json.get('legenda', 'Foto') if request.is_json else request.form.get('legenda', 'Foto')
        foto.descricao = request.form.get('descricao', '') if not request.is_json else request.json.get('descricao', '')
        foto.tipo_servico = request.form.get('categoria', 'Geral') if not request.is_json else request.json.get('categoria', 'Geral')
        foto.ordem = FotoRelatorio.query.filter_by(relatorio_id=relatorio_id).count() + 1
        
        # Salvar imagem binária e metadados
        foto.imagem = image_data
        foto.imagem_hash = imagem_hash
        foto.content_type = content_type
        foto.imagem_size = len(image_data)
        
        current_app.logger.info(f"📸 Salvando foto: legenda='{foto.legenda}', hash={imagem_hash[:12]}, size={len(image_data)}")
        
        db.session.add(foto)
        db.session.commit()
        
        current_app.logger.info(f"✅ FOTO SALVA COM SUCESSO! ID={foto.id}")
        
        return jsonify({
            'success': True,
            'message': 'Foto salva com sucesso',
            'foto_id': foto.id,
            'hash': imagem_hash,
            'size': len(image_data),
            'url': url_for('api_get_photo', foto_id=foto.id)
        }), 201
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"❌ ERRO UPLOAD SIMPLE: {str(e)}")
        import traceback
        current_app.logger.error(f"❌ Traceback: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

        # Validar dados
        if not data:
            return jsonify({
                'status': 'error',
                'message': 'Dados vazios ou inválidos'
            }), 400
        
        lat = data.get('lat')
        lng = data.get('lng')
        
        # Validação de latitude e longitude
        if not lat or not lng:
            return jsonify({
                'status': 'error',
                'message': 'Localização inválida: latitude e longitude são obrigatórias'
            }), 400
        
        try:
            lat = float(lat)
            lng = float(lng)
        except (ValueError, TypeError):
            return jsonify({
                'status': 'error',
                'message': 'Latitude e longitude devem ser números válidos'
            }), 400
        
        # Validar ranges
        if not (-90 <= lat <= 90):
            return jsonify({
                'status': 'error',
                'message': 'Latitude deve estar entre -90 e 90'
            }), 400
        
        if not (-180 <= lng <= 180):
            return jsonify({
                'status': 'error',
                'message': 'Longitude deve estar entre -180 e 180'
            }), 400
        
        # Dados adicionais opcionais
        accuracy = data.get('accuracy')
        source = data.get('source', 'gps')  # 'gps' ou 'ip'
        address = data.get('address')
        projeto_id = data.get('projeto_id')
        relatorio_id = data.get('relatorio_id')
        
        # Log da localização
        current_app.logger.info(
            f"📍 Localização recebida de {current_user.username}: "
            f"Lat={lat}, Lng={lng}, Source={source}, "
            f"Accuracy={accuracy}m, Address={address}"
        )
        
        # Aqui você pode salvar no banco de dados se necessário
        # Por exemplo, associar a um relatório ou projeto
        # if relatorio_id:
        #     relatorio = Relatorio.query.get(relatorio_id)
        #     if relatorio:
        #         relatorio.latitude = lat
        #         relatorio.longitude = lng
        #         db.session.commit()
        
        # Retornar sucesso
        return jsonify({
            'status': 'success',
            'message': 'Localização salva com sucesso',
            'lat': lat,
            'lng': lng,
            'source': source,
            'user': current_user.username
        }), 200
        
    except Exception as e:
        current_app.logger.error(f"❌ Erro ao salvar localização: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Erro ao salvar localização: {str(e)}'
        }), 500

# Authentication routes
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user and user.ativo and check_password_hash(user.password_hash, form.password.data):
            login_user(user, remember=form.remember_me.data)

            # Verificar se é o primeiro login
            if hasattr(user, 'primeiro_login') and user.primeiro_login:
                return redirect(url_for('first_login'))

            next_page = request.args.get('next')
            if not next_page or urlparse(next_page).netloc != '':
                next_page = url_for('index')
            return redirect(next_page)
        flash('Usuário ou senha inválidos.', 'error')

    return render_template('auth/login.html', form=form)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/register', methods=['GET', 'POST'])
@login_required
def register():
    if not current_user.is_master:
        flash('Acesso negado. Apenas usuários master podem cadastrar novos usuários.', 'error')
        return redirect(url_for('index'))

    form = RegisterForm()
    if form.validate_on_submit():
        if User.query.filter_by(username=form.username.data).first():
            flash('Nome de usuário já existe.', 'error')
            return render_template('auth/register.html', form=form)

        if User.query.filter_by(email=form.email.data).first():
            flash('Email já cadastrado.', 'error')
            return render_template('auth/register.html', form=form)

        user = User(
            username=form.username.data,
            email=form.email.data,
            nome_completo=form.nome_completo.data,
            cargo=form.cargo.data,
            telefone=form.telefone.data,
            password_hash=generate_password_hash(form.password.data),
            is_master=form.is_master.data
        )

        try:
            db.session.add(user)
            db.session.commit()
            flash('Usuário cadastrado com sucesso!', 'success')
            return redirect(url_for('users_list'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao cadastrar usuário: {str(e)}', 'error')

    return render_template('auth/register.html', form=form)

# ==================== NEW UNIFIED IMAGE UPLOAD API ====================
@app.route('/api/fotos/upload', methods=['POST'])
@login_required
@csrf.exempt
def api_upload_photo():
    """
    API unificada para upload de imagens via multipart/form-data
    Compatível com todos os dispositivos: iPhone, Android, Desktop, APK
    
    Aceita:
    - multipart/form-data com campo 'imagem' (arquivo binário)
    - metadados: relatorio_id, legenda, descricao, categoria, filename_original
    
    Recursos:
    - Calcula SHA-256 hash da imagem
    - Detecta e armazena content_type (MIME type)
    - Armazena tamanho da imagem
    - Detecta duplicatas por hash
    """
    try:
        # Log DETALHADO para debug
        current_app.logger.info(f"📸 ========== API UPLOAD CHAMADA ==========")
        current_app.logger.info(f"📸 Usuário: {current_user.id} ({current_user.username})")
        current_app.logger.info(f"📸 Content-Type: {request.content_type}")
        current_app.logger.info(f"📸 Form keys: {list(request.form.keys())}")
        current_app.logger.info(f"📸 Files keys: {list(request.files.keys())}")
        current_app.logger.info(f"📸 Headers: {dict(request.headers)}")
        
        # Validar que temos um arquivo
        if 'imagem' not in request.files:
            current_app.logger.error(f"❌ Campo 'imagem' não encontrado nos files")
            return jsonify({
                'success': False,
                'error': 'Campo "imagem" é obrigatório'
            }), 400
        
        file = request.files['imagem']
        
        # Validar que o arquivo não está vazio
        if not file or not file.filename:
            current_app.logger.error(f"❌ Arquivo vazio ou sem nome")
            return jsonify({
                'success': False,
                'error': 'Arquivo de imagem inválido'
            }), 400
        
        # Validar extensão
        allowed_extensions = {'.jpg', '.jpeg', '.png', '.gif', '.webp'}
        file_ext = os.path.splitext(file.filename.lower())[1]
        if file_ext not in allowed_extensions:
            current_app.logger.error(f"❌ Extensão não permitida: {file_ext}")
            return jsonify({
                'success': False,
                'error': f'Formato não suportado. Use: {", ".join(allowed_extensions)}'
            }), 400
        
        # Ler dados binários do arquivo
        file_bytes = file.read()
        file_size = len(file_bytes)
        
        current_app.logger.info(f"📸 Arquivo lido: {file.filename}, {file_size} bytes")
        
        # Validar tamanho (máximo 10MB)
        max_size = 10 * 1024 * 1024  # 10MB
        if file_size > max_size:
            current_app.logger.error(f"❌ Arquivo muito grande: {file_size} bytes")
            return jsonify({
                'success': False,
                'error': f'Arquivo muito grande. Máximo: {max_size // (1024*1024)}MB'
            }), 400
        
        if file_size == 0:
            current_app.logger.error(f"❌ Arquivo vazio")
            return jsonify({
                'success': False,
                'error': 'Arquivo de imagem está vazio'
            }), 400
        
        # Calcular hash SHA-256 da imagem
        imagem_hash = hashlib.sha256(file_bytes).hexdigest()
        current_app.logger.info(f"🔐 Hash calculado: {imagem_hash}")
        
        # Detectar content_type (MIME type)
        content_type = file.mimetype
        if not content_type or content_type == 'application/octet-stream':
            # Fallback: tentar detectar pelo nome do arquivo
            guessed_type, _ = mimetypes.guess_type(file.filename)
            content_type = guessed_type or 'image/jpeg'
        current_app.logger.info(f"📄 Content-Type detectado: {content_type}")
        
        # Obter metadados do form
        relatorio_id = request.form.get('relatorio_id')
        legenda = request.form.get('legenda', '').strip()
        descricao = request.form.get('descricao', '').strip()
        categoria = request.form.get('categoria', '').strip()
        local = request.form.get('local', '').strip()
        filename_original = request.form.get('filename_original', file.filename)
        
        # Validar relatorio_id
        if not relatorio_id:
            current_app.logger.error(f"❌ relatorio_id não fornecido")
            return jsonify({
                'success': False,
                'error': 'relatorio_id é obrigatório'
            }), 400
        
        try:
            relatorio_id = int(relatorio_id)
        except (ValueError, TypeError):
            current_app.logger.error(f"❌ relatorio_id inválido: {relatorio_id}")
            return jsonify({
                'success': False,
                'error': 'relatorio_id deve ser um número'
            }), 400
        
        # Verificar se o relatório existe e o usuário tem permissão
        relatorio = Relatorio.query.get(relatorio_id)
        if not relatorio:
            current_app.logger.error(f"❌ Relatório {relatorio_id} não encontrado")
            return jsonify({
                'success': False,
                'error': 'Relatório não encontrado'
            }), 404
        
        # Verificar permissão
        if not current_user.is_master and relatorio.autor_id != current_user.id:
            current_app.logger.error(f"❌ Usuário {current_user.id} sem permissão para relatório {relatorio_id}")
            return jsonify({
                'success': False,
                'error': 'Você não tem permissão para adicionar fotos a este relatório'
            }), 403
        
        # Checar se já existe uma foto com o mesmo hash para este relatório (evitar duplicatas)
        foto_existente = FotoRelatorio.query.filter_by(
            relatorio_id=relatorio_id,
            imagem_hash=imagem_hash
        ).first()
        
        if foto_existente:
            current_app.logger.warning(f"⚠️ Imagem duplicada detectada! Retornando foto existente ID={foto_existente.id}")
            return jsonify({
                'success': True,
                'message': 'Imagem já existe (duplicada)',
                'foto_id': foto_existente.id,
                'filename': foto_existente.filename,
                'file_size': foto_existente.imagem_size,
                'hash': foto_existente.imagem_hash,
                'is_duplicate': True,
                'url': url_for('api_get_photo', foto_id=foto_existente.id)
            }), 200
        
        # Validar campos obrigatórios
        if not legenda:
            current_app.logger.error(f"❌ Legenda não fornecida")
            return jsonify({
                'success': False,
                'error': 'Legenda é obrigatória'
            }), 400
        
        # Gerar nome único para o arquivo baseado no hash
        unique_filename = f"{imagem_hash}{file_ext}"
        
        # CRÍTICO: Criar registro da foto com dados binários ANTES de adicionar ao session
        foto = FotoRelatorio()
        foto.relatorio_id = relatorio_id
        foto.filename = unique_filename
        foto.filename_original = filename_original
        foto.legenda = legenda
        foto.descricao = descricao
        foto.tipo_servico = categoria
        foto.local = local
        foto.ordem = FotoRelatorio.query.filter_by(relatorio_id=relatorio_id).count() + 1
        
        # GARANTIR que os dados binários sejam bytes puros
        if isinstance(file_bytes, memoryview):
            file_bytes = bytes(file_bytes)
        elif not isinstance(file_bytes, bytes):
            file_bytes = bytes(file_bytes)
        
        # LOG DETALHADO PRÉ-SAVE
        current_app.logger.info(f"📊 PRÉ-SAVE: tipo={type(file_bytes).__name__}, tamanho={len(file_bytes)}, hash={imagem_hash[:12]}")
        
        # ATRIBUIR dados binários ao modelo SQLAlchemy
        foto.imagem = file_bytes
        foto.imagem_hash = imagem_hash
        foto.content_type = content_type
        foto.imagem_size = file_size
        
        # LOG: Verificar atribuição
        current_app.logger.info(f"💾 APÓS ATRIBUIÇÃO: foto.imagem type={type(foto.imagem).__name__}, size={len(foto.imagem) if foto.imagem else 0}")
        
        # Adicionar ao session
        db.session.add(foto)
        
        # FLUSH para obter ID
        try:
            db.session.flush()
            current_app.logger.info(f"🔄 FLUSH OK: foto.id={foto.id}, imagem_presente={foto.imagem is not None}")
        except Exception as flush_error:
            current_app.logger.error(f"❌ ERRO NO FLUSH: {flush_error}")
            db.session.rollback()
            return jsonify({
                'success': False,
                'error': f'Erro ao preparar salvamento: {str(flush_error)}'
            }), 500
        
        # COMMIT - Forçar salvamento no PostgreSQL
        try:
            db.session.commit()
            current_app.logger.info(f"✅ COMMIT EXECUTADO para foto.id={foto.id}")
        except Exception as commit_error:
            current_app.logger.error(f"❌ ERRO NO COMMIT: {commit_error}")
            db.session.rollback()
            return jsonify({
                'success': False,
                'error': f'Erro ao salvar no banco: {str(commit_error)}'
            }), 500
        
        # VERIFICAÇÃO PÓS-COMMIT - Buscar do banco novamente
        db.session.expire_all()  # Forçar reload do PostgreSQL
        foto_verificada = FotoRelatorio.query.get(foto.id)
        
        if not foto_verificada:
            current_app.logger.error(f"❌ CRÍTICO: Foto {foto.id} não encontrada após commit!")
            return jsonify({
                'success': False,
                'error': 'Foto não foi salva no banco de dados'
            }), 500
        
        # Verificar dados binários no PostgreSQL
        if foto_verificada.imagem:
            if isinstance(foto_verificada.imagem, (bytes, bytearray)):
                imagem_size_db = len(foto_verificada.imagem)
            elif isinstance(foto_verificada.imagem, memoryview):
                imagem_size_db = len(bytes(foto_verificada.imagem))
            else:
                imagem_size_db = 0
        else:
            imagem_size_db = 0
        
        current_app.logger.info(f"✅ VERIFICAÇÃO POSTGRESQL: foto.id={foto.id}, imagem_size_db={imagem_size_db}, imagem_size_original={file_size}")
        
        # VALIDAÇÃO FINAL
        if imagem_size_db == 0:
            current_app.logger.error(f"❌ FALHA: Imagem NÃO foi salva no PostgreSQL! foto.id={foto.id}")
            db.session.rollback()
            return jsonify({
                'success': False,
                'error': 'FALHA: Imagem não foi gravada no banco PostgreSQL'
            }), 500
        
        if imagem_size_db != file_size:
            current_app.logger.warning(f"⚠️ ATENÇÃO: Tamanho difere! Enviado={file_size}, PostgreSQL={imagem_size_db}")
        
        # DIAGNÓSTICO FINAL: Query SQL direta no PostgreSQL
        try:
            from sqlalchemy import text
            sql_check = text("SELECT id, LENGTH(imagem) as img_size FROM fotos_relatorio WHERE id = :foto_id")
            resultado = db.session.execute(sql_check, {'foto_id': foto.id}).fetchone()
            
            if resultado:
                sql_img_size = resultado.img_size if resultado.img_size else 0
                current_app.logger.info(f"🔍 SQL DIRETO: foto.id={foto.id}, LENGTH(imagem)={sql_img_size}")
                
                if sql_img_size == 0:
                    current_app.logger.error(f"❌ POSTGRESQL VAZIO! foto.id={foto.id} tem imagem NULL ou vazia")
            else:
                current_app.logger.error(f"❌ SQL DIRETO: foto.id={foto.id} não encontrada!")
        except Exception as sql_error:
            current_app.logger.error(f"❌ ERRO SQL DIRETO: {sql_error}")
        
        return jsonify({
            'success': True,
            'message': 'Imagem enviada com sucesso',
            'foto_id': foto.id,
            'filename': unique_filename,
            'file_size': file_size,
            'db_size': imagem_size_db,
            'hash': imagem_hash,
            'content_type': content_type,
            'is_duplicate': False,
            'url': url_for('api_get_photo', foto_id=foto.id)
        }), 201
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"❌ Erro ao fazer upload da foto: {str(e)}")
        import traceback
        current_app.logger.error(f"❌ Traceback: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'error': f'Erro ao processar upload: {str(e)}'
        }), 500

@app.route('/api/fotos/<int:foto_id>', methods=['GET'])
@login_required
def api_get_photo(foto_id):
    """
    API para recuperar imagem do banco de dados PostgreSQL
    Serve a imagem diretamente do campo BYTEA com Content-Type correto
    """
    try:
        foto = FotoRelatorio.query.get_or_404(foto_id)
        
        # Verificar se tem dados binários
        if not foto.imagem:
            current_app.logger.warning(f"⚠️ Foto {foto_id} sem dados binários no campo imagem")
            
            # Retornar imagem placeholder
            from flask import send_file
            import io
            placeholder = generate_placeholder_image(foto.filename)
            return send_file(
                io.BytesIO(placeholder),
                mimetype='image/png',
                as_attachment=False
            )
        
        # Usar content_type do banco de dados (prioridade)
        mimetype = foto.content_type or 'image/jpeg'
        
        # Fallback: se content_type não estiver no DB, detectar pelo filename
        if not foto.content_type and foto.filename:
            if foto.filename.lower().endswith('.png'):
                mimetype = 'image/png'
            elif foto.filename.lower().endswith('.gif'):
                mimetype = 'image/gif'
            elif foto.filename.lower().endswith('.webp'):
                mimetype = 'image/webp'
            else:
                mimetype = 'image/jpeg'
        
        current_app.logger.info(f"📤 Servindo foto {foto_id}: size={len(foto.imagem)} bytes, type={mimetype}")
        
        # CRÍTICO: Garantir que foto.imagem seja bytes
        image_data = foto.imagem
        if isinstance(image_data, memoryview):
            image_data = bytes(image_data)
        
        # Retornar a imagem com cabeçalhos corretos
        response = make_response(image_data)
        response.headers['Content-Type'] = mimetype
        response.headers['Content-Disposition'] = f'inline; filename="{foto.filename}"'
        response.headers['Cache-Control'] = 'public, max-age=31536000'
        response.headers['Content-Length'] = str(len(image_data))
        
        return response
        
    except Exception as e:
        current_app.logger.error(f"❌ Erro ao servir foto {foto_id}: {str(e)}")
        import traceback
        current_app.logger.error(f"❌ Traceback: {traceback.format_exc()}")
        
        # Retornar placeholder em caso de erro
        from flask import send_file
        import io
        placeholder = generate_placeholder_image(f"Erro: {foto_id}")
        return send_file(
            io.BytesIO(placeholder),
            mimetype='image/png',
            as_attachment=False
        )

# ==================== END NEW UNIFIED IMAGE UPLOAD API ====================

# Main routes
@app.route('/')
def index():
    # Se não estiver logado, redirecionar para login
    if not current_user.is_authenticated:
        return redirect(url_for('login'))

    try:
        # BUSCAR DADOS REAIS DO POSTGRESQL COM FALLBACK
        projetos_ativos = Projeto.query.filter_by(status='Ativo').count()
        visitas_agendadas = Visita.query.filter_by(status='Agendada').count()
        relatorios_pendentes = Relatorio.query.filter(
            Relatorio.status.in_(['Rascunho', 'Aguardando Aprovação'])
        ).count()

        # Usuários ativos
        try:
            usuarios_ativos = User.query.filter_by(ativo=True).count()
        except:
            usuarios_ativos = 0

        stats = {
            'projetos_ativos': projetos_ativos,
            'visitas_agendadas': visitas_agendadas,
            'relatorios_pendentes': relatorios_pendentes,
            'usuarios_ativos': usuarios_ativos
        }

        # Log para monitoramento
        print(f"REAL STATS FROM DB: P={projetos_ativos}, V={visitas_agendadas}, R={relatorios_pendentes}, U={usuarios_ativos}")

        # Get recent reports com fallback
        try:
            relatorios_recentes = Relatorio.query.order_by(Relatorio.created_at.desc()).limit(5).all()
        except:
            relatorios_recentes = []

    except Exception as e:
        # FALLBACK em caso de erro de conexão
        print(f"ERRO DB: {e} - Usando dados fallback")
        stats = {
            'projetos_ativos': 6,
            'visitas_agendadas': 4,
            'relatorios_pendentes': 0,
            'usuarios_ativos': 2
        }
        relatorios_recentes = []

    return render_template('dashboard_simple.html',
                         stats=stats,
                         relatorios_recentes=relatorios_recentes)

# User management routes
@app.route('/users')
@login_required
def users_list():
    if not current_user.is_master:
        flash('Acesso negado.', 'error')
        return redirect(url_for('index'))

    users = User.query.all()
    return render_template('users/list.html', users=users)

@app.route('/users/<int:user_id>/edit', methods=['GET', 'POST'])
@login_required
def user_edit(user_id):
    if not current_user.is_master:
        flash('Acesso negado.', 'error')
        return redirect(url_for('index'))

    user = User.query.get_or_404(user_id)
    form = UserForm(obj=user)

    if form.validate_on_submit():
        # Check for username conflicts
        existing_user = User.query.filter_by(username=form.username.data).first()
        if existing_user and existing_user.id != user.id:
            flash('Nome de usuário já existe.', 'error')
            return render_template('users/form.html', form=form, user=user)

        # Check for email conflicts
        existing_email = User.query.filter_by(email=form.email.data).first()
        if existing_email and existing_email.id != user.id:
            flash('Email já cadastrado.', 'error')
            return render_template('users/form.html', form=form, user=user)

        user.username = form.username.data
        user.email = form.email.data
        user.nome_completo = form.nome_completo.data
        user.cargo = form.cargo.data
        user.telefone = form.telefone.data
        user.is_master = form.is_master.data
        user.ativo = form.ativo.data

        if form.password.data:
            user.password_hash = generate_password_hash(form.password.data)

        try:
            db.session.commit()
            flash('Usuário atualizado com sucesso!', 'success')
            return redirect(url_for('users_list'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar usuário: {str(e)}', 'error')

    return render_template('users/form.html', form=form, user=user)

# Project management routes
@app.route('/projects')
@login_required
def projects_list():
    from geopy.distance import geodesic
    
    # Try to get user location from session or request
    user_lat = request.args.get('lat', type=float)
    user_lon = request.args.get('lon', type=float)

    # Get search query parameter
    q = request.args.get('q')

    # Start with base query
    query = Projeto.query

    # Apply intelligent search if query provided
    if q and q.strip():
        from sqlalchemy import or_
        search_term = f"%{q.strip()}%"
        query = query.filter(or_(
            Projeto.nome.ilike(search_term),
            Projeto.numero.ilike(search_term),
            Projeto.endereco.ilike(search_term),
            Projeto.construtora.ilike(search_term),
            Projeto.nome_funcionario.ilike(search_term),
            Projeto.descricao.ilike(search_term),
            Projeto.tipo_obra.ilike(search_term)
        ))

    projects = query.all()

    # Calculate distance for each project and prepare for sorting
    projects_with_distance = []
    for project in projects:
        # Calculate distance if user location and project coordinates are available
        if user_lat and user_lon and project.latitude and project.longitude:
            distance = geodesic(
                (user_lat, user_lon),
                (float(project.latitude), float(project.longitude))
            ).km
        else:
            # Projects without coordinates go to the end
            distance = float('inf')
        
        projects_with_distance.append((project, distance))

    # Sort by status (Ativo first) and then by distance
    projects_with_distance.sort(
        key=lambda x: (
            0 if x[0].status.lower() == 'ativo' else 1,
            x[1]
        )
    )

    # Extract only the sorted projects
    projects = [p[0] for p in projects_with_distance]

    return render_template('projects/list.html', projects=projects)

# Reports routes - Versão DEFINITIVA para PostgreSQL Railway
@app.route('/reports')
@login_required  
def reports():
    """Listar relatórios de obra - versão corrigida com paginação"""
    try:
        # Obter parâmetros de busca e paginação
        page = request.args.get('page', 1, type=int)
        search_query = request.args.get('q', '')
        per_page = 20  # Relatórios por página

        # Query básica com joins
        query = db.session.query(Relatorio).join(
            User, Relatorio.autor_id == User.id
        ).outerjoin(
            Projeto, Relatorio.projeto_id == Projeto.id
        )

        # Aplicar filtro de busca se fornecido
        if search_query and search_query.strip():
            from sqlalchemy import or_
            search_term = f"%{search_query.strip()}%"
            query = query.filter(or_(
                Relatorio.numero.ilike(search_term),
                Relatorio.titulo.ilike(search_term),
                Projeto.nome.ilike(search_term),
                User.nome_completo.ilike(search_term)
            ))

        # Ordenar por data de criação (mais recente primeiro)
        query = query.order_by(Relatorio.created_at.desc())

        # Aplicar paginação
        relatorios = query.paginate(
            page=page,
            per_page=per_page,
            error_out=False
        )

        current_app.logger.info(f"✅ Relatórios carregados: {relatorios.total} total, página {page}")
        return render_template("reports/list.html", relatorios=relatorios)

    except Exception as e:
        current_app.logger.exception(f"❌ Erro ao carregar relatórios: {str(e)}")
        
        # Fallback com SQL direto
        try:
            from sqlalchemy import text
            offset = (page - 1) * per_page if 'page' in locals() else 0
            
            sql_query = """
                SELECT r.*, p.nome as projeto_nome, u.nome_completo as autor_nome
                FROM relatorios r
                LEFT JOIN projetos p ON r.projeto_id = p.id
                LEFT JOIN users u ON r.autor_id = u.id
                ORDER BY r.created_at DESC
                LIMIT :limit OFFSET :offset
            """
            
            rows = db.session.execute(text(sql_query), {
                'limit': per_page if 'per_page' in locals() else 20,
                'offset': offset
            }).fetchall()

            # Criar objeto mock para paginação
            class MockPagination:
                def __init__(self, items, current_page=1, per_page=20):
                    self.items = items
                    self.total = len(items)
                    self.page = current_page
                    self.per_page = per_page
                    self.pages = max(1, (self.total // self.per_page) + (1 if self.total % self.per_page > 0 else 0))
                    self.has_prev = self.page > 1
                    self.has_next = self.page < self.pages

                def iter_pages(self):
                    """Método para compatibilidade com template"""
                    return range(1, min(self.pages + 1, 6))  # Máximo 5 páginas

            current_page = page if 'page' in locals() else 1
            per_page_value = per_page if 'per_page' in locals() else 20
            relatorios = MockPagination(rows, current_page, per_page_value)
            current_app.logger.warning(f"⚠️ Usando fallback SQL: {len(rows)} relatórios")
            return render_template("reports/list.html", relatorios=relatorios, fallback=True)

        except Exception as fallback_error:
            current_app.logger.error(f"❌ Fallback também falhou: {str(fallback_error)}")
            return render_template("reports/list.html", relatorios=MockPagination([]), fallback=True, error=str(e))


@app.route('/reports/autosave/<int:report_id>', methods=['POST'])
@login_required
def autosave_report(report_id):
    """
    Rota AJAX segura e idempotente para auto-save de relatórios
    Aceita JSON e atualiza apenas campos permitidos (whitelist)
    """
    try:
        current_app.logger.info(f"💾 AUTOSAVE: Usuário {current_user.username} salvando relatório {report_id}")

        # Verificar se o JSON é válido - usar silent=True conforme especificação
        data = request.get_json(silent=True)
        if not data:
            current_app.logger.error("❌ AUTOSAVE: JSON vazio ou inválido")
            return jsonify({"success": False, "error": "JSON vazio ou inválido"}), 400

        # Buscar o relatório
        relatorio = Relatorio.query.get(report_id)
        if not relatorio:
            current_app.logger.warning(f"⚠️ AUTOSAVE: Relatório {report_id} não encontrado")
            return jsonify({"success": False, "error": "Relatório não encontrado"}), 404

        # Verificar permissão (autor ou master)
        if relatorio.autor_id != current_user.id and not current_user.is_master:
            current_app.logger.warning(f"🚫 AUTOSAVE: Usuário {current_user.username} sem permissão para relatório {report_id}")
            return jsonify({"success": False, "error": "Sem permissão para editar este relatório"}), 403

        # Whitelist de campos permitidos conforme especificação
        allowed_fields = [
            'titulo', 'observacoes', 'latitude', 'longitude', 
            'endereco', 'checklist_data', 'last_edited_at',
            'descricao', 'categoria', 'local', 'observacoes_finais', 
            'conteudo', 'lembrete_proxima_visita', 'acompanhantes'
        ]

        # Aplicar updates apenas nos campos permitidos
        changes_made = False
        for field, value in data.items():
            if field in allowed_fields:
                # Validações específicas por campo
                if field == 'checklist_data':
                    # Validar se é um JSON válido
                    if value is not None:
                        try:
                            if isinstance(value, dict):
                                import json
                                value = json.dumps(value)
                            elif isinstance(value, str):
                                # Verificar se é JSON válido
                                json.loads(value)
                            else:
                                current_app.logger.warning(f"⚠️ AUTOSAVE: checklist_data tipo inválido: {type(value)}")
                                continue
                        except json.JSONDecodeError:
                            current_app.logger.warning(f"⚠️ AUTOSAVE: checklist_data JSON inválido")
                            continue
                
                elif field == 'lembrete_proxima_visita':
                    # Converter string ISO para datetime
                    if value is not None and value != '':
                        try:
                            if isinstance(value, str):
                                value = datetime.fromisoformat(value.replace('Z', '+00:00'))
                            elif not isinstance(value, datetime):
                                current_app.logger.warning(f"⚠️ AUTOSAVE: lembrete_proxima_visita tipo inválido: {type(value)}")
                                continue
                        except (ValueError, AttributeError) as e:
                            current_app.logger.warning(f"⚠️ AUTOSAVE: Erro ao converter lembrete_proxima_visita: {e}")
                            continue
                    else:
                        value = None
                
                elif field == 'acompanhantes':
                    # Validar que seja uma lista válida (JSONB)
                    if value is not None:
                        if isinstance(value, list):
                            # Já é uma lista, pode usar diretamente
                            pass
                        elif isinstance(value, str):
                            try:
                                value = json.loads(value)
                                if not isinstance(value, list):
                                    current_app.logger.warning(f"⚠️ AUTOSAVE: acompanhantes não é uma lista após parsing")
                                    continue
                            except json.JSONDecodeError as e:
                                current_app.logger.warning(f"⚠️ AUTOSAVE: Erro ao parsear acompanhantes JSON: {e}")
                                continue
                        else:
                            current_app.logger.warning(f"⚠️ AUTOSAVE: acompanhantes tipo inválido: {type(value)}")
                            continue
                    else:
                        value = None

                # Aplicar a mudança se o valor for diferente
                current_value = getattr(relatorio, field, None)
                if current_value != value:
                    setattr(relatorio, field, value)
                    changes_made = True
                    current_app.logger.info(f"📝 AUTOSAVE: Campo '{field}' atualizado")

        # Se houve mudanças, atualizar status conforme especificação
        if changes_made:
            # Se report.status != 'Aprovado', definir report.status = 'preenchimento'
            if relatorio.status != 'Aprovado':
                relatorio.status = 'preenchimento'
                current_app.logger.debug("📝 AUTOSAVE: Status alterado para 'preenchimento'")

            # Atualizar timestamp
            relatorio.updated_at = datetime.utcnow()

            # Commit com try/except e rollback
            try:
                db.session.commit()
                current_app.logger.info(f"✅ AUTOSAVE: Relatório {report_id} salvo com sucesso")
                return jsonify({
                    "success": True, 
                    "message": "Rascunho salvo automaticamente",
                    "status": relatorio.status,
                    "timestamp": relatorio.updated_at.isoformat()
                }), 200
            except Exception as e:
                db.session.rollback()
                current_app.logger.error(f"❌ AUTOSAVE: Erro ao salvar no banco - {str(e)}")
                return jsonify({"success": False, "error": "Erro ao salvar no banco de dados"}), 500
        else:
            current_app.logger.debug(f"🔄 AUTOSAVE: Nenhuma mudança detectada para relatório {report_id}")
            return jsonify({
                "success": True, 
                "message": "Nenhuma alteração para salvar",
                "status": relatorio.status
            }), 200

    except Exception as e:
        # Log completo do erro
        current_app.logger.exception(f"❌ AUTOSAVE CRÍTICO: Erro inesperado no relatório {report_id}")
        db.session.rollback()
        return jsonify({"success": False, "error": "Erro interno do servidor"}), 500

@app.route('/reports/new', methods=['GET', 'POST'])
@login_required
def create_report():
    # CORREÇÃO: Se edit=X na URL, redirecionar para a rota de edição completa
    edit_id = request.args.get('edit', type=int)
    if edit_id:
        current_app.logger.info(f"🔀 Redirecionando /reports/new?edit={edit_id} → /reports/{edit_id}/editarrel")
        return redirect(url_for('report_edit_complete', report_id=edit_id))
    
    # Verificar se há projeto pré-selecionado via URL
    preselected_project_id = request.args.get('projeto_id', type=int)
    disable_fields = bool(preselected_project_id)

    if request.method == 'POST':
        # Debug: Log dos dados recebidos
        current_app.logger.info(f"📝 FORM DATA RECEBIDO: {list(request.form.keys())}")
        current_app.logger.info(f"📁 FILES RECEBIDOS: {list(request.files.keys())}")

        # Processar dados do formulário diretamente
        projeto_id = request.form.get('projeto_id')
        titulo = request.form.get('titulo', 'Relatório de visita')
        conteudo = request.form.get('conteudo', '')
        aprovador_nome = request.form.get('aprovador_nome', '')
        data_relatorio_str = request.form.get('data_relatorio')

        # Validações básicas
        if not projeto_id:
            flash('Obra é obrigatória.', 'error')
            return redirect(url_for('create_report'))

        try:
            projeto_id = int(projeto_id)
            # Convert date string to datetime object
            if data_relatorio_str:
                data_relatorio = datetime.strptime(data_relatorio_str, '%Y-%m-%d')
            else:
                data_relatorio = datetime.now()
        except (ValueError, TypeError):
            flash('Dados inválidos no formulário.', 'error')
            return redirect(url_for('create_report'))
        try:
            # Check if we're editing an existing report
            edit_report_id = request.form.get('edit_report_id')
            if edit_report_id:
                # Update existing report
                relatorio = Relatorio.query.get(int(edit_report_id))
                if not relatorio:
                    flash('Relatório não encontrado para edição.', 'error')
                    return redirect(url_for('create_report'))
                
                # Check permissions
                if relatorio.autor_id != current_user.id and not current_user.is_master:
                    flash('Você não tem permissão para editar este relatório.', 'error')
                    return redirect(url_for('reports'))
                
                # Update fields
                relatorio.titulo = titulo
                relatorio.projeto_id = projeto_id
                relatorio.updated_at = datetime.utcnow()
                current_app.logger.info(f"📝 Updating existing report {relatorio.numero}")
            else:
                # Create new report
                relatorio = Relatorio()
                
                # Check if numero was manually provided in the form
                manual_numero = request.form.get('numero', '').strip()
                
                if manual_numero:
                    # Check if this numero already exists for this project
                    existing_report = Relatorio.query.filter_by(
                        projeto_id=projeto_id,
                        numero=manual_numero
                    ).first()
                    
                    if existing_report:
                        flash(f'Número de relatório "{manual_numero}" já existe para esta obra. Por favor, use outro número.', 'error')
                        return redirect(url_for('create_report', projeto_id=projeto_id))
                    
                    # User manually edited the numero - use it
                    relatorio.numero = manual_numero
                    
                    # Extract numero_projeto from the numero string (e.g., "REL-0005" -> 5)
                    try:
                        if '-' in manual_numero:
                            numero_projeto_str = manual_numero.split('-')[1]
                            relatorio.numero_projeto = int(numero_projeto_str)
                        else:
                            # Fallback: calculate next numero_projeto
                            ultimo_numero = db.session.query(
                                db.func.max(Relatorio.numero_projeto)
                            ).filter_by(projeto_id=projeto_id).scalar()
                            relatorio.numero_projeto = (ultimo_numero or 0) + 1
                    except (ValueError, IndexError):
                        # If manual numero is invalid format, calculate next numero_projeto
                        ultimo_numero = db.session.query(
                            db.func.max(Relatorio.numero_projeto)
                        ).filter_by(projeto_id=projeto_id).scalar()
                        relatorio.numero_projeto = (ultimo_numero or 0) + 1
                    
                    current_app.logger.info(f"📝 Creating report with manual numero: {manual_numero} (numero_projeto: {relatorio.numero_projeto})")
                else:
                    # Auto-generate numero based on project sequence with numeracao_inicial
                    projeto = Projeto.query.get(projeto_id)
                    if not projeto:
                        flash('Projeto não encontrado.', 'error')
                        return redirect(url_for('create_report'))
                    
                    # Get numeracao_inicial from project (default to 1 if not set)
                    numeracao_inicial = getattr(projeto, 'numeracao_inicial', 1) or 1
                    
                    # Count existing reports for this project
                    relatorios_count = Relatorio.query.filter_by(projeto_id=projeto_id).count()
                    
                    # Calculate next number based on numeracao_inicial + count
                    proximo_numero = numeracao_inicial + relatorios_count
                    
                    # Double-check this numero doesn't exist (race condition protection)
                    tentativas = 0
                    while tentativas < 10:
                        numero_candidato = f"REL-{proximo_numero:04d}"
                        existing = Relatorio.query.filter_by(
                            projeto_id=projeto_id,
                            numero=numero_candidato
                        ).first()
                        
                        if not existing:
                            relatorio.numero_projeto = proximo_numero
                            relatorio.numero = numero_candidato
                            break
                        
                        proximo_numero += 1
                        tentativas += 1
                    
                    if tentativas >= 10:
                        flash('Erro ao gerar número do relatório. Tente novamente.', 'error')
                        return redirect(url_for('create_report', projeto_id=projeto_id))
                    
                    current_app.logger.info(f"📝 Creating report with auto-generated numero: {relatorio.numero} (based on numeracao_inicial: {numeracao_inicial})")
                
                relatorio.titulo = titulo
                relatorio.projeto_id = projeto_id
                relatorio.autor_id = current_user.id
            # Process checklist data from form
            checklist_text = ""
            checklist_items = []

            # Check for standard checklist items from form_complete.html
            checklist_fields = [
                ('estrutura', 'Estrutura / Fundação', 'obs_estrutura'),
                ('alvenaria', 'Alvenaria / Vedação', 'obs_alvenaria'), 
                ('instalacoes', 'Instalações (Elétrica/Hidráulica)', 'obs_instalacoes'),
                ('acabamento', 'Acabamentos', 'obs_acabamento'),
                ('limpeza', 'Limpeza / Organização', 'obs_limpeza')
            ]

            checklist_has_items = False
            for field_name, field_label, obs_field in checklist_fields:
                is_checked = request.form.get(field_name) == 'on'
                observation = request.form.get(obs_field, '').strip()

                if is_checked or observation:
                    checklist_has_items = True
                    status = "✓" if is_checked else "○"
                    checklist_items.append({
                        'status': status,
                        'item': field_label,
                        'observation': observation
                    })

            # Also check for JSON checklist data (legacy support)
            json_checklist = request.form.get('checklist_data')
            if json_checklist and not checklist_has_items:
                try:
                    import json
                    legacy_items = json.loads(json_checklist)
                    for item_data in legacy_items:
                        status = "✓" if item_data.get('completed') else "○"
                        checklist_items.append({
                            'status': status,
                            'item': item_data.get('item', ''),
                            'observation': item_data.get('observations', '')
                        })
                    checklist_has_items = True
                except Exception as e:
                    print(f"Error parsing JSON checklist data: {e}")

            # Format checklist text for PDF
            if checklist_has_items:
                checklist_text = "CHECKLIST DA OBRA:\n\n"
                for item in checklist_items:
                    checklist_text += f"{item['status']} {item['item']}\n"
                    if item['observation']:
                        checklist_text += f"   Observações: {item['observation']}\n"
                    checklist_text += "\n"

            # Process location data with address conversion
            latitude = request.form.get('latitude')
            longitude = request.form.get('longitude')
            location_text = ""
            if latitude and longitude:
                try:
                    # Use the same geocoding service to get formatted address
                    import requests
                    url = f"https://nominatim.openstreetmap.org/reverse?format=json&lat={latitude}&lon={longitude}&addressdetails=1&language=pt-BR"
                    headers = {'User-Agent': 'SistemaObras/1.0'}
                    response = requests.get(url, headers=headers, timeout=10)

                    if response.status_code == 200:
                        data = response.json()
                        addr = data.get('address', {})

                        # Build formatted address
                        address_parts = []
                        if addr.get('house_number'):
                            address_parts.append(f"{addr.get('road', '')} {addr['house_number']}")
                        elif addr.get('road'):
                            address_parts.append(addr['road'])
                        if addr.get('suburb') or addr.get('neighbourhood'):
                            address_parts.append(addr.get('suburb') or addr.get('neighbourhood'))
                        city = addr.get('city') or addr.get('town') or addr.get('village')
                        if city:
                            state = addr.get('state')
                            if state:
                                address_parts.append(f"{city} - {state}")
                            else:
                                address_parts.append(city)

                        formatted_address = ', '.join(filter(None, address_parts))
                        location_display = formatted_address or data.get('display_name', f"Lat: {latitude}, Lng: {longitude}")
                    else:
                        location_display = f"Lat: {latitude}, Lng: {longitude}"
                except:
                    location_display = f"Lat: {latitude}, Lng: {longitude}"

                location_text = f"\n\nLOCALIZAÇÃO DO RELATÓRIO:\n{location_display}\nCoordenadas GPS capturadas durante a visita."

            # Combine content with checklist and location
            final_content = ""
            if conteudo:
                final_content += conteudo
            if checklist_text:
                if final_content:
                    final_content += "\n\n" + checklist_text
                else:
                    final_content = checklist_text
            if location_text:
                final_content += location_text

            relatorio.conteudo = final_content
            relatorio.data_relatorio = data_relatorio
            relatorio.status = 'preenchimento'  # Status inicial - muda para "Aguardando Aprovação" ao finalizar
            relatorio.created_at = datetime.utcnow()
            relatorio.updated_at = datetime.utcnow()
            
            # Save lembrete_proxima_visita
            lembrete = request.form.get('lembrete_proxima_visita', '').strip()
            if lembrete:
                relatorio.lembrete_proxima_visita = lembrete
                current_app.logger.info(f"✅ Lembrete para próxima visita salvo: {lembrete[:50]}...")

            # Process acompanhantes (visit attendees)
            acompanhantes_data = request.form.get('acompanhantes')
            if acompanhantes_data:
                try:
                    import json
                    acompanhantes_list = json.loads(acompanhantes_data)
                    if isinstance(acompanhantes_list, list):
                        relatorio.acompanhantes = acompanhantes_list
                        current_app.logger.info(f"✅ Acompanhantes salvos: {len(acompanhantes_list)} registros")
                except Exception as e:
                    current_app.logger.error(f"❌ Erro ao processar acompanhantes: {e}")
                    relatorio.acompanhantes = None

            # Set approver automatically based on project
            # Priority: Temporary Approver for project > Global Approver
            aprovador = get_aprovador_padrao_para_projeto(projeto_id)
            if aprovador:
                relatorio.aprovador_id = aprovador.id
                relatorio.aprovador_nome = aprovador.nome_completo
                current_app.logger.info(f"✅ Aprovador automático definido: {aprovador.nome_completo} (ID={aprovador.id})")
            else:
                current_app.logger.warning(f"⚠️ Nenhum aprovador configurado para projeto {projeto_id}")

            db.session.add(relatorio)
            db.session.flush()  # Get the ID
            current_app.logger.info(f"✅ RELATÓRIO ID={relatorio.id} NÚMERO={relatorio.numero}")

            # VERIFICAR SE JÁ EXISTEM FOTOS (do autosave) - Evitar duplicação
            fotos_existentes = FotoRelatorio.query.filter_by(relatorio_id=relatorio.id).count()
            if fotos_existentes > 0:
                current_app.logger.info(f"📸 AUTOSAVE: {fotos_existentes} fotos já existem - PULANDO processamento de fotos para evitar duplicação")
                photo_count = fotos_existentes
                skip_photo_processing = True
            else:
                skip_photo_processing = False

            # Handle photo uploads if any - APENAS pasta uploads
            upload_folder = 'uploads'  # Sempre uploads
            if not os.path.exists(upload_folder):
                os.makedirs(upload_folder)

            current_app.logger.info(f"📁 SALVANDO FOTOS EM: {upload_folder}")
            photo_count = 0 if not skip_photo_processing else photo_count

            # Process photos from sessionStorage (via form data) - APENAS se não houver fotos do autosave
            if not skip_photo_processing:
                photos_data = request.form.get('photos_data')
                if photos_data:
                    try:
                        import json
                        photos_list = json.loads(photos_data)
                        for i, photo_data in enumerate(photos_list):
                            # Processo simplificado - apenas salvar referência
                            foto = FotoRelatorio()
                            foto.relatorio_id = relatorio.id
                            foto.filename = f"sessao_foto_{i+1}.jpg"
                            foto.legenda = photo_data.get('caption', f'Foto {i+1}')
                            foto.tipo_servico = photo_data.get('category', 'Geral')
                            foto.ordem = i + 1

                            db.session.add(foto)
                            photo_count += 1
                    except Exception as e:
                        pass  # Ignore session storage errors

            # Process mobile photos with mandatory caption validation - APENAS se não houver fotos do autosave
            if not skip_photo_processing:
                mobile_photos_data = request.form.get('mobile_photos_data')
                current_app.logger.info(f"🔍 MOBILE_PHOTOS_DATA PRESENTE? {mobile_photos_data is not None}")
            else:
                mobile_photos_data = None
            
            if mobile_photos_data and not skip_photo_processing:
                current_app.logger.info(f"📦 TAMANHO DO JSON: {len(mobile_photos_data)} caracteres")
                try:
                    import json
                    mobile_photos = json.loads(mobile_photos_data)
                    photos_list = mobile_photos.get('photos', [])

                    current_app.logger.info(f"📱 PROCESSANDO {len(photos_list)} FOTOS MOBILE")

                    # Validate and filter photos with captions (Item 19 - Mandatory captions)
                    valid_photos = []
                    for photo_data in photos_list:
                        caption = photo_data.get('caption', '').strip()
                        if not caption:
                            current_app.logger.warning(f"⚠️ Foto mobile sem legenda será ignorada: {photo_data.get('filename')}")
                            continue
                        valid_photos.append(photo_data)
                    
                    if len(valid_photos) < len(photos_list):
                        flash(f'⚠️ {len(photos_list) - len(valid_photos)} fotos sem legenda foram ignoradas.', 'warning')
                    
                    photos_list = valid_photos

                    # If validation passes, save mobile photos
                    for i, photo_data in enumerate(photos_list):
                        try:
                            # DEBUG: Log all keys in photo_data
                            current_app.logger.info(f"🔍 DEBUG Foto {i+1}: Keys disponíveis = {list(photo_data.keys())}")
                            
                            foto = FotoRelatorio()
                            foto.relatorio_id = relatorio.id
                            foto.filename = photo_data.get('filename', f'mobile_foto_{photo_count + i + 1}.jpg')
                            foto.legenda = photo_data.get('caption', '').strip()  # Already validated as non-empty
                            foto.descricao = photo_data.get('description', '').strip()  # Adicionar descrição
                            foto.tipo_servico = photo_data.get('category', 'Geral')
                            foto.ordem = photo_count + i + 1

                            # CRÍTICO: Salvar dados binários da imagem
                            has_data_field = photo_data.get('data') is not None
                            current_app.logger.info(f"🔍 DEBUG Foto {i+1}: Campo 'data' existe? {has_data_field}")
                            
                            if has_data_field:
                                try:
                                    import base64
                                    image_data_b64 = photo_data['data']
                                    data_preview = image_data_b64[:100] if isinstance(image_data_b64, str) else str(type(image_data_b64))
                                    current_app.logger.info(f"🔍 DEBUG Foto {i+1}: Preview dos dados = {data_preview}")
                                    
                                    if ',' in image_data_b64:
                                        image_data_b64 = image_data_b64.split(',')[1]

                                    # Decodificar e salvar dados binários
                                    image_binary = base64.b64decode(image_data_b64)
                                    foto.imagem = image_binary
                                    current_app.logger.info(f"✅ IMAGEM BINÁRIA SALVA: {len(image_binary)} bytes para foto {i+1}")
                                except Exception as e:
                                    current_app.logger.error(f"❌ Erro ao processar dados binários da foto mobile {i+1}: {e}")
                                    import traceback
                                    current_app.logger.error(f"❌ Traceback: {traceback.format_exc()}")
                                    # Continuar sem a imagem binária
                            else:
                                current_app.logger.warning(f"⚠️ Foto mobile {i+1} sem dados binários - 'data' field não encontrado")

                            # Salvar anotações se disponível (JSONB aceita dict diretamente)
                            if photo_data.get('annotations'):
                                foto.anotacoes_dados = photo_data['annotations']

                            # Salvar coordenadas se disponível (JSONB aceita dict diretamente)
                            if photo_data.get('coordinates'):
                                foto.coordenadas_anotacao = photo_data['coordinates']

                            db.session.add(foto)
                            current_app.logger.info(f"✅ Foto mobile {i+1} completa: legenda='{foto.legenda}', tipo='{foto.tipo_servico}', imagem={len(foto.imagem) if foto.imagem else 0} bytes")

                        except Exception as foto_error:
                            current_app.logger.error(f"❌ Erro ao processar foto mobile {i+1}: {foto_error}")
                            continue

                    photo_count += len(photos_list)
                except Exception as e:
                    current_app.logger.error(f"❌ Erro ao processar JSON de fotos mobile: {e}")
                    import traceback
                    current_app.logger.error(f"❌ Traceback completo: {traceback.format_exc()}")
                    flash('⚠️ Algumas fotos mobile podem não ter sido processadas corretamente.', 'warning')

            # Process regular file uploads - APENAS se não houver fotos do autosave
            if not skip_photo_processing:
                for i in range(50):  # Support up to 50 photos
                    photo_key = f'photo_{i}'
                    edited_photo_key = f'edited_photo_{i}'

                    # Check if this photo was edited
                    has_edited_version = edited_photo_key in request.form

                    if has_edited_version:
                        # Process only the edited version, ignore the original
                        try:
                            import base64
                            from io import BytesIO
                            from PIL import Image

                            edited_data = request.form[edited_photo_key]
                            # Remove data:image/jpeg;base64, prefix
                            if ',' in edited_data:
                                edited_data = edited_data.split(',')[1]

                            image_data = base64.b64decode(edited_data)
                            image = Image.open(BytesIO(image_data))

                            # Save edited image
                            filename = f"{uuid.uuid4().hex}_edited.jpg"
                            filepath = os.path.join(upload_folder, filename)
                            image.save(filepath, 'JPEG', quality=85)

                            # Get metadata
                            photo_caption = request.form.get(f'photo_caption_{i}', f'Foto {photo_count + 1}')
                            photo_category = request.form.get(f'photo_category_{i}', 'Geral')
                            photo_description = request.form.get(f'photo_description_{i}', '')

                            # Create photo record for edited version only
                            foto = FotoRelatorio()
                            foto.relatorio_id = relatorio.id
                            foto.filename = filename
                            foto.filename_anotada = filename  # Mark as annotated version
                            foto.legenda = photo_caption or f'Foto {photo_count + 1}'
                            foto.descricao = photo_description  # Adicionar descrição
                            foto.tipo_servico = photo_category or 'Geral'
                            foto.ordem = photo_count + 1
                            foto.imagem = image_data  # Salvar dados binários da imagem editada

                            # Salvar anotações se disponível (parse JSON string)
                            annotations = request.form.get(f'photo_annotations_{i}')
                            if annotations:
                                try:
                                    foto.anotacoes_dados = json.loads(annotations)
                                except (json.JSONDecodeError, TypeError):
                                    current_app.logger.warning(f"⚠️ Anotações inválidas para foto {i}, ignorando")
                                    foto.anotacoes_dados = None

                            db.session.add(foto)
                            photo_count += 1
                            current_app.logger.info(f"✅ Foto editada {photo_count} salva: {filename}, {len(image_data)} bytes")
                        except Exception as e:
                            current_app.logger.error(f"❌ Erro ao processar foto editada {i}: {e}")
                            continue

                    elif photo_key in request.files:
                        # Process original photo only if no edited version exists
                        file = request.files[photo_key]
                        if file and file.filename and file.filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
                            try:
                                filename = secure_filename(f"{uuid.uuid4().hex}_{file.filename}")
                                filepath = os.path.join(upload_folder, filename)

                                # Ler dados do arquivo antes de salvar
                                file_data = file.read()
                                file.seek(0)  # Reset para salvar o arquivo também
                                file.save(filepath)
                                current_app.logger.info(f"✅ FOTO SALVA: {filepath}")

                                # Get metadata
                                photo_caption = request.form.get(f'photo_caption_{i}', f'Foto {photo_count + 1}')
                                photo_category = request.form.get(f'photo_category_{i}', 'Geral')
                                photo_description = request.form.get(f'photo_description_{i}', '')

                                # Create photo record for original
                                foto = FotoRelatorio()
                                foto.relatorio_id = relatorio.id
                                foto.filename = filename
                                foto.filename_original = filename  # Mark as original version
                                foto.legenda = photo_caption or f'Foto {photo_count + 1}'
                                foto.descricao = photo_description  # Adicionar descrição
                                foto.tipo_servico = photo_category or 'Geral'
                                foto.ordem = photo_count + 1
                                foto.imagem = file_data  # Salvar dados binários da imagem original

                                # Salvar anotações se disponível (parse JSON string)
                                annotations = request.form.get(f'photo_annotations_{i}')
                                if annotations:
                                    try:
                                        foto.anotacoes_dados = json.loads(annotations)
                                    except (json.JSONDecodeError, TypeError):
                                        current_app.logger.warning(f"⚠️ Anotações inválidas para foto {i}, ignorando")
                                        foto.anotacoes_dados = None

                                db.session.add(foto)
                                photo_count += 1
                                current_app.logger.info(f"✅ Foto original {photo_count} salva: {filename}, {len(file_data)} bytes")
                            except Exception as e:
                                current_app.logger.error(f"❌ Erro ao processar foto {i}: {e}")
                                continue

            current_app.logger.info(f"📊 RESUMO FINAL: {photo_count} fotos processadas para relatório {relatorio.numero}")

            # Debug: Verificar fotos antes do commit
            fotos_debug = FotoRelatorio.query.filter_by(relatorio_id=relatorio.id).all()
            for foto_debug in fotos_debug:
                current_app.logger.info(f"🔍 FOTO PRÉ-COMMIT: ID={foto_debug.id}, filename='{foto_debug.filename}', legenda='{foto_debug.legenda}', descricao='{foto_debug.descricao}', tipo='{foto_debug.tipo_servico}', imagem_size={len(foto_debug.imagem) if foto_debug.imagem else 0}")

            current_app.logger.info(f"🔧 Fazendo COMMIT de {photo_count} fotos para relatório {relatorio.id}")
            
            db.session.commit()
            
            current_app.logger.info(f"✅ COMMIT REALIZADO COM SUCESSO")
            
            # VERIFICAÇÃO PÓS-COMMIT: Contar imagens salvas com dados binários
            fotos_com_imagem = db.session.query(FotoRelatorio).filter(
                FotoRelatorio.relatorio_id == relatorio.id,
                FotoRelatorio.imagem != None
            ).count()
            current_app.logger.info(f"📊 VERIFICAÇÃO: {fotos_com_imagem} de {photo_count} fotos têm dados binários salvos")

            # Debug: Verificar fotos após o commit diretamente do banco
            fotos_post = FotoRelatorio.query.filter_by(relatorio_id=relatorio.id).all()
            current_app.logger.info(f"✅ PÓS-COMMIT: {len(fotos_post)} fotos encontradas no banco para relatório {relatorio.id}")
            
            for foto_post in fotos_post:
                imagem_size = len(foto_post.imagem) if foto_post.imagem else 0
                current_app.logger.info(f"💾 FOTO ID={foto_post.id}: legenda='{foto_post.legenda}', filename='{foto_post.filename}', imagem_bytes={imagem_size}, imagem_presente={foto_post.imagem is not None}")
                
                # Verificar dados JSON
                if foto_post.anotacoes_dados:
                    current_app.logger.info(f"   📝 Anotações: {type(foto_post.anotacoes_dados).__name__}")
                if foto_post.coordenadas_anotacao:
                    current_app.logger.info(f"   📍 Coordenadas: {type(foto_post.coordenadas_anotacao).__name__}")

            # SEMPRE mantém status como "preenchimento" (comportamento igual ao autosave)
            should_finalize = request.form.get('should_finalize') == 'true'
            if should_finalize:
                current_app.logger.info(f"✅ Relatório {relatorio.numero} salvo em preenchimento (should_finalize ignorado)")

            flash('Relatório criado com sucesso!', 'success')

            # Return JSON response for AJAX submission
            if request.content_type and 'multipart/form-data' in request.content_type:
                return jsonify({
                    'success': True, 
                    'redirect': '/reports',
                    'report_id': relatorio.id,
                    'report_number': relatorio.numero
                })
            else:
                return redirect(url_for('reports'))

        except Exception as e:
            db.session.rollback()
            print(f"Erro detalhado ao criar relatório: {e}")
            import traceback
            traceback.print_exc()
            flash(f'Erro ao criar relatório: {str(e)}', 'error')

            # Return JSON error response for AJAX submission  
            if request.content_type and 'multipart/form-data' in request.content_type:
                return jsonify({'success': False, 'error': str(e)}), 400
            else:
                return redirect(url_for('create_report'))

    projetos = Projeto.query.filter_by(status='Ativo').all()
    # Get admin users for approver selection
    admin_users = User.query.filter_by(is_master=True).all()

    # Check if we're editing an existing report
    edit_report_id = request.args.get('edit')
    existing_report = None
    existing_fotos = []
    existing_checklist = {}
    
    if edit_report_id:
        try:
            edit_report_id = int(edit_report_id)
            existing_report = Relatorio.query.get(edit_report_id)
            if existing_report:
                # Check permissions - only author or master can edit
                if existing_report.autor_id != current_user.id and not current_user.is_master:
                    flash('Você não tem permissão para editar este relatório.', 'error')
                    return redirect(url_for('reports'))
                
                # Load existing photos
                try:
                    existing_fotos = FotoRelatorio.query.filter_by(relatorio_id=existing_report.id).order_by(FotoRelatorio.ordem).all()
                    current_app.logger.info(f"📸 Loaded {len(existing_fotos)} photos for report {edit_report_id}")
                except Exception as e:
                    current_app.logger.error(f"❌ Error loading photos for report {edit_report_id}: {str(e)}")
                    existing_fotos = []
                
                # Load existing checklist
                try:
                    if existing_report.checklist_data:
                        import json
                        existing_checklist = json.loads(existing_report.checklist_data)
                        current_app.logger.info(f"✅ Loaded checklist for report {edit_report_id}")
                    else:
                        existing_checklist = {}
                except Exception as e:
                    current_app.logger.error(f"❌ Error loading checklist for report {edit_report_id}: {str(e)}")
                    existing_checklist = {}
                    
                current_app.logger.info(f"📝 Loading existing report {existing_report.numero} for editing")
            else:
                current_app.logger.warning(f"⚠️ Report {edit_report_id} not found")
                flash('Relatório não encontrado.', 'error')
                return redirect(url_for('reports'))
        except (ValueError, TypeError) as e:
            current_app.logger.error(f"❌ Invalid report ID format: {edit_report_id} - {str(e)}")
            flash('ID de relatório inválido.', 'error')
            return redirect(url_for('reports'))
        except Exception as e:
            current_app.logger.error(f"❌ Unexpected error loading report {edit_report_id}: {str(e)}")
            import traceback
            traceback.print_exc()
            flash('Erro ao carregar relatório para edição.', 'error')
            return redirect(url_for('reports'))

    # Auto-preenchimento: Verificar se projeto_id foi passado como parâmetro da URL
    selected_project = None
    selected_aprovador = None
    selected_aprovador_nome = ''
    next_numero = None
    lembrete_anterior = None
    
    # If editing, use the existing report's project
    if existing_report:
        selected_project = existing_report.projeto
        if selected_project:
            aprovador_obj = get_aprovador_padrao_para_projeto(selected_project.id)
            selected_aprovador_nome = aprovador_obj.nome_completo if aprovador_obj else ''
            selected_aprovador = selected_aprovador_nome
    else:
        projeto_id_param = request.args.get('projeto_id')
        if projeto_id_param:
            try:
                projeto_id_param = int(projeto_id_param)
                selected_project = Projeto.query.get(projeto_id_param)
                # Buscar aprovador padrão para este projeto
                if selected_project:
                    aprovador_obj = get_aprovador_padrao_para_projeto(selected_project.id)
                    selected_aprovador_nome = aprovador_obj.nome_completo if aprovador_obj else ''
                    selected_aprovador = selected_aprovador_nome
                    
                    # Calculate next report number: max of (numeracao_inicial-1, highest existing numero_projeto) + 1
                    numeracao_inicial = selected_project.numeracao_inicial or 1
                    max_numero_existente = db.session.query(
                        db.func.max(Relatorio.numero_projeto)
                    ).filter_by(projeto_id=projeto_id_param).scalar()
                    
                    if max_numero_existente is None:
                        # No reports yet, use numeracao_inicial
                        proximo_numero_projeto = numeracao_inicial
                    else:
                        # Ensure we never go below numeracao_inicial and always increment from max
                        proximo_numero_projeto = max(numeracao_inicial - 1, max_numero_existente) + 1
                    
                    next_numero = f"REL-{proximo_numero_projeto:04d}"
                    current_app.logger.info(f"📋 Next numero for project {projeto_id_param}: {next_numero} (numeracao_inicial: {numeracao_inicial}, max_existente: {max_numero_existente})")
                    
                    # Buscar lembrete do último relatório deste projeto
                    ultimo_relatorio = Relatorio.query.filter_by(
                        projeto_id=projeto_id_param
                    ).order_by(Relatorio.created_at.desc()).first()
                    
                    if ultimo_relatorio and ultimo_relatorio.lembrete_proxima_visita:
                        lembrete_anterior = {
                            'numero': ultimo_relatorio.numero,
                            'texto': ultimo_relatorio.lembrete_proxima_visita
                        }
                        current_app.logger.info(f"📝 Lembrete encontrado do relatório {ultimo_relatorio.numero}")
            except (ValueError, TypeError):
                selected_project = None
        else:
            # Se não há projeto específico, buscar aprovador global
            aprovador_obj = get_aprovador_padrao_para_projeto(None)
            selected_aprovador_nome = aprovador_obj.nome_completo if aprovador_obj else ''
            selected_aprovador = selected_aprovador_nome

    # ===== SERIALIZAÇÃO DE OBJETOS PARA JSON =====
    # Converter todos os objetos ORM em dicionários simples
    
    # Serializar projetos
    projetos_data = []
    for projeto in projetos:
        projetos_data.append({
            'id': projeto.id,
            'nome': projeto.nome,
            'cliente': projeto.construtora or '',
            'status': projeto.status or 'Ativo',
            'numeracao_inicial': projeto.numeracao_inicial or 1
        })
    
    # Serializar usuários admin
    admin_users_data = []
    for user in admin_users:
        admin_users_data.append({
            'id': user.id,
            'nome': user.nome_completo,
            'email': user.email,
            'is_master': user.is_master
        })
    
    # Serializar projeto selecionado
    selected_project_data = None
    if selected_project:
        selected_project_data = {
            'id': selected_project.id,
            'nome': selected_project.nome,
            'cliente': selected_project.construtora or '',
            'status': selected_project.status or 'Ativo',
            'numeracao_inicial': selected_project.numeracao_inicial or 1
        }
    
    # Serializar relatório existente
    existing_report_data = None
    if existing_report:
        existing_report_data = {
            'id': existing_report.id,
            'numero': existing_report.numero,
            'titulo': existing_report.titulo or '',
            'projeto_id': existing_report.projeto_id,
            'autor_id': existing_report.autor_id,
            'status': existing_report.status or 'preenchimento'
        }
    
    # Serializar fotos existentes
    existing_fotos_data = []
    for foto in existing_fotos:
        existing_fotos_data.append({
            'id': foto.id,
            'url': url_for('api_get_photo', foto_id=foto.id),
            'legenda': foto.legenda or '',
            'categoria': foto.categoria or '',
            'local': foto.local or '',
            'ordem': foto.ordem or 0
        })
    
    # Preparar report_data para autosave
    if existing_report:
        # Serialize acompanhantes to ensure it's JSON-safe (no User objects)
        acompanhantes_safe = []
        if existing_report.acompanhantes:
            try:
                # If it's already a list of simple types, use it
                if isinstance(existing_report.acompanhantes, list):
                    for item in existing_report.acompanhantes:
                        # Convert User objects to strings (names)
                        if hasattr(item, 'nome_completo'):
                            acompanhantes_safe.append(item.nome_completo)
                        elif isinstance(item, (str, dict)):
                            acompanhantes_safe.append(item)
                        else:
                            acompanhantes_safe.append(str(item))
                else:
                    acompanhantes_safe = []
            except Exception as e:
                current_app.logger.warning(f"Error serializing acompanhantes: {e}")
                acompanhantes_safe = []
        
        # Modo de edição: usar dados existentes
        report_data = {
            'id': existing_report.id,
            'projeto_id': existing_report.projeto_id,
            'titulo': existing_report.titulo or '',
            'conteudo': existing_report.conteudo or '',
            'data_relatorio': existing_report.data_relatorio.isoformat() if existing_report.data_relatorio else date.today().isoformat(),
            'aprovador_nome': existing_report.aprovador_nome or '',
            'observacoes_finais': existing_report.observacoes_finais or '',
            'lembrete_proxima_visita': existing_report.lembrete_proxima_visita.isoformat() if existing_report.lembrete_proxima_visita else '',
            'latitude': existing_report.latitude,
            'longitude': existing_report.longitude,
            'checklist_data': existing_checklist if existing_checklist else {},
            'acompanhantes': acompanhantes_safe,
            'fotos': existing_fotos_data
        }
    else:
        # Modo de criação: estrutura vazia
        report_data = {
            'id': None,
            'projeto_id': selected_project_data['id'] if selected_project_data else None,
            'titulo': '',
            'conteudo': '',
            'data_relatorio': date.today().isoformat(),
            'aprovador_nome': selected_aprovador if selected_aprovador else '',
            'observacoes_finais': '',
            'lembrete_proxima_visita': '',
            'latitude': None,
            'longitude': None,
            'checklist_data': {},
            'acompanhantes': [],
            'fotos': []
        }
    
    # Log de sucesso
    current_app.logger.info(f"✅ Dados prontos para template: {len(existing_fotos_data)} fotos, {len(existing_checklist)} checklist items, {len(report_data.get('acompanhantes', []))} acompanhantes")
    
    # Determinar se está em modo de edição
    edit_mode = existing_report is not None
    
    # IMPORTANTE: Passar report_id para o template para inicializar o autosave corretamente
    report_id = existing_report.id if existing_report else None
    
    # Render the form for GET requests - TODOS OS OBJETOS AGORA SÃO DICIONÁRIOS
    return render_template('reports/form_complete.html', 
                         projetos=projetos_data, 
                         admin_users=admin_users_data, 
                         selected_project=selected_project_data,
                         selected_aprovador=selected_aprovador,
                         disable_fields=disable_fields,
                         preselected_project_id=preselected_project_id,
                         existing_report=existing_report_data,
                         existing_fotos=existing_fotos_data,
                         existing_checklist=existing_checklist,
                         next_numero=next_numero,
                         lembrete_anterior=lembrete_anterior,
                         report_data=report_data,
                         edit_mode=edit_mode,
                         report_id=report_id,
                         today=date.today().isoformat())

# Removed duplicate function - using the more comprehensive version below at line 7415

# Função deprecated removida para evitar conflitos

# Photo annotation system routes
@app.route('/photo-annotation')
@login_required
def photo_annotation():
    photo_path = request.args.get('photo')
    report_id = request.args.get('report_id')
    photo_id = request.args.get('photo_id')

    if not photo_path:
        flash('Foto não especificada.', 'error')
        return redirect(url_for('reports'))

    return render_template('reports/photo_annotation.html')

@app.route('/photo-editor', methods=['GET', 'POST'])
@login_required
def photo_editor():
    """Editor de fotos professional com Fabric.js"""
    photo_id = request.args.get('photoId') or request.form.get('photoId')
    image_url = request.args.get('imageUrl', '')

    if request.method == 'POST':
        # Processar imagem editada
        edited_image = request.form.get('edited_image')
        legend = request.form.get('legend', '')

        if edited_image and photo_id:
            try:
                # Processar base64
                if ',' in edited_image:
                    edited_image = edited_image.split(',')[1]

                import base64
                image_binary = base64.b64decode(edited_image)

                # Salvar imagem editada
                upload_folder = app.config.get('UPLOAD_FOLDER', 'uploads')
                if not os.path.exists(upload_folder):
                    os.makedirs(upload_folder)

                filename = f"edited_{photo_id}_{uuid.uuid4().hex}.jpg"
                filepath = os.path.join(upload_folder, filename)

                with open(filepath, 'wb') as f:
                    f.write(image_binary)

                # Atualizar banco se necessário
                if photo_id != 'temp':
                    foto = FotoRelatorio.query.get(photo_id)
                    if foto:
                        foto.filename_anotada = filename
                        foto.legenda = legend
                        foto.imagem = image_binary  # Salvar dados binários da imagem editada
                        db.session.commit()

                flash('Imagem editada com sucesso!', 'success')
                return redirect(request.referrer or url_for('reports'))

            except Exception as e:
                flash(f'Erro ao salvar imagem: {str(e)}', 'error')

    return render_template('reports/fabric_photo_editor.html', 
                         photo_id=photo_id, 
                         image_url=image_url)

@app.route('/reports/<int:photo_id>/annotate', methods=['POST'])
@login_required
def annotate_photo(photo_id):
    """Salvar anotações em uma foto"""
    try:
        foto = FotoRelatorio.query.get_or_404(photo_id)
        data = request.get_json()

        if 'image_data' not in data:
            return jsonify({'success': False, 'error': 'Dados da imagem não encontrados'})

        # Process base64 image data
        image_data = data['image_data']
        if ',' in image_data:
            image_data = image_data.split(',')[1]

        # Save annotated image
        import base64
        image_binary = base64.b64decode(image_data)

        upload_folder = app.config.get('UPLOAD_FOLDER', 'uploads')
        if not os.path.exists(upload_folder):
            os.makedirs(upload_folder)

        # Generate new filename for annotated version
        filename_parts = foto.filename.rsplit('.', 1)
        annotated_filename = f"{filename_parts[0]}_annotated.{filename_parts[1] if len(filename_parts) > 1 else 'jpg'}"

        filepath = os.path.join(upload_folder, annotated_filename)
        with open(filepath, 'wb') as f:
            f.write(image_binary)

        # Update photo record with binary data
        foto.filename = annotated_filename
        # Salvar annotations como JSON (não como string vazia)
        annotations_data = data.get('annotations')
        if annotations_data:
            foto.coordenadas_anotacao = annotations_data if isinstance(annotations_data, (dict, list)) else None
        foto.imagem = image_binary  # Salvar dados binários da imagem anotada
        db.session.commit()

        return jsonify({'success': True})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/reports/<int:photo_id>/delete', methods=['POST'])
@login_required
def delete_photo(photo_id):
    """Excluir uma foto (legacy endpoint - use /api/fotos/<foto_id>/delete)"""
    try:
        foto = FotoRelatorio.query.get_or_404(photo_id)

        # Check permissions
        if not current_user.is_master and foto.relatorio.autor_id != current_user.id:
            return jsonify({'success': False, 'error': 'Permissão negada'})

        # Delete file
        upload_folder = app.config.get('UPLOAD_FOLDER', 'uploads')
        filepath = os.path.join(upload_folder, foto.filename)
        if os.path.exists(filepath):
            os.remove(filepath)

        # Delete record
        db.session.delete(foto)
        db.session.commit()

        return jsonify({'success': True})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/fotos/<int:foto_id>/delete', methods=['POST'])
@login_required
@csrf.exempt
def api_delete_foto(foto_id):
    """API dedicada para exclusão de fotos - evita conflito com rota de exclusão de relatórios"""
    try:
        current_app.logger.info(f"🗑️ API DELETE FOTO: Usuário {current_user.username} deletando foto {foto_id}")
        
        foto = FotoRelatorio.query.get_or_404(foto_id)

        # Check permissions
        if not current_user.is_master and foto.relatorio.autor_id != current_user.id:
            current_app.logger.warning(f"❌ Permissão negada para deletar foto {foto_id}")
            return jsonify({'success': False, 'error': 'Permissão negada'}), 403

        # Store info for logging
        relatorio_id = foto.relatorio_id
        filename = foto.filename

        # Delete binary data (if stored in database)
        # The file deletion is optional since we're using database storage
        try:
            upload_folder = app.config.get('UPLOAD_FOLDER', 'uploads')
            if filename:
                filepath = os.path.join(upload_folder, filename)
                if os.path.exists(filepath):
                    os.remove(filepath)
                    current_app.logger.info(f"📁 Arquivo deletado: {filepath}")
        except Exception as file_error:
            current_app.logger.warning(f"⚠️ Erro ao deletar arquivo físico: {file_error}")

        # Delete record from database
        db.session.delete(foto)
        db.session.commit()
        
        current_app.logger.info(f"✅ Foto {foto_id} deletada com sucesso do relatório {relatorio_id}")

        return jsonify({
            'success': True,
            'message': 'Foto deletada com sucesso',
            'foto_id': foto_id
        }), 200

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"❌ Erro ao deletar foto {foto_id}: {str(e)}")
        import traceback
        current_app.logger.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({'success': False, 'error': f'Erro ao deletar foto: {str(e)}'}), 500

@app.route('/reports/<int:id>/status', methods=['POST'])
@login_required
def update_report_status(id):
    """Atualizar status do relatório"""
    try:
        relatorio = Relatorio.query.get_or_404(id)
        data = request.get_json()

        # Check permissions
        if not current_user.is_master and relatorio.autor_id != current_user.id:
            return jsonify({'success': False, 'error': 'Permissão negada'})

        new_status = data.get('status')
        valid_statuses = ['Rascunho', 'Aguardando Aprovação', 'Aprovado', 'Rejeitado']

        if new_status not in valid_statuses:
            return jsonify({'success': False, 'error': 'Status inválido'})

        relatorio.status = new_status
        db.session.commit()

        return jsonify({'success': True})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/reports/<int:report_id>/review')
@login_required
def review_report(report_id):
    """Página de revisão do relatório - versão robusta conforme especificação"""
    try:
        # Logging inicial conforme especificação
        current_app.logger.info(f"🔍 /reports/{report_id}/review: Usuário {current_user.id} acessando revisão")

        # Buscar relatório com validação robusta
        try:
            report = Relatorio.query.get_or_404(report_id)
            current_app.logger.info(f"✅ Relatório {report_id} encontrado para revisão: Status={report.status}")
        except Exception as e:
            current_app.logger.exception(f"ERRO CRÍTICO ao buscar relatório {report_id} para revisão: {str(e)}")
            abort(500, description="Erro interno ao carregar relatório.")

        # Validação defensiva de atributos None
        if not report:
            current_app.logger.error(f"❌ Relatório {report_id} é None após get_or_404")
            abort(404, description="Relatório não encontrado.")

        # Proteger contra JSON malformado no checklist com validação defensiva
        try:
            import json
            checklist = json.loads(report.checklist_data) if report.checklist_data else {}
            current_app.logger.info(f"✅ Checklist review carregado: {len(checklist)} itens")
        except (json.JSONDecodeError, TypeError, AttributeError) as e:
            current_app.logger.exception(f"ERRO JSON REVIEW relatório {report_id}: {str(e)}")
            checklist = {}
        except Exception as e:
            current_app.logger.exception(f"ERRO GERAL CHECKLIST review {report_id}: {str(e)}")
            checklist = {}

        # Buscar fotos com validação defensiva
        try:
            fotos = FotoRelatorio.query.filter_by(relatorio_id=report_id).order_by(FotoRelatorio.ordem).all()
            current_app.logger.info(f"✅ Fotos review carregadas: {len(fotos)} arquivos")
        except Exception as e:
            current_app.logger.error(f"❌ Erro ao buscar fotos review relatório {report_id}: {str(e)}")
            fotos = []

        # Verificar se usuário é aprovador com validação defensiva
        try:
            # Proteger contra projeto_id None
            projeto_id_safe = getattr(report, 'projeto_id', None) if report else None
            user_is_approver = current_user_is_aprovador(projeto_id_safe)
            current_app.logger.info(f"🔐 Usuário {current_user.id} é aprovador: {user_is_approver} (projeto_id={projeto_id_safe})")
        except AttributeError as e:
            current_app.logger.error(f"❌ ATRIBUTO NONE: Erro ao verificar aprovador para relatório {report_id}: {str(e)}")
            user_is_approver = False
        except Exception as e:
            current_app.logger.error(f"❌ ERRO GERAL aprovador para relatório {report_id}: {str(e)}")
            user_is_approver = False

        return render_template('reports/review.html', 
                             report=report,  # Padronizado conforme especificação
                             relatorio=report,  # Manter compatibilidade
                             fotos=fotos, 
                             checklist=checklist,
                             user_is_approver=user_is_approver)

    except Exception as e:
        current_app.logger.exception(f"ERRO GERAL REVIEW /reports/{report_id}/review: {str(e)}")
        abort(500, description="Erro interno ao carregar página de revisão.")

@app.route('/reports/<int:id>/approve', methods=['POST'])
@login_required
def approve_report(id):
    """Aprova relatório, gera PDF e envia e-mail via Resend."""
    relatorio = db.session.get(Relatorio, id)
    if not relatorio:
        flash('Relatório não encontrado.', 'error')
        return redirect(url_for('reports'))

    try:
        # Atualizar status do relatório ANTES de enviar e-mail
        relatorio.status = "Aprovado"
        relatorio.aprovado_por = current_user.id
        relatorio.data_aprovacao = datetime.utcnow()
        db.session.commit()
        
        current_app.logger.info(f"✅ Relatório {relatorio.numero} aprovado no banco de dados")
        
        # Criar notificação para o autor do relatório
        try:
            from notification_service import notification_service
            notification_service.criar_notificacao_relatorio_aprovado(relatorio.id, current_user.id)
            current_app.logger.info(f"✅ Notificação de aprovação enviada ao autor")
        except Exception as notif_error:
            current_app.logger.error(f"⚠️ Erro ao criar notificação de aprovação: {notif_error}")

        # Gerar PDF usando WeasyPrint
        from pdf_generator_weasy import WeasyPrintReportGenerator
        generator = WeasyPrintReportGenerator()
        
        # Buscar fotos do relatório
        fotos = FotoRelatorio.query.filter_by(relatorio_id=relatorio.id).order_by(FotoRelatorio.ordem).all()
        
        # Gerar PDF
        pdf_filename = f"relatorio_{relatorio.numero}.pdf"
        pdf_path = os.path.join('static', 'reports', pdf_filename)
        os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
        
        generator.generate_report_pdf(relatorio, fotos, output_path=pdf_path)
        current_app.logger.info(f"📄 PDF gerado: {pdf_path}")

        # Coletar destinatários de e-mail
        destinatarios = []
        
        # Adicionar autor do relatório
        if relatorio.autor and relatorio.autor.email:
            destinatarios.append(relatorio.autor.email)
        
        # Adicionar responsável do projeto
        if relatorio.projeto and relatorio.projeto.responsavel:
            if relatorio.projeto.responsavel.email and relatorio.projeto.responsavel.email not in destinatarios:
                destinatarios.append(relatorio.projeto.responsavel.email)
        
        # Adicionar acompanhantes (JSONB array)
        if relatorio.acompanhantes:
            try:
                acompanhantes_list = relatorio.acompanhantes if isinstance(relatorio.acompanhantes, list) else []
                for acomp in acompanhantes_list:
                    if isinstance(acomp, dict) and acomp.get('email'):
                        if acomp['email'] not in destinatarios:
                            destinatarios.append(acomp['email'])
            except Exception as e:
                current_app.logger.warning(f"⚠️ Erro ao processar acompanhantes: {e}")
        
        # Adicionar clientes da obra que devem receber relatórios
        try:
            clientes = EmailCliente.query.filter_by(
                projeto_id=relatorio.projeto_id,
                ativo=True,
                receber_relatorios=True
            ).all()
            
            for cliente in clientes:
                if cliente.email and cliente.email not in destinatarios:
                    destinatarios.append(cliente.email)
        except Exception as e:
            current_app.logger.warning(f"⚠️ Erro ao buscar clientes: {e}")

        if not destinatarios:
            current_app.logger.warning(f"⚠️ Nenhum destinatário encontrado para relatório {relatorio.numero}")
            flash('✅ Relatório aprovado com sucesso! Nenhum destinatário de e-mail configurado.', 'warning')
            return redirect(url_for('review_report', report_id=id))

        current_app.logger.info(f"📧 Destinatários: {destinatarios}")

        # Preparar assunto e corpo do e-mail
        assunto = f"Relatório {relatorio.numero} - {relatorio.projeto.nome}"
        responsavel_email = relatorio.projeto.responsavel.email if relatorio.projeto and relatorio.projeto.responsavel else "elp@elpconsultoria.eng.br"
        
        corpo_html = f"""
        <html>
        <body>
            <p>Olá,</p>
            <p>O relatório <strong>{relatorio.numero}</strong> referente à obra <strong>{relatorio.projeto.nome}</strong> foi aprovado.</p>
            <p>O PDF está anexado a este e-mail.</p>
            <p>Qualquer dúvida, entre em contato com <strong>{responsavel_email}</strong>.</p>
            <br>
            <p>Atenciosamente,<br><strong>ELP Consultoria</strong></p>
        </body>
        </html>
        """

        # Enviar e-mail via Resend
        from email_service import EmailServiceRelatorio
        email_service = EmailServiceRelatorio()
        enviado = email_service.enviar_relatorio_por_email(
            relatorio.id, destinatarios, assunto, corpo_html, pdf_path
        )

        if enviado:
            current_app.logger.info(f"✅ E-mail enviado com sucesso para {len(destinatarios)} destinatário(s)")
            flash(f'✅ Relatório aprovado com sucesso! E-mail enviado para {len(destinatarios)} destinatário(s).', 'success')
        else:
            current_app.logger.warning(f"⚠️ Falha ao enviar e-mail")
            flash('✅ Relatório aprovado com sucesso! Não foi possível enviar o e-mail de notificação.', 'warning')
        
        return redirect(url_for('review_report', report_id=id))
            
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"❌ Erro ao aprovar relatório: {str(e)}")
        import traceback
        current_app.logger.error(traceback.format_exc())
        flash(f'Erro ao aprovar relatório: {str(e)}', 'error')
        return redirect(url_for('review_report', report_id=id))

@app.route('/reports/<int:id>/reject', methods=['POST'])
@login_required
def reject_report(id):
    """Rejeitar relatório - apenas usuários aprovadores"""
    relatorio = Relatorio.query.get_or_404(id)

    # Verificar se usuário é aprovador para este projeto
    if not current_user_is_aprovador(relatorio.projeto_id):
        flash('Acesso negado. Apenas usuários aprovadores podem rejeitar relatórios.', 'error')
        return redirect(url_for('reports'))

    # Obter comentário da reprovação (obrigatório)
    comentario = request.form.get('comentario_reprovacao')
    if not comentario or not comentario.strip():
        flash('Comentário de reprovação é obrigatório.', 'error')
        return redirect(url_for('review_report', report_id=id))

    # Guardar informações do autor antes da mudança
    autor = relatorio.autor
    aprovador = current_user
    projeto = relatorio.projeto
    projeto_nome = projeto.nome if projeto else 'N/A'

    try:
        # Atualizar status do relatório
        relatorio.status = 'Rejeitado'  # Status correto para relatórios rejeitados
        relatorio.aprovado_por = current_user.id
        relatorio.data_aprovacao = datetime.utcnow()
        relatorio.comentario_aprovacao = comentario.strip()

        # Commit único de todas as alterações
        db.session.commit()
        
        # Criar notificação para o autor usando o novo serviço
        from notification_service import notification_service
        try:
            notification_service.criar_notificacao_relatorio_reprovado(relatorio.id)
        except Exception as notif_error:
            current_app.logger.error(f"⚠️ Erro ao criar notificação de reprovação: {notif_error}")
        
        current_app.logger.info(f"✅ Relatório {relatorio.numero} rejeitado com sucesso - Notificação criada")
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"❌ Erro ao rejeitar relatório: {str(e)}")
        flash(f'Erro ao rejeitar relatório: {str(e)}', 'error')
        return redirect(url_for('review_report', report_id=id))

    # Log da ação de reprovação
    import logging
    logging.info(f"Relatório {relatorio.numero} rejeitado por {current_user.nome_completo} (ID: {current_user.id}). "
                f"Projeto: {projeto_nome}. Motivo: {comentario.strip()[:100]}...")

    # TODO: Implementar envio de notificação por email ao autor com Resend

    flash(f'Relatório {relatorio.numero} rejeitado e devolvido para edição. '
          f'O autor {autor.nome_completo} deve fazer as correções solicitadas.', 'warning')
    return redirect(url_for('reports'))

@app.route('/reports/pending')
@login_required
def pending_reports():
    """Painel de relatórios pendentes de aprovação - apenas usuários aprovadores"""
    # Verificar se usuário é aprovador (global ou de algum projeto)
    if not current_user_is_aprovador():
        flash('Acesso negado. Apenas usuários aprovadores podem ver relatórios pendentes.', 'error')
        return redirect(url_for('reports'))

    page = request.args.get('page', 1, type=int)
    relatorios = Relatorio.query.filter_by(status='Aguardando Aprovação').order_by(Relatorio.created_at.desc()).paginate(
        page=page, per_page=10, error_out=False)

    return render_template('reports/pending.html', relatorios=relatorios)

# Função autosave_report duplicada removida - mantendo apenas a implementação principal robusta (linha ~625)

@app.route('/reports/<int:report_id>/finalize', methods=['POST'])
@login_required
def finalize_report(report_id):
    """Finalizar relatório em preenchimento e eliminar duplicados"""
    try:
        relatorio = Relatorio.query.get_or_404(report_id)

        # Verificar permissões
        if not current_user.is_master and relatorio.autor_id != current_user.id:
            return jsonify({'success': False, 'error': 'Acesso negado'}), 403

        # Tornar idempotente: se já está aguardando aprovação, limpar duplicados e retornar sucesso
        if relatorio.status == 'Aguardando Aprovação':
            # Limpar TODOS os outros relatórios em "preenchimento" do mesmo projeto
            duplicados = Relatorio.query.filter(
                Relatorio.id != relatorio.id,
                Relatorio.projeto_id == relatorio.projeto_id,
                Relatorio.status == 'preenchimento'
            ).all()
            
            for dup in duplicados:
                # Deletar fotos associadas
                fotos_dup = FotoRelatorio.query.filter_by(relatorio_id=dup.id).all()
                for foto in fotos_dup:
                    db.session.delete(foto)
                db.session.delete(dup)
                current_app.logger.info(f"🗑️ Deletado relatório duplicado ID={dup.id} (estava em preenchimento)")
            
            if duplicados:
                db.session.commit()
                current_app.logger.info(f"✅ {len(duplicados)} relatório(s) duplicado(s) removido(s)")
            
            return jsonify({
                'success': True,
                'message': 'Relatório já foi finalizado e enviado para aprovação',
                'redirect': url_for('reports')
            })

        # Verificar se o relatório está em preenchimento
        if relatorio.status != 'preenchimento':
            return jsonify({'success': False, 'error': 'Relatório não está em preenchimento'}), 400

        # SEMPRE mantém como "em preenchimento" ao concluir (igual ao autosave)
        relatorio.status = 'preenchimento'
        relatorio.updated_at = datetime.utcnow()
        current_app.logger.info(f"✅ Relatório {relatorio.numero} salvo em preenchimento")

        # IMPORTANTE: Deletar TODOS os outros relatórios em "preenchimento" do mesmo projeto
        # Isso garante que apenas 1 relatório existirá após a conclusão
        duplicados = Relatorio.query.filter(
            Relatorio.id != relatorio.id,
            Relatorio.projeto_id == relatorio.projeto_id,
            Relatorio.status == 'preenchimento'
        ).all()
        
        for dup in duplicados:
            # Deletar fotos associadas ao relatório duplicado
            fotos_dup = FotoRelatorio.query.filter_by(relatorio_id=dup.id).all()
            for foto in fotos_dup:
                db.session.delete(foto)
            db.session.delete(dup)
            current_app.logger.info(f"🗑️ Deletado relatório duplicado ID={dup.id} (estava em preenchimento)")

        # COMMIT da alteração
        db.session.commit()

        if duplicados:
            current_app.logger.info(f"✅ Relatório {relatorio.numero} salvo em preenchimento - {len(duplicados)} duplicado(s) removido(s)")
            message = f'Relatório salvo em preenchimento ({len(duplicados)} duplicado(s) removido(s))'
        else:
            current_app.logger.info(f"✅ Relatório {relatorio.numero} salvo em preenchimento")
            message = 'Relatório salvo em preenchimento'

        return jsonify({
            'success': True,
            'message': message,
            'status': 'preenchimento',
            'redirect': url_for('reports')
        })

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"❌ Erro ao finalizar relatório: {e}")
        import traceback
        current_app.logger.error(f"Traceback: {traceback.format_exc()}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/reports/<int:id>/delete', methods=['GET', 'POST', 'DELETE'])
@login_required
def delete_report(id):
    """
    Excluir relatório - apenas usuários master
    Implementado conforme Item 23: Correção das Funcionalidades de Exclusão
    """
    if not current_user.is_master:
        flash('Acesso negado. Apenas usuários master podem excluir relatórios.', 'error')
        return redirect(url_for('reports'))

    relatorio = Relatorio.query.get_or_404(id)
    numero = relatorio.numero
    projeto_nome = relatorio.projeto.nome if relatorio.projeto else 'N/A'
    
    try:
        # ========== BLOCO 1: DELETAR REGISTROS DO BANCO ==========
        # Todas as operações de exclusão DEVEM ser concluídas ANTES do redirect
        
        # 1. Deletar fotos físicas e registros
        fotos = FotoRelatorio.query.filter_by(relatorio_id=id).all()
        fotos_deletadas = 0
        
        for foto in fotos:
            try:
                # Deletar arquivo físico
                upload_folder = app.config.get('UPLOAD_FOLDER', 'uploads')
                filepath = os.path.join(upload_folder, foto.filename)
                if os.path.exists(filepath):
                    os.remove(filepath)
                    current_app.logger.info(f"📁 Arquivo deletado: {filepath}")
                    fotos_deletadas += 1
            except Exception as e:
                current_app.logger.warning(f"⚠️ Erro ao deletar arquivo {foto.filename}: {str(e)}")
            
            # Deletar registro da foto
            db.session.delete(foto)
        
        # 2. Deletar notificações relacionadas (via cascade ou manual)
        try:
            from models import Notificacao
            notificacoes = Notificacao.query.filter_by(relatorio_id=id).all()
            for notif in notificacoes:
                db.session.delete(notif)
            current_app.logger.info(f"🔔 {len(notificacoes)} notificação(ões) deletada(s)")
        except Exception as e:
            current_app.logger.warning(f"⚠️ Erro ao deletar notificações: {str(e)}")
        
        # 3. Deletar logs de envio de email relacionados
        try:
            from models import LogEnvioEmail
            logs_email = LogEnvioEmail.query.filter_by(relatorio_id=id).all()
            for log in logs_email:
                db.session.delete(log)
            current_app.logger.info(f"📧 {len(logs_email)} log(s) de email deletado(s)")
        except Exception as e:
            current_app.logger.warning(f"⚠️ Erro ao deletar logs de email: {str(e)}")
        
        # 4. Deletar o relatório
        db.session.delete(relatorio)
        
        # 5. COMMIT ÚNICO de todas as exclusões
        db.session.commit()
        
        current_app.logger.info(f"✅ Relatório {numero} (Projeto: {projeto_nome}) excluído com sucesso pelo usuário {current_user.nome_completo} (ID: {current_user.id})")
        current_app.logger.info(f"   - {fotos_deletadas} foto(s) deletada(s)")
        
        flash(f'Relatório {numero} excluído com sucesso.', 'success')
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"❌ Erro ao excluir relatório {numero}: {str(e)}")
        import traceback
        current_app.logger.error(f"Traceback: {traceback.format_exc()}")
        flash(f'Erro ao excluir relatório: {str(e)}', 'error')
        return redirect(url_for('reports'))
    
    # Retornar com status 303 (See Other) conforme especificação
    from flask import redirect as flask_redirect, Response
    return flask_redirect(url_for('reports'), code=303)

@app.route('/reports/<int:report_id>/pdf')
@login_required
def generate_pdf_report(report_id):
    """Gerar PDF do relatório usando WeasyPrint (modelo Artesano) para visualização"""
    try:
        relatorio = Relatorio.query.get_or_404(report_id)
        fotos = FotoRelatorio.query.filter_by(relatorio_id=report_id).order_by(FotoRelatorio.ordem).all()

        from pdf_generator_weasy import WeasyPrintReportGenerator
        generator = WeasyPrintReportGenerator()

        # Generate PDF
        pdf_data = generator.generate_report_pdf(relatorio, fotos)

        # Create response for inline viewing
        from flask import Response
        filename = f"relatorio_{relatorio.numero.replace('/', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"

        response = Response(
            pdf_data,
            mimetype='application/pdf',
            headers={
                'Content-Disposition': f'inline; filename="{filename}"',
                'Content-Type': 'application/pdf'
            }
        )

        return response

    except Exception as e:
        flash(f'Erro ao gerar PDF: {str(e)}', 'error')
        return redirect(url_for('report_edit', report_id=report_id))

@app.route('/reports/<int:id>/pdf/download')
@login_required
def generate_report_pdf_download(id):
    """Baixar PDF do relatório usando WeasyPrint (mesmo formato da visualização)"""
    try:
        relatorio = Relatorio.query.get_or_404(id)
        fotos = FotoRelatorio.query.filter_by(relatorio_id=id).order_by(FotoRelatorio.ordem).all()

        from pdf_generator_weasy import WeasyPrintReportGenerator
        generator = WeasyPrintReportGenerator()

        # Generate PDF (mesmo conteúdo da visualização)
        pdf_data = generator.generate_report_pdf(relatorio, fotos)

        # Create response for download
        from flask import Response
        filename = f"relatorio_{relatorio.numero.replace('/', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"

        response = Response(
            pdf_data,
            mimetype='application/pdf',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"',
                'Content-Type': 'application/pdf'
            }
        )

        return response

    except Exception as e:
        flash(f'Erro ao gerar PDF: {str(e)}', 'error')
        return redirect(url_for('report_edit', report_id=report_id))

@app.route('/reports/<int:id>/pdf/legacy')
@login_required
def generate_pdf_report_legacy(id):
    """Gerar PDF do relatório usando ReportLab (versão legacy)"""
    try:
        relatorio = Relatorio.query.get_or_404(id)
        fotos = FotoRelatorio.query.filter_by(relatorio_id=id).order_by(FotoRelatorio.ordem).all()

        from pdf_generator_artesano import ArtesanoPDFGenerator
        generator = ArtesanoPDFGenerator()

        # Generate PDF
        pdf_data = generator.generate_report_pdf(relatorio, fotos)

        # Create response
        from flask import Response
        filename = f"relatorio_legacy_{relatorio.numero.replace('/', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"

        response = Response(
            pdf_data,
            mimetype='application/pdf',
            headers={
                'Content-Disposition': f'inline; filename="{filename}"',
                'Content-Type': 'application/pdf'
            }
        )

        return response

    except Exception as e:
        flash(f'Erro ao gerar PDF: {str(e)}', 'error')
        return redirect(url_for('report_edit', report_id=report_id))

def calculate_distance(lat1, lon1, lat2, lon2):
    """Calculate distance between two points using geopy"""
    from geopy.distance import geodesic
    
    # Calculate distance using geodesic (more accurate than Haversine)
    distance_km = geodesic((lat1, lon1), (lat2, lon2)).km
    
    return distance_km

def normalizar_endereco(endereco):
    """Normaliza endereços expandindo abreviações comuns"""
    import re
    
    if not endereco:
        return ''
    
    endereco = endereco.strip()
    
    # Normalizar abreviações comuns de logradouros
    endereco = re.sub(r'^(Av\.?|Avenida)\b', 'Avenida', endereco, flags=re.IGNORECASE)
    endereco = re.sub(r'^(R\.?|Rua)\b', 'Rua', endereco, flags=re.IGNORECASE)
    endereco = re.sub(r'^(Estr\.?|Estrada)\b', 'Estrada', endereco, flags=re.IGNORECASE)
    endereco = re.sub(r'^(Rod\.?|Rodovia)\b', 'Rodovia', endereco, flags=re.IGNORECASE)
    endereco = re.sub(r'^(Trav\.?|Travessa)\b', 'Travessa', endereco, flags=re.IGNORECASE)
    endereco = re.sub(r'^(Pç\.?|Praça)\b', 'Praça', endereco, flags=re.IGNORECASE)
    endereco = re.sub(r'^(Al\.?|Alameda)\b', 'Alameda', endereco, flags=re.IGNORECASE)
    
    return endereco

# ==================== NOTIFICATION API ENDPOINTS ====================

@app.route('/api/notifications/subscribe', methods=['POST'])
@login_required
def api_notifications_subscribe():
    """API para registrar subscription de push notifications"""
    try:
        data = request.get_json()
        subscription = data.get('subscription')
        user_agent = data.get('user_agent')
        timestamp = data.get('timestamp')
        
        # Log the subscription (in production, save to database)
        print(f"✅ NOTIFICATION SUBSCRIBE: User {current_user.id} ({current_user.username})")
        print(f"   Endpoint: {subscription.get('endpoint', 'N/A')[:80]}...")
        print(f"   User Agent: {user_agent}")
        print(f"   Timestamp: {timestamp}")
        
        # TODO: Save subscription to database for future use
        # For now, just acknowledge the subscription
        
        return jsonify({
            'success': True,
            'message': 'Subscription registrada com sucesso',
            'user_id': current_user.id
        })
        
    except Exception as e:
        print(f"❌ NOTIFICATION SUBSCRIBE ERROR: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/notifications/unsubscribe', methods=['POST'])
@login_required
def api_notifications_unsubscribe():
    """API para remover subscription de push notifications"""
    try:
        print(f"🔕 NOTIFICATION UNSUBSCRIBE: User {current_user.id} ({current_user.username})")
        
        # TODO: Remove subscription from database
        # For now, just acknowledge the unsubscription
        
        return jsonify({
            'success': True,
            'message': 'Subscription removida com sucesso',
            'user_id': current_user.id
        })
        
    except Exception as e:
        print(f"❌ NOTIFICATION UNSUBSCRIBE ERROR: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/notifications/check-updates')
@login_required
def api_notifications_check_updates():
    """API para verificar se há atualizações/novidades no sistema"""
    try:
        # TODO: Check for real updates in the database
        # For now, return no updates
        
        print(f"🔍 NOTIFICATION CHECK: User {current_user.id} ({current_user.username})")
        
        # Example: Check for pending reports, new visits, etc.
        # For demo, return no updates
        
        return jsonify({
            'success': True,
            'has_updates': False,
            'updates': []
        })
        
    except Exception as e:
        print(f"❌ NOTIFICATION CHECK ERROR: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ==================== END NOTIFICATION API ENDPOINTS ====================

@app.route('/api/save-annotated-photo', methods=['POST'])
@login_required  
def save_annotated_photo():
    """API para salvar foto anotada (legacy)"""
    try:
        image_data = request.form.get('annotated_image_data')
        caption = request.form.get('caption', '')
        category = request.form.get('category', '')
        description = request.form.get('description', '')
        annotations_data = request.form.get('annotations_data', '{}')

        # Para retornar via postMessage para a janela pai
        return f"""
        <script>
            if (window.opener) {{
                window.opener.postMessage({{
                    type: 'photo-edited',
                    photoId: new URLSearchParams(window.location.search).get('photoId'),
                    imageData: '{image_data}',
                    caption: '{caption}',
                    category: '{category}',
                    description: '{description}',
                    annotations: {annotations_data}
                }}, '*');
                window.close();
            }} else {{
                alert('Foto salva com sucesso!');
                window.history.back();
            }}
        </script>
        """

    except Exception as e:
        return f"""
        <script>
            alert('Erro ao salvar foto: {str(e)}');
            window.history.back();
        </script>
        """



@app.route('/debug/image/<filename>')
@login_required
def debug_image(filename):
    """Rota de diagnóstico para investigar problemas de imagem"""
    if not current_user.is_master:
        return jsonify({'error': 'Acesso negado'}), 403

    try:
        diagnostic_info = {
            'filename': filename,
            'search_results': [],
            'database_info': {},
            'directories_checked': [],
            'file_system_scan': []
        }

        # Verificar diretórios
        search_directories = [
            os.path.join(os.getcwd(), 'uploads'),
            os.path.join(os.getcwd(), 'attached_assets'),
            os.path.join(os.getcwd(), 'static', 'uploads'),
            os.path.join(os.getcwd(), 'static', 'img')
        ]

        for directory in search_directories:
            dir_info = {
                'path': directory,
                'exists': os.path.exists(directory),
                'files_count': 0,
                'target_file_found': False
            }

            if os.path.exists(directory):
                try:
                    files = os.listdir(directory)
                    dir_info['files_count'] = len([f for f in files if f.lower().endswith(('.jpg', '.jpeg', '.png'))])
                    dir_info['target_file_found'] = filename in files

                    if filename in files:
                        filepath = os.path.join(directory, filename)
                        file_stat = os.stat(filepath)
                        dir_info['file_details'] = {
                            'size': file_stat.st_size,
                            'modified': datetime.fromtimestamp(file_stat.st_mtime).isoformat(),
                            'readable': os.access(filepath, os.R_OK)
                        }
                except Exception as e:
                    dir_info['error'] = str(e)

            diagnostic_info['directories_checked'].append(dir_info)

        # Busca recursiva por arquivos similares
        base_pattern = filename[:20] if len(filename) > 20 else filename.split('.')[0]

        for root_dir in search_directories:
            if not os.path.exists(root_dir):
                continue

            for root, dirs, files in os.walk(root_dir):
                for file in files:
                    if (base_pattern in file and 
                        file.lower().endswith(('.jpg', '.jpeg', '.png', '.gif'))):
                        diagnostic_info['file_system_scan'].append({
                            'found_file': file,
                            'location': root,
                            'similarity': 'exact' if file == filename else 'pattern_match'
                        })

        # Verificar banco de dados
        try:
            from models import FotoRelatorio, FotoRelatorioExpress

            foto_relatorio = FotoRelatorio.query.filter_by(filename=filename).first()
            foto_express = FotoRelatorioExpress.query.filter_by(filename=filename).first()

            if foto_relatorio:
                diagnostic_info['database_info']['foto_relatorio'] = {
                    'id': foto_relatorio.id,
                    'relatorio_id': foto_relatorio.relatorio_id,
                    'legenda': foto_relatorio.legenda,
                    'created_at': foto_relatorio.created_at.isoformat() if foto_relatorio.created_at else None
                }

            if foto_express:
                diagnostic_info['database_info']['foto_express'] = {
                    'id': foto_express.id,
                    'relatorio_express_id': foto_express.relatorio_express_id,
                    'legenda': foto_express.legenda,
                    'created_at': foto_express.created_at.isoformat() if foto_express.created_at else None
                }

        except Exception as db_error:
            diagnostic_info['database_info']['error'] = str(db_error)

        return jsonify(diagnostic_info)

    except Exception as e:
        return jsonify({'error': str(e)}), 500




@app.route('/check-specific-image')
@login_required
def check_specific_image():
    """Verificar especificamente a imagem que está dando problema"""
    if not current_user.is_master:
        return jsonify({'error': 'Acesso negado'}), 403

    target_filename = "017050397c414b519d1df38fd32e78c8_1758143134942281964204884259198.jpg"

    try:
        results = {
            'target_filename': target_filename,
            'found_locations': [],
            'database_status': {},
            'filesystem_scan': [],
            'recommendations': []
        }

        # Verificar banco de dados primeiro
        from models import FotoRelatorio, FotoRelatorioExpress

        foto_relatorio = FotoRelatorio.query.filter_by(filename=target_filename).first()
        if foto_relatorio:
            results['database_status']['relatorio'] = {
                'id': foto_relatorio.id,
                'relatorio_id': foto_relatorio.relatorio_id,
                'legenda': foto_relatorio.legenda,
                'created_at': foto_relatorio.created_at.isoformat() if foto_relatorio.created_at else None,
                'ordem': foto_relatorio.ordem
            }

        # Buscar arquivo no filesystem
        search_locations = [
            'uploads',
            'attached_assets', 
            'static/uploads',
            'static/img',
            '.'
        ]

        for location in search_locations:
            full_path = os.path.join(os.getcwd(), location)
            if os.path.exists(full_path):
                # Busca exata
                target_path = os.path.join(full_path, target_filename)
                if os.path.exists(target_path):
                    file_stat = os.stat(target_path)
                    results['found_locations'].append({
                        'location': location,
                        'full_path': target_path,
                        'size': file_stat.st_size,
                        'modified': datetime.fromtimestamp(file_stat.st_mtime).isoformat(),
                        'readable': os.access(target_path, os.R_OK)
                    })

                # Busca recursiva
                for root, dirs, files in os.walk(full_path):
                    if target_filename in files:
                        file_path = os.path.join(root, target_filename)
                        if file_path not in [loc['full_path'] for loc in results['found_locations']]:
                            file_stat = os.stat(file_path)
                            results['found_locations'].append({
                                'location': f"{location} (subdir: {os.path.relpath(root, full_path)})",
                                'full_path': file_path,
                                'size': file_stat.st_size,
                                'modified': datetime.fromtimestamp(file_stat.st_mtime).isoformat(),
                                'readable': os.access(file_path, os.R_OK)
                            })

        # Buscar por padrão similar
        base_pattern = target_filename[:30]
        for location in search_locations:
            full_path = os.path.join(os.getcwd(), location)
            if os.path.exists(full_path):
                for root, dirs, files in os.walk(full_path):
                    for file in files:
                        if (base_pattern in file and 
                            file.lower().endswith(('.jpg', '.jpeg')) and
                            file != target_filename):
                            results['filesystem_scan'].append({
                                'similar_file': file,
                                'location': os.path.relpath(root, os.getcwd()),
                                'similarity_score': len(set(target_filename) & set(file)) / len(set(target_filename) | set(file))
                            })

        # Gerar recomendações
        if results['database_status'] and not results['found_locations']:
            results['recommendations'].append("Arquivo existe no banco mas não no filesystem - possível problema de upload ou migração")

        if results['found_locations']:
            results['recommendations'].append(f"Arquivo encontrado em {len(results['found_locations'])} localização(ões)")

        if results['filesystem_scan']:
            results['recommendations'].append(f"Encontrados {len(results['filesystem_scan'])} arquivos similares")
            results['recommendations'].append("Verificar se houve corrupção de nome durante upload")

        return jsonify(results)

    except Exception as e:
        return jsonify({'error': str(e)}), 500




@app.route('/reports/<int:report_id>/photos/upload', methods=['POST'])
@login_required
def upload_report_photos(report_id):
    relatorio = Relatorio.query.get_or_404(report_id)

    # Check permissions
    if not current_user.is_master and relatorio.autor_id != current_user.id:
        return jsonify({'success': False, 'error': 'Acesso negado'}), 403

    try:
        files = request.files.getlist('photos')
        uploaded_count = 0
        duplicated_count = 0

        for file in files:
            if file.filename and file.filename.lower().endswith(('.png', '.jpg', '.jpeg')):
                # Read file data before saving
                file_data = file.read()
                
                # 🔧 CORREÇÃO CRÍTICA: Calcular hash SHA-256 para prevenir duplicação
                import hashlib
                imagem_hash = hashlib.sha256(file_data).hexdigest()
                
                # Verificar se imagem JÁ EXISTE no banco (prevenir duplicação)
                foto_existente = FotoRelatorio.query.filter_by(
                    relatorio_id=relatorio.id,
                    imagem_hash=imagem_hash
                ).first()
                
                if foto_existente:
                    # Imagem já existe - não duplicar
                    current_app.logger.info(f"⚠️ Imagem duplicada detectada (hash={imagem_hash[:12]}...) - ID existente={foto_existente.id}. Não será duplicada.")
                    duplicated_count += 1
                    continue
                
                # Imagem NÃO existe - criar nova
                file.seek(0)  # Reset for saving to disk
                
                # Generate unique filename
                filename = f"{uuid.uuid4().hex}_{secure_filename(file.filename)}"
                filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)

                # Save file
                file.save(filepath)
                current_app.logger.info(f"✅ FOTO SALVA: {filepath}")

                # Create photo record with binary data
                foto = FotoRelatorio()
                foto.relatorio_id = relatorio.id
                foto.filename = filename
                foto.legenda = f'Foto {uploaded_count + 1}'
                foto.ordem = uploaded_count + 1
                foto.imagem = file_data  # Salvar dados binários da imagem
                foto.imagem_hash = imagem_hash
                foto.imagem_size = len(file_data)

                db.session.add(foto)
                uploaded_count += 1

        db.session.commit()

        message = f'{uploaded_count} foto(s) enviada(s) com sucesso!'
        if duplicated_count > 0:
            message += f' ({duplicated_count} duplicada(s) ignorada(s))'

        return jsonify({
            'success': True, 
            'message': message,
            'uploaded': uploaded_count,
            'duplicated': duplicated_count
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500





@app.route('/projects/new', methods=['GET', 'POST'])
@login_required
def project_new():
    # Only master users can create new projects
    if not current_user.is_master:
        flash('Apenas o usuário master pode criar novas obras.', 'danger')
        return redirect(url_for('projects_list'))
    
    form = ProjetoForm()

    if request.method == 'POST':
        print(f"🔍 DEBUG: Form data received: {dict(request.form)}")
        print(f"🔍 DEBUG: Form validation: {form.validate_on_submit()}")
        if form.errors:
            print(f"🔍 DEBUG: Form errors: {form.errors}")

    if form.validate_on_submit():
        try:
            # Extract contatos unificados do formulário
            contatos = []

            # Processar contatos
            for key in request.form.keys():
                if key.startswith('contatos[') and key.endswith('][nome]'):
                    index = key.split('[')[1].split(']')[0]
                    nome = request.form.get(f'contatos[{index}][nome]')
                    cargo = request.form.get(f'contatos[{index}][cargo]', '')
                    empresa = request.form.get(f'contatos[{index}][empresa]', '')
                    email = request.form.get(f'contatos[{index}][email]', '')
                    telefone = request.form.get(f'contatos[{index}][telefone]', '')

                    if nome or email:  # Pelo menos nome ou email deve estar preenchido
                        contatos.append({
                            'nome': nome,
                            'cargo': cargo,
                            'empresa': empresa,
                            'email': email,
                            'telefone': telefone
                        })

            # Process categorias - Item 16
            categorias_adicionais = []
            for key in request.form.keys():
                if key.startswith('categorias[') and key.endswith('][nome]'):
                    index = key.split('[')[1].split(']')[0]
                    nome_categoria = request.form.get(f'categorias[{index}][nome]')
                    ordem = request.form.get(f'categorias[{index}][ordem]', 0)

                    if nome_categoria:
                        categorias_adicionais.append({
                            'nome': nome_categoria.strip(),
                            'ordem': int(ordem) if ordem else 0
                        })

            print(f"🔍 DEBUG: Found {len(contatos)} contatos unificados")
            print(f"🔍 DEBUG: Found {len(categorias_adicionais)} categorias - Item 16")

            # Check if project with same name already exists
            existing_project = Projeto.query.filter_by(nome=form.nome.data).first()

            if existing_project:
                # Project consolidation: add employee and email to existing project
                print(f"🔍 DEBUG: Project '{form.nome.data}' already exists. Consolidating...")
                projeto = existing_project
                funcionarios_adicionados = 0
                emails_adicionados = 0

                # Adicionar contatos ao projeto existente
                contatos_adicionados = 0
                for contato_data in contatos:
                    # Verificar se contato já existe (por email se fornecido)
                    existing_email = None
                    if contato_data.get('email'):
                        existing_email = EmailCliente.query.filter_by(
                            projeto_id=projeto.id,
                            email=contato_data['email']
                        ).first()

                    if not existing_email:
                        novo_contato = EmailCliente(
                            projeto_id=projeto.id,
                            email=contato_data.get('email') or None,  # Permitir None se não houver email
                            nome_contato=contato_data.get('nome', 'Sem nome'),
                            cargo=contato_data.get('cargo', ''),
                            empresa=contato_data.get('empresa', ''),
                            telefone=contato_data.get('telefone', ''),  # Adicionar telefone
                            ativo=True
                        )
                        db.session.add(novo_contato)
                        contatos_adicionados += 1

                # Add categorias to existing project - Item 16
                categorias_adicionadas = 0
                for categoria_data in categorias_adicionais:
                    # Check if category already exists (by name)
                    existing_categoria = CategoriaObra.query.filter_by(
                        projeto_id=projeto.id,
                        nome_categoria=categoria_data['nome']
                    ).first()

                    if not existing_categoria:
                        nova_categoria = CategoriaObra(
                            projeto_id=projeto.id,
                            nome_categoria=categoria_data['nome'],
                            ordem=categoria_data['ordem']
                        )
                        db.session.add(nova_categoria)
                        categorias_adicionadas += 1

                flash(f'Obra consolidada! Adicionados {contatos_adicionados} contato(s) e {categorias_adicionadas} categoria(s) à obra existente: {projeto.nome}', 'success')

            else:
                # Create new project
                print(f"🔍 DEBUG: Creating new project: {form.nome.data}")
                projeto = Projeto()
                projeto.numero = generate_project_number()
                projeto.nome = form.nome.data
                projeto.descricao = 'Projeto criado através do sistema ELP'  # Default value since field was removed
                projeto.endereco = form.endereco.data
                projeto.latitude = float(form.latitude.data) if form.latitude.data else None
                projeto.longitude = float(form.longitude.data) if form.longitude.data else None

                # Automatic geocoding: if no GPS coordinates but address exists, convert address to coordinates
                if not projeto.latitude or not projeto.longitude:
                    if projeto.endereco and projeto.endereco.strip():
                        print(f"🔍 GEOCODING: Tentando converter endereço '{projeto.endereco}' para coordenadas GPS...")
                        lat, lng = get_coordinates_from_address(projeto.endereco)
                        if lat and lng:
                            projeto.latitude = lat
                            projeto.longitude = lng
                            print(f"✅ GEOCODING: Sucesso! Coordenadas: {lat}, {lng}")
                        else:
                            print(f"❌ GEOCODING: Não foi possível converter o endereço")

                projeto.tipo_obra = 'Geral'  # Default value since field was removed
                projeto.construtora = form.construtora.data
                projeto.nome_funcionario = contatos[0].get('nome', '') if contatos else ''  # Primeiro contato como funcionário padrão
                projeto.responsavel_id = form.responsavel_id.data
                projeto.email_principal = contatos[0].get('email', '') if contatos else ''  # Primeiro contato como email padrão
                projeto.data_inicio = form.data_inicio.data
                projeto.data_previsao_fim = form.data_previsao_fim.data
                projeto.status = form.status.data
                projeto.numeracao_inicial = form.numeracao_inicial.data or 1

                db.session.add(projeto)
                db.session.flush()  # Get the project ID

                # Adicionar contatos unificados
                contatos_adicionados = 0
                for i, contato_data in enumerate(contatos):
                    novo_contato = EmailCliente(
                        projeto_id=projeto.id,
                        email=contato_data.get('email') or None,  # Permitir None se não houver email
                        nome_contato=contato_data.get('nome', 'Sem nome'),
                        cargo=contato_data.get('cargo', ''),
                        empresa=contato_data.get('empresa', ''),
                        telefone=contato_data.get('telefone', ''),  # Adicionar telefone
                        ativo=True
                    )
                    db.session.add(novo_contato)
                    contatos_adicionados += 1

                # Add categorias - Item 16
                for categoria_data in categorias_adicionais:
                    categoria = CategoriaObra(
                        projeto_id=projeto.id,
                        nome_categoria=categoria_data['nome'],
                        ordem=categoria_data['ordem']
                    )
                    db.session.add(categoria)

                total_contatos = len(contatos)
                total_categorias = len(categorias_adicionais)
                flash(f'Obra cadastrada com sucesso! {total_contatos} contato(s) e {total_categorias} categoria(s) adicionados.', 'success')

            db.session.commit()
            print(f"🔍 DEBUG: Trying to save projeto: {projeto.nome}")
            print(f"✅ DEBUG: Projeto saved successfully!")
            
            # Criar notificações de obra criada apenas para novos projetos
            if not existing_project:
                from notification_service import notification_service
                try:
                    notification_service.criar_notificacao_obra_criada(projeto.id)
                    current_app.logger.info(f"✅ Notificações de obra criada enviadas para projeto {projeto.id}")
                except Exception as notif_error:
                    current_app.logger.error(f"⚠️ Erro ao criar notificações de obra criada: {notif_error}")
            
            return redirect(url_for('projects_list'))
        except Exception as e:
            print(f"❌ DEBUG: Error saving projeto: {e}")
            db.session.rollback()
            flash(f'Erro ao salvar obra: {str(e)}', 'error')

    return render_template('projects/form.html', form=form, contatos_existentes=[])

@app.route('/projects/<int:project_id>')
@login_required
def project_view(project_id):
    project = Projeto.query.get_or_404(project_id)
    contatos = ContatoProjeto.query.filter_by(projeto_id=project_id).all()
    visitas = Visita.query.filter_by(projeto_id=project_id).order_by(Visita.data_inicio.desc()).all()
    relatorios = Relatorio.query.filter_by(projeto_id=project_id).order_by(Relatorio.created_at.desc()).all()
    relatorios_express = RelatorioExpress.query.order_by(RelatorioExpress.created_at.desc()).all()

    # Get communications from all visits of this project
    comunicacoes = []
    for visita in visitas:
        visit_comunicacoes = ComunicacaoVisita.query.filter_by(visita_id=visita.id).order_by(ComunicacaoVisita.created_at.desc()).all()
        for com in visit_comunicacoes:
            comunicacoes.append({
                'comunicacao': com,
                'visita': visita
            })

    # Sort all communications by date
    comunicacoes.sort(key=lambda x: x['comunicacao'].created_at, reverse=True)

    return render_template('projects/view.html', 
                         project=project, 
                         visitas=visitas, 
                         relatorios=relatorios,
                         relatorios_express=relatorios_express,
                         comunicacoes=comunicacoes[:10])  # Show last 10 communications

@app.route('/projects/<int:project_id>/reactivate', methods=['POST'])
@login_required
def reactivate_project(project_id):
    project = Projeto.query.get_or_404(project_id)
    
    # Permitir apenas usuário master
    if not current_user.is_master:
        flash('Acesso negado. Apenas o usuário master pode reativar obras.', 'danger')
        return redirect(url_for('project_view', project_id=project_id))
    
    # Verifica status atual
    if project.status.lower() == 'concluído':
        project.status = 'Ativo'
        db.session.commit()
        flash('Obra reativada com sucesso!', 'success')
    else:
        flash('A obra já está ativa.', 'info')
    
    return redirect(url_for('project_view', project_id=project_id))

@app.route('/projects/<int:project_id>/edit', methods=['GET', 'POST'])
@login_required
def project_edit(project_id):
    project = Projeto.query.get_or_404(project_id)
    
    # Only master users can edit projects
    if not current_user.is_master:
        flash('Apenas o usuário master pode editar obras.', 'danger')
        return redirect(url_for('project_view', project_id=project_id))
    
    form = ProjetoForm(obj=project)
    
    # Buscar categorias existentes do projeto - Item 16 (Fix)
    categorias_existentes = CategoriaObra.query.filter_by(projeto_id=project_id).order_by(CategoriaObra.ordem).all()
    
    # Serializar corretamente com fallback para evitar undefined
    categorias_serializadas = []
    for c in categorias_existentes:
        categorias_serializadas.append({
            "id": c.id,
            "nome": getattr(c, "nome", None) or getattr(c, "nome_categoria", None) or "",  # fallback de campo
            "ordem": getattr(c, "ordem", 0),
            "project_id": c.projeto_id
        })
    
    # Carregar contatos existentes (para GET e POST)
    contatos_existentes = EmailCliente.query.filter_by(projeto_id=project_id).all()

    if form.validate_on_submit():
        try:
            project.nome = form.nome.data
            project.descricao = 'Projeto atualizado através do sistema ELP'  # Default value since field was removed
            project.endereco = form.endereco.data
            project.latitude = float(form.latitude.data) if form.latitude.data else None
            project.longitude = float(form.longitude.data) if form.longitude.data else None

            # Automatic geocoding: if no GPS coordinates but address exists, convert address to coordinates
            if not project.latitude or not project.longitude:
                if project.endereco and project.endereco.strip():
                    print(f"🔍 GEOCODING: Tentando converter endereço '{project.endereco}' para coordenadas GPS...")
                    lat, lng = get_coordinates_from_address(project.endereco)
                    if lat and lng:
                        project.latitude = lat
                        project.longitude = lng
                        print(f"✅ GEOCODING: Sucesso! Coordenadas: {lat}, {lng}")
                    else:
                        print(f"❌ GEOCODING: Não foi possível converter o endereço")

            project.tipo_obra = 'Geral'  # Default value since field was removed
            project.construtora = form.construtora.data
            project.responsavel_id = form.responsavel_id.data
            project.data_inicio = form.data_inicio.data
            project.data_previsao_fim = form.data_previsao_fim.data
            project.status = form.status.data
            project.numeracao_inicial = form.numeracao_inicial.data or 1
            
            # Processar exclusões de contatos
            contatos_excluidos = request.form.getlist('contatos_excluidos[]')
            for contato_id in contatos_excluidos:
                if contato_id:
                    contato = EmailCliente.query.get(int(contato_id))
                    if contato and contato.projeto_id == project.id:
                        db.session.delete(contato)
            
            # Processar contatos usando indexed notation (contatos[0][nome], contatos[0][email], etc.)
            # Extract all unique contact indices from form keys
            contatos = []
            contact_indices = set()
            for key in request.form.keys():
                if key.startswith('contatos[') and '][' in key:
                    index = key.split('[')[1].split(']')[0]
                    contact_indices.add(index)
            
            # Process each contact by index
            for index in sorted(contact_indices):
                nome = request.form.get(f'contatos[{index}][nome]', '')
                cargo = request.form.get(f'contatos[{index}][cargo]', '')
                empresa = request.form.get(f'contatos[{index}][empresa]', '')
                email = request.form.get(f'contatos[{index}][email]', '')
                telefone = request.form.get(f'contatos[{index}][telefone]', '')
                contato_id = request.form.get(f'contatos[{index}][id]', None)

                if nome or email:  # Pelo menos nome ou email deve estar preenchido
                    contatos.append({
                        'nome': nome,
                        'cargo': cargo,
                        'empresa': empresa,
                        'email': email,
                        'telefone': telefone,
                        'id': contato_id
                    })
            
            # Processar cada contato
            for contato_data in contatos:
                if contato_data.get('id'):
                    # Atualizar contato existente
                    contato = EmailCliente.query.get(int(contato_data['id']))
                    if contato and contato.projeto_id == project.id:
                        contato.nome_contato = contato_data['nome']
                        contato.cargo = contato_data['cargo']
                        contato.empresa = contato_data['empresa']
                        contato.email = contato_data['email'] or None
                        contato.telefone = contato_data['telefone']
                else:
                    # Criar novo contato
                    novo_contato = EmailCliente(
                        projeto_id=project.id,
                        nome_contato=contato_data['nome'],
                        cargo=contato_data['cargo'],
                        empresa=contato_data['empresa'],
                        email=contato_data['email'] or None,
                        telefone=contato_data['telefone'],
                        ativo=True
                    )
                    db.session.add(novo_contato)
            
            # Atualizar campos legados do projeto com primeiro contato
            primeiro_contato = EmailCliente.query.filter_by(projeto_id=project.id).first()
            if primeiro_contato:
                project.nome_funcionario = primeiro_contato.nome_contato
                project.email_principal = primeiro_contato.email or ''
            else:
                project.nome_funcionario = ''
                project.email_principal = ''

            # Process categorias from form - Item 16 (Fix)
            categorias_adicionais = []
            for key in request.form.keys():
                if key.startswith('categorias[') and key.endswith('][nome]'):
                    index = key.split('[')[1].split(']')[0]
                    nome_categoria = request.form.get(f'categorias[{index}][nome]')
                    ordem = request.form.get(f'categorias[{index}][ordem]', 0)

                    if nome_categoria:
                        categorias_adicionais.append({
                            'nome': nome_categoria.strip(),
                            'ordem': int(ordem) if ordem else 0
                        })

            # Add new categorias to the project
            categorias_adicionadas = 0
            for categoria_data in categorias_adicionais:
                # Check if category already exists (by name)
                existing_categoria = CategoriaObra.query.filter_by(
                    projeto_id=project.id,
                    nome_categoria=categoria_data['nome']
                ).first()

                if not existing_categoria:
                    nova_categoria = CategoriaObra(
                        projeto_id=project.id,
                        nome_categoria=categoria_data['nome'],
                        ordem=categoria_data['ordem']
                    )
                    db.session.add(nova_categoria)
                    categorias_adicionadas += 1

            db.session.commit()
            
            if categorias_adicionadas > 0:
                flash(f'Obra atualizada com sucesso! {categorias_adicionadas} nova(s) categoria(s) adicionada(s).', 'success')
            else:
                flash('Obra atualizada com sucesso!', 'success')
            
            return redirect(url_for('project_view', project_id=project.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar obra: {str(e)}', 'error')
    
    return render_template('projects/form.html', form=form, project=project, categorias=categorias_serializadas, contatos_existentes=contatos_existentes)

# Category management routes - Item 16
@app.route('/projects/<int:project_id>/categorias')
@login_required
def project_categorias_list(project_id):
    """Lista categorias de um projeto"""
    project = Projeto.query.get_or_404(project_id)
    categorias = CategoriaObra.query.filter_by(projeto_id=project_id).order_by(CategoriaObra.ordem).all()
    return jsonify({
        'success': True,
        'categorias': [{
            'id': c.id,
            'nome_categoria': c.nome_categoria,
            'ordem': c.ordem
        } for c in categorias]
    })

@app.route('/projects/<int:project_id>/categorias/add', methods=['POST'])
@login_required
@csrf.exempt
def project_categoria_add(project_id):
    """Adiciona uma nova categoria ao projeto"""
    project = Projeto.query.get_or_404(project_id)
    
    try:
        data = request.get_json()
        nome_categoria = data.get('nome_categoria', '').strip()
        ordem = data.get('ordem', 0)
        
        if not nome_categoria:
            return jsonify({'error': 'Nome da categoria é obrigatório'}), 400
        
        # Verificar duplicação
        existe = CategoriaObra.query.filter_by(
            projeto_id=project_id,
            nome_categoria=nome_categoria
        ).first()
        
        if existe:
            return jsonify({'error': 'Categoria já existe para este projeto'}), 400
        
        # Se não informou ordem, usar a próxima sequencial
        if not ordem:
            max_ordem = db.session.query(db.func.max(CategoriaObra.ordem)).filter_by(projeto_id=project_id).scalar() or 0
            ordem = max_ordem + 1
        
        categoria = CategoriaObra(
            projeto_id=project_id,
            nome_categoria=nome_categoria,
            ordem=ordem
        )
        
        db.session.add(categoria)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'categoria': {
                'id': categoria.id,
                'nome_categoria': categoria.nome_categoria,
                'ordem': categoria.ordem
            }
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/projects/<int:project_id>/categorias/<int:categoria_id>/edit', methods=['PUT'])
@login_required
@csrf.exempt
def project_categoria_edit(project_id, categoria_id):
    """Edita uma categoria existente"""
    categoria = CategoriaObra.query.filter_by(id=categoria_id, projeto_id=project_id).first_or_404()
    
    try:
        data = request.get_json()
        nome_categoria = data.get('nome_categoria', '').strip()
        ordem = data.get('ordem')
        
        if not nome_categoria:
            return jsonify({'error': 'Nome da categoria é obrigatório'}), 400
        
        # Verificar duplicação (exceto a própria categoria)
        existe = CategoriaObra.query.filter(
            CategoriaObra.projeto_id == project_id,
            CategoriaObra.nome_categoria == nome_categoria,
            CategoriaObra.id != categoria_id
        ).first()
        
        if existe:
            return jsonify({'error': 'Categoria já existe para este projeto'}), 400
        
        categoria.nome_categoria = nome_categoria
        if ordem is not None:
            categoria.ordem = ordem
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'categoria': {
                'id': categoria.id,
                'nome_categoria': categoria.nome_categoria,
                'ordem': categoria.ordem
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/projects/<int:project_id>/categorias/<int:categoria_id>/delete', methods=['DELETE'])
@login_required
@csrf.exempt
def project_categoria_delete(project_id, categoria_id):
    """Remove uma categoria (mantém histórico se houver fotos vinculadas)"""
    categoria = CategoriaObra.query.filter_by(id=categoria_id, projeto_id=project_id).first_or_404()
    
    try:
        # Aqui poderíamos verificar se há fotos vinculadas e manter histórico
        # Por enquanto, apenas removemos
        db.session.delete(categoria)
        db.session.commit()
        
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@app.route('/api/projects/<int:project_id>/categorias')
@login_required
def api_project_categorias(project_id):
    """API para obter categorias de um projeto para uso em forms"""
    categorias = CategoriaObra.query.filter_by(projeto_id=project_id).order_by(CategoriaObra.ordem).all()
    
    # Retorna categorias do banco de dados ou lista vazia
    return jsonify({
        'categorias': [{
            'id': c.id,
            'nome': c.nome_categoria
        } for c in categorias],
        'has_categories': len(categorias) > 0
    })

# AJAX routes for category management - conforme orientações do prompt
@app.route('/api/categorias/<int:id>/update', methods=['POST'])
@login_required
@csrf.exempt
def update_categoria(id):
    """Atualiza uma categoria via AJAX"""
    categoria = CategoriaObra.query.get_or_404(id)
    data = request.get_json()
    # Aceita tanto 'nome' quanto 'nome_categoria' para compatibilidade
    nome = data.get('nome') or data.get('nome_categoria')
    if nome:
        categoria.nome_categoria = nome
    if 'ordem' in data:
        categoria.ordem = data.get('ordem')
    db.session.commit()
    return jsonify({"success": True, "message": "Categoria atualizada com sucesso"}), 200

@app.route('/api/categorias/<int:id>/delete', methods=['DELETE'])
@login_required
@csrf.exempt
def delete_categoria(id):
    """Exclui uma categoria via AJAX"""
    categoria = CategoriaObra.query.get_or_404(id)
    db.session.delete(categoria)
    db.session.commit()
    return jsonify({"success": True, "message": "Categoria excluída com sucesso"}), 200

# Contact management routes
@app.route('/contacts')
@login_required
def contacts_list():
    contacts = Contato.query.all()
    return render_template('contacts/list.html', contacts=contacts)

@app.route('/contacts/new', methods=['GET', 'POST'])
@login_required
def contact_new():
    form = ContatoForm()

    if form.validate_on_submit():
        contato = Contato(
            nome=form.nome.data,
            email=form.email.data,
            telefone=form.telefone.data,
            empresa=form.empresa.data,
            cargo=form.cargo.data,
            observacoes=form.observacoes.data
        )

        db.session.add(contato)
        db.session.commit()
        flash('Contato cadastrado com sucesso!', 'success')
        return redirect(url_for('contacts_list'))

    return render_template('contacts/form.html', form=form)

@app.route('/contacts/<int:contact_id>/edit', methods=['GET', 'POST'])
@login_required
def contact_edit(contact_id):
    contact = Contato.query.get_or_404(contact_id)


@app.route('/admin/diagnostico-imagens')
@login_required
def diagnostico_imagens():
    """Diagnóstico completo das imagens no sistema"""
    if not current_user.is_master:
        return jsonify({'error': 'Acesso negado'}), 403

    try:
        from models import FotoRelatorio, FotoRelatorioExpress

        # Buscar todas as fotos do banco
        fotos_normais = FotoRelatorio.query.all()
        fotos_express = FotoRelatorioExpress.query.all()

        diagnostico = {
            'total_banco': len(fotos_normais) + len(fotos_express),
            'fotos_normais': len(fotos_normais),
            'fotos_express': len(fotos_express),
            'existem_fisicamente': 0,
            'nao_existem_fisicamente': 0,
            'arquivos_perdidos': [],
            'arquivos_ok': [],
            'em_attached_assets': 0,
            'timestamp': datetime.utcnow().isoformat()
        }

        upload_folder = app.config.get('UPLOAD_FOLDER', 'uploads')

        # Verificar fotos normais
        for foto in fotos_normais:
            filepath = os.path.join(upload_folder, foto.filename)
            attached_path = os.path.join('attached_assets', foto.filename)

            if os.path.exists(filepath):
                diagnostico['existem_fisicamente'] += 1
                diagnostico['arquivos_ok'].append({
                    'filename': foto.filename,
                    'tipo': 'normal',
                    'relatorio_id': foto.relatorio_id,
                    'localizacao': 'uploads'
                })
            elif os.path.exists(attached_path):
                diagnostico['em_attached_assets'] += 1
                diagnostico['arquivos_ok'].append({
                    'filename': foto.filename,
                    'tipo': 'normal',
                    'relatorio_id': foto.relatorio_id,
                    'localizacao': 'attached_assets'
                })
            else:
                diagnostico['nao_existem_fisicamente'] += 1
                diagnostico['arquivos_perdidos'].append({
                    'filename': foto.filename,
                    'tipo': 'normal',
                    'relatorio_id': foto.relatorio_id,
                    'legenda': foto.legenda
                })

        # Verificar fotos express
        for foto in fotos_express:
            filepath = os.path.join(upload_folder, foto.filename)
            attached_path = os.path.join('attached_assets', foto.filename)

            if os.path.exists(filepath):
                diagnostico['existem_fisicamente'] += 1
                diagnostico['arquivos_ok'].append({
                    'filename': foto.filename,
                    'tipo': 'express',
                    'relatorio_id': foto.relatorio_express_id,
                    'localizacao': 'uploads'
                })
            elif os.path.exists(attached_path):
                diagnostico['em_attached_assets'] += 1
                diagnostico['arquivos_ok'].append({
                    'filename': foto.filename,
                    'tipo': 'express',
                    'relatorio_id': foto.relatorio_express_id,
                    'localizacao': 'attached_assets'
                })
            else:
                diagnostico['nao_existem_fisicamente'] += 1
                diagnostico['arquivos_perdidos'].append({
                    'filename': foto.filename,
                    'tipo': 'express',
                    'relatorio_id': foto.relatorio_express_id,
                    'legenda': foto.legenda
                })

        current_app.logger.info(f"📊 DIAGNÓSTICO: {diagnostico['total_banco']} total, {diagnostico['existem_fisicamente']} OK, {diagnostico['nao_existem_fisicamente']} perdidas")

        return jsonify(diagnostico)

    except Exception as e:
        current_app.logger.error(f"❌ Erro no diagnóstico: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/admin/migrar-attached-assets')
@login_required
def migrar_attached_assets():
    """Migrar arquivos de attached_assets para uploads"""
    if not current_user.is_master:
        return jsonify({'error': 'Acesso negado'}), 403

    try:
        from models import FotoRelatorio, FotoRelatorioExpress
        import shutil

        upload_folder = app.config.get('UPLOAD_FOLDER', 'uploads')
        if not os.path.exists(upload_folder):
            os.makedirs(upload_folder)

        migradas = []
        erros = []

        # Buscar todas as fotos do banco
        fotos_normais = FotoRelatorio.query.all()
        fotos_express = FotoRelatorioExpress.query.all()

        todas_fotos = []
        for foto in fotos_normais:
            todas_fotos.append({
                'filename': foto.filename,
                'tipo': 'normal',
                'relatorio_id': foto.relatorio_id
            })

        for foto in fotos_express:
            todas_fotos.append({
                'filename': foto.filename,
                'tipo': 'express',
                'relatorio_id': foto.relatorio_express_id
            })

        for foto_info in todas_fotos:
            filename = foto_info['filename']
            upload_path = os.path.join(upload_folder, filename)
            attached_path = os.path.join('attached_assets', filename)

            # Se não existe em uploads mas existe em attached_assets
            if not os.path.exists(upload_path) and os.path.exists(attached_path):
                try:
                    shutil.copy2(attached_path, upload_path)
                    migradas.append({
                        'filename': filename,
                        'tipo': foto_info['tipo'],
                        'relatorio_id': foto_info['relatorio_id'],
                        'origem': attached_path,
                        'destino': upload_path
                    })
                    current_app.logger.info(f"✅ MIGRADA: {filename}")
                except Exception as e:
                    erros.append({
                        'filename': filename,
                        'erro': str(e)
                    })
                    current_app.logger.error(f"❌ Erro ao migrar {filename}: {str(e)}")

        resultado = {
            'success': True,
            'migradas': migradas,
            'erros': erros,
            'total_migradas': len(migradas),
            'total_erros': len(erros),
            'message': f'{len(migradas)} arquivos migrados, {len(erros)} erros'
        }

        current_app.logger.info(f"📦 MIGRAÇÃO COMPLETA: {len(migradas)} arquivos migrados")

        return jsonify(resultado)

    except Exception as e:
        current_app.logger.error(f"❌ Erro na migração: {str(e)}")
        return jsonify({'error': str(e)}), 500

    form = ContatoForm(obj=contact)

    if form.validate_on_submit():
        contact.nome = form.nome.data
        contact.email = form.email.data
        contact.telefone = form.telefone.data
        contact.empresa = form.empresa.data
        contact.cargo = form.cargo.data
        contact.observacoes = form.observacoes.data

        db.session.commit()
        flash('Contato atualizado com sucesso!', 'success')
        return redirect(url_for('contacts_list'))

# Contact functionality removed as requested

# Visit management routes
@app.route('/visits')
@login_required
def visits_list():
    """Lista de visitas - versão ultra-robusta com tratamento completo de erros"""
    try:
        current_app.logger.info(f"📋 /visits: Usuário {current_user.username} acessando lista de visitas")

        # Check if user wants calendar view
        view_type = request.args.get('view', 'list')
        if view_type == 'calendar':
            try:
                return render_template('visits/calendar.html')
            except Exception as template_error:
                current_app.logger.error(f"❌ Erro no template calendar: {template_error}")
                flash('Erro ao carregar calendário. Exibindo lista.', 'warning')
                view_type = 'list'

        # Get search query parameter
        q = request.args.get('q', '').strip()

        # Query robusta com múltiplos fallbacks
        visits = []
        try:
            # Tentativa 1: Query completa
            base_query = Visita.query

            if q:
                from sqlalchemy import or_
                search_term = f"%{q}%"
                base_query = base_query.filter(or_(
                    Visita.numero.ilike(search_term),
                    Visita.observacoes.ilike(search_term),
                    Visita.projeto_outros.ilike(search_term)
                ))

            visits = base_query.order_by(Visita.data_inicio.desc()).limit(50).all()
            current_app.logger.info(f"✅ {len(visits)} visitas carregadas com sucesso")

        except Exception as query_error:
            current_app.logger.error(f"❌ Erro na query principal: {str(query_error)}")
            try:
                db.session.rollback()
                # Tentativa 2: Query simples sem filtros
                visits = Visita.query.order_by(Visita.id.desc()).limit(10).all()
                current_app.logger.info(f"🔄 Fallback 1: {len(visits)} visitas carregadas")
            except Exception as fallback_error:
                current_app.logger.error(f"❌ Erro no fallback 1: {str(fallback_error)}")
                try:
                    db.session.rollback()
                    # Tentativa 3: Query mínima
                    visits = Visita.query.limit(5).all()
                    current_app.logger.info(f"🔄 Fallback 2: {len(visits)} visitas carregadas")
                except Exception:
                    # Tentativa 4: Lista vazia (última opção)
                    visits = []
                    current_app.logger.error("❌ Todos os fallbacks falharam - retornando lista vazia")
                    flash('Não foi possível carregar as visitas. Tente novamente.', 'warning')

        # Garantir que sempre temos uma lista válida
        if not isinstance(visits, list):
            visits = list(visits) if visits else []

        # Verificar cada visita para garantir que as propriedades funcionam
        safe_visits = []
        for visit in visits:
            try:
                # Testar acesso às propriedades críticas
                _ = visit.numero
                _ = visit.status or 'Agendada'
                safe_visits.append(visit)
            except Exception as prop_error:
                current_app.logger.warning(f"⚠️ Visita {visit.id} com propriedades inválidas: {prop_error}")
                # Pular esta visita específica
                continue

        visits = safe_visits

        # Renderizar template com tratamento de erro
        try:
            return render_template('visits/list.html', visits=visits)
        except Exception as template_error:
            current_app.logger.error(f"❌ Erro no template visits/list.html: {template_error}")
            # Template de emergência em caso de erro
            emergency_html = f'''
            <!DOCTYPE html>
            <html>
            <head><title>Visitas - ELP</title></head>
            <body>
                <h1>Sistema de Visitas</h1>
                <p>Encontradas {len(visits)} visitas.</p>
                <p>Erro no template principal. <a href="/visits">Tentar novamente</a></p>
            </body>
            </html>
            '''
            return emergency_html, 200

    except Exception as e:
        current_app.logger.exception(f"❌ ERRO CRÍTICO na rota /visits: {str(e)}")

        # Resposta de emergência absoluta
        try:
            flash('Erro temporário ao carregar visitas. Tente novamente.', 'error')
            return render_template('visits/list.html', visits=[])
        except Exception:
            # Se até o template de erro falhar, retornar HTML simples
            return '''
            <!DOCTYPE html>
            <html>
            <head><title>Erro - ELP</title></head>
            <body>
                <h1>Erro Temporário</h1>
                <p>Por favor, <a href="/">volte à página inicial</a> e tente novamente.</p>
            </body>
            </html>
            ''', 500

@app.route('/visits/calendar')
@login_required
def visits_calendar():
    """Calendar view for visits"""
    return render_template('visits/calendar.html')

@app.route('/visits/new', methods=['GET', 'POST'])
@login_required  
def visit_new():
    form = VisitaForm(current_user_id=current_user.id)

    if form.validate_on_submit():
        try:
            from datetime import datetime
            from models import VisitaParticipante

            # Handle project selection - 'Others' option
            final_projeto_id = None
            final_projeto_outros = None

            if form.projeto_id.data == -1:  # 'Others'
                final_projeto_outros = form.projeto_outros.data
            else:
                final_projeto_id = form.projeto_id.data

            # Convert datetime-local strings to datetime objects
            dt_inicio = datetime.fromisoformat(form.data_inicio.data)
            dt_fim = datetime.fromisoformat(form.data_fim.data)

            # Create visit with new structure
            visita = Visita(
                numero=generate_visit_number(),
                projeto_id=final_projeto_id,
                projeto_outros=final_projeto_outros,
                responsavel_id=form.responsavel_id.data,  # Responsável selecionado no formulário
                data_inicio=dt_inicio,
                data_fim=dt_fim,
                observacoes=form.observacoes.data,
                is_pessoal=form.is_pessoal.data,  # Item 31: Compromisso pessoal
                criado_por=current_user.id  # Item 31: Usuário criador (para auditoria)
            )

            db.session.add(visita)
            db.session.flush()  # Get the ID

            # Rastrear participantes adicionados para evitar duplicatas
            participantes_adicionados = set()

            # Add selected participants using form data
            if form.participantes.data:
                current_app.logger.info(f"🔧 Processando {len(form.participantes.data)} participantes")
                for user_id in form.participantes.data:
                    try:
                        # Validar se user_id é válido
                        user_id_int = int(user_id)
                        
                        # Pular se já foi adicionado
                        if user_id_int in participantes_adicionados:
                            current_app.logger.warning(f"⚠️ Participante {user_id_int} já processado, pulando")
                            continue
                        
                        user_exists = User.query.get(user_id_int)

                        if user_exists and user_exists.ativo:
                            # Marcar como confirmado se for o responsável
                            is_responsavel = (user_id_int == visita.responsavel_id)
                            
                            participante = VisitaParticipante(
                                visita_id=visita.id,
                                user_id=user_id_int,
                                confirmado=is_responsavel  # Responsável já confirmado automaticamente
                            )
                            db.session.add(participante)
                            participantes_adicionados.add(user_id_int)
                            current_app.logger.info(f"✅ Participante adicionado: {user_exists.nome_completo} (confirmado={is_responsavel})")
                        else:
                            current_app.logger.error(f"❌ Usuário inválido ou inativo: {user_id}")
                    except (ValueError, TypeError) as e:
                        current_app.logger.error(f"❌ Erro ao processar participante {user_id}: {e}")
                        continue

            # Adicionar responsável como participante se ainda não foi adicionado
            if visita.responsavel_id not in participantes_adicionados:
                responsavel = User.query.get(visita.responsavel_id)
                responsavel_participante = VisitaParticipante(
                    visita_id=visita.id,
                    user_id=visita.responsavel_id,
                    confirmado=True  # Responsável já confirmado automaticamente
                )
                db.session.add(responsavel_participante)
                participantes_adicionados.add(visita.responsavel_id)
                current_app.logger.info(f"✅ Responsável adicionado como participante: {responsavel.nome_completo if responsavel else 'Desconhecido'}")

            # Add default checklist items from templates if available
            try:
                templates = ChecklistTemplate.query.filter_by(ativo=True).order_by(ChecklistTemplate.ordem).all()
                for template in templates:
                    checklist_item = ChecklistItem(
                        visita_id=visita.id,
                        template_id=template.id,
                        pergunta=template.descricao if hasattr(template, 'descricao') else template.nome,
                        obrigatorio=getattr(template, 'obrigatorio', False),
                        ordem=getattr(template, 'ordem', 0)
                    )
                    db.session.add(checklist_item)
            except Exception as template_error:
                current_app.logger.warning(f"⚠️ Erro ao adicionar templates de checklist: {template_error}")

            db.session.commit()
            flash('Visita agendada com sucesso!', 'success')
            return redirect(url_for('visits_list'))

        except Exception as e:
            db.session.rollback()
            current_app.logger.exception(f"❌ Erro ao criar visita: {str(e)}")
            flash(f'Erro ao agendar visita: {str(e)}', 'error')

    # Handle GET request - pre-fill form data from calendar if available
    data_param = request.args.get('data')
    hora_param = request.args.get('hora')

    # Preparar dados iniciais para o template
    form_data = {}
    if data_param:
        try:
            # Processar data do calendário (formato YYYY-MM-DD)
            if hora_param:
                # Combinar data e hora para datetime
                datetime_str = f"{data_param}T{hora_param}"
                form_data['data_inicio'] = datetime_str
                # Definir fim como 1 hora depois
                from datetime import datetime, timedelta
                dt = datetime.fromisoformat(datetime_str)
                dt_fim = dt + timedelta(hours=1)
                form_data['data_fim'] = dt_fim.strftime('%Y-%m-%dT%H:%M')
            else:
                # Apenas data, usar horário padrão (08:00-09:00)
                form_data['data_inicio'] = f"{data_param}T08:00"
                form_data['data_fim'] = f"{data_param}T09:00"
        except Exception as e:
            current_app.logger.error(f"Erro ao processar parâmetros de data: {e}")

    return render_template('visits/form.html', form=form, form_data=form_data)


@app.route('/visits/<int:visit_id>/realize', methods=['GET', 'POST'])
@login_required
def visit_realize(visit_id):
    visit = Visita.query.get_or_404(visit_id)

    if visit.status == 'Realizada':
        flash('Esta visita já foi realizada.', 'info')
        return redirect(url_for('visit_checklist', visit_id=visit_id))

    form = VisitaRealizadaForm()

    if form.validate_on_submit():
        visit.data_realizada = datetime.utcnow()
        visit.status = 'Realizada'
        visit.atividades_realizadas = form.atividades_realizadas.data
        visit.observacoes = form.observacoes.data

        if form.latitude.data and form.longitude.data:
            visit.latitude = float(form.latitude.data)
            visit.longitude = float(form.longitude.data)
            visit.endereco_gps = form.endereco_gps.data

        db.session.commit()
        flash('Visita registrada com sucesso!', 'success')
        return redirect(url_for('visit_checklist', visit_id=visit_id))

    return render_template('visits/form.html', form=form, visit=visit, action='realize')

@app.route('/visits/<int:visit_id>/checklist', methods=['GET', 'POST'])
@login_required
def visit_checklist(visit_id):
    visit = Visita.query.get_or_404(visit_id)
    checklist_items = ChecklistItem.query.filter_by(visita_id=visit_id).order_by(ChecklistItem.ordem).all()

    if request.method == 'POST':
        for item in checklist_items:
            resposta = request.form.get(f'resposta_{item.id}')
            concluido = f'concluido_{item.id}' in request.form

            item.resposta = resposta
            item.concluido = concluido

        db.session.commit()
        flash('Checklist atualizado com sucesso!', 'success')

        # Check if all mandatory items are completed
        mandatory_incomplete = ChecklistItem.query.filter_by(
            visita_id=visit_id,
            obrigatorio=True,
            concluido=False
        ).count()

        if mandatory_incomplete > 0:
            flash(f'Atenção: {mandatory_incomplete} itens obrigatórios ainda não foram concluídos.', 'warning')

    return render_template('visits/checklist.html', visit=visit, checklist_items=checklist_items)

@app.route('/visits/<int:visit_id>')
@login_required
def visit_view(visit_id):
    """View visit details"""
    visit = Visita.query.get_or_404(visit_id)
    checklist_items = ChecklistItem.query.filter_by(visita_id=visit_id).order_by(ChecklistItem.ordem).all()
    comunicacoes = ComunicacaoVisita.query.filter_by(visita_id=visit_id).order_by(ComunicacaoVisita.created_at.desc()).limit(5).all()

    # Buscar participantes da visita
    participantes = []
    if hasattr(visit, 'participantes'):
        for participante in visit.participantes:
            if participante.user:
                participantes.append(participante.user)

    return render_template('visits/checklist.html', visit=visit, checklist_items=checklist_items, comunicacoes=comunicacoes, participantes=participantes)

@app.route('/visits/<int:visit_id>/cancel', methods=['POST'])
@login_required
def visit_cancel(visit_id):
    """Cancelar uma visita"""
    try:
        visit = Visita.query.get_or_404(visit_id)

        # Verificar permissões
        if not (current_user.is_master or visit.responsavel_id == current_user.id or visit.criado_por == current_user.id):
            flash('Acesso negado para cancelar esta visita.', 'error')
            return redirect(url_for('visits_list'))

        # Não permitir cancelar visitas já realizadas
        if visit.status == 'Realizada':
            flash('Não é possível cancelar uma visita já realizada.', 'error')
            return redirect(url_for('visit_view', visit_id=visit_id))

        visit.status = 'Cancelado'
        db.session.commit()

        flash('Visita cancelada com sucesso!', 'success')
        return redirect(url_for('visits_list'))

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao cancelar visita: {str(e)}', 'error')
        return redirect(url_for('visit_view', visit_id=visit_id))

@app.route('/visits/<int:visit_id>/edit', methods=['GET', 'POST'])
@login_required
def visit_edit(visit_id):
    """Alterar data e hora de uma visita"""
    try:
        from models import VisitaParticipante
        visit = Visita.query.get_or_404(visit_id)

        # Verificar permissões
        if not (current_user.is_master or visit.responsavel_id == current_user.id or visit.criado_por == current_user.id):
            flash('Acesso negado para alterar esta visita.', 'error')
            return redirect(url_for('visits_list'))

        # Não permitir alterar visitas já realizadas
        if visit.status == 'Realizada':
            flash('Não é possível alterar uma visita já realizada.', 'error')
            return redirect(url_for('visit_view', visit_id=visit_id))

        form = VisitaForm(visit=visit, current_user_id=current_user.id)

        if request.method == 'GET':
            try:
                # Preencher formulário com dados atuais - convert to datetime-local format
                form.data_inicio.data = visit.data_inicio.strftime('%Y-%m-%dT%H:%M') if visit.data_inicio else ''
                form.data_fim.data = visit.data_fim.strftime('%Y-%m-%dT%H:%M') if visit.data_fim else ''
                form.observacoes.data = visit.observacoes or ''
                form.responsavel_id.data = visit.responsavel_id  # Preencher responsável atual
                form.is_pessoal.data = visit.is_pessoal or False  # Preencher flag pessoal

                # Preencher projeto
                if visit.projeto_id:
                    form.projeto_id.data = visit.projeto_id
                else:
                    form.projeto_id.data = -1  # 'Outros'
                    form.projeto_outros.data = visit.projeto_outros or ''

                # Preencher participantes com tratamento de erro robusto
                try:
                    participantes_existentes = VisitaParticipante.query.filter_by(visita_id=visit_id).all()
                    participante_ids = []
                    for p in participantes_existentes:
                        if p.user_id:
                            participante_ids.append(str(p.user_id))
                    form.participantes.data = participante_ids
                    current_app.logger.info(f"✅ Participantes carregados: {len(participante_ids)}")
                except Exception as part_error:
                    current_app.logger.error(f"❌ Erro ao carregar participantes: {part_error}")
                    form.participantes.data = []

            except Exception as form_error:
                current_app.logger.error(f"❌ Erro ao preencher formulário: {form_error}")
                flash('Erro ao carregar dados da visita.', 'error')
                return redirect(url_for('visits_list'))

        if form.validate_on_submit():
            try:
                # Convert datetime-local strings to datetime objects
                dt_inicio = datetime.fromisoformat(form.data_inicio.data)
                dt_fim = datetime.fromisoformat(form.data_fim.data)

                # Atualizar campos
                visit.data_inicio = dt_inicio
                visit.data_fim = dt_fim
                visit.observacoes = form.observacoes.data or ''
                visit.responsavel_id = form.responsavel_id.data  # Atualizar responsável
                visit.is_pessoal = form.is_pessoal.data or False  # Atualizar flag pessoal

                # Atualizar projeto
                if form.projeto_id.data == -1:  # 'Outros'
                    visit.projeto_id = None
                    visit.projeto_outros = form.projeto_outros.data or ''
                else:
                    visit.projeto_id = form.projeto_id.data
                    visit.projeto_outros = None

                # Atualizar participantes com tratamento de erro robusto
                try:
                    from models import VisitaParticipante

                    # Primeiro, remover participantes existentes
                    VisitaParticipante.query.filter_by(visita_id=visit_id).delete()
                    current_app.logger.info(f"🗑️ Participantes existentes removidos da visita {visit_id}")

                    # Rastrear participantes adicionados para evitar duplicatas
                    participantes_adicionados = set()

                    # Adicionar novos participantes selecionados
                    if form.participantes.data:
                        current_app.logger.info(f"🔧 Processando {len(form.participantes.data)} participantes para edição")

                        for user_id in form.participantes.data:
                            try:
                                user_id_int = int(user_id)

                                # Pular se já foi adicionado
                                if user_id_int in participantes_adicionados:
                                    current_app.logger.warning(f"⚠️ Participante {user_id_int} já processado, pulando")
                                    continue

                                # Verificar se usuário existe e está ativo
                                user_exists = User.query.get(user_id_int)
                                if user_exists and user_exists.ativo:
                                    # Marcar como confirmado se for o responsável
                                    is_responsavel = (user_id_int == visit.responsavel_id)
                                    
                                    participante = VisitaParticipante(
                                        visita_id=visit_id,
                                        user_id=user_id_int,
                                        confirmado=is_responsavel
                                    )
                                    db.session.add(participante)
                                    participantes_adicionados.add(user_id_int)
                                    current_app.logger.info(f"✅ Participante readicionado: {user_exists.nome_completo} (confirmado={is_responsavel})")
                                else:
                                    current_app.logger.warning(f"⚠️ Usuário inválido ou inativo ignorado: {user_id}")

                            except (ValueError, TypeError) as e:
                                current_app.logger.warning(f"⚠️ ID de usuário inválido ignorado: {user_id} - {e}")
                                continue

                    # Adicionar responsável como participante se ainda não foi adicionado
                    if visit.responsavel_id not in participantes_adicionados:
                        try:
                            responsavel_participante = VisitaParticipante(
                                visita_id=visit_id,
                                user_id=visit.responsavel_id,
                                confirmado=True  # Responsável já confirmado automaticamente
                            )
                            db.session.add(responsavel_participante)
                            participantes_adicionados.add(visit.responsavel_id)
                            current_app.logger.info(f"✅ Responsável readicionado como participante")
                        except Exception as resp_error:
                            current_app.logger.error(f"❌ Erro ao readicionar responsável: {resp_error}")

                except Exception as part_error:
                    current_app.logger.error(f"❌ Erro crítico ao processar participantes: {part_error}")
                    # Manter pelo menos o responsável como participante em caso de erro
                    try:
                        VisitaParticipante.query.filter_by(visita_id=visit_id).delete()
                        responsavel_participante = VisitaParticipante(
                            visita_id=visit_id,
                            user_id=visit.responsavel_id,
                            confirmado=True
                        )
                        db.session.add(responsavel_participante)
                        current_app.logger.info("🔄 Fallback: Apenas responsável mantido como participante")
                    except Exception as fallback_error:
                        current_app.logger.error(f"❌ Erro no fallback de participantes: {fallback_error}")

                db.session.commit()
                current_app.logger.info(f"✅ Visita {visit_id} alterada com sucesso")
                flash('Visita alterada com sucesso!', 'success')
                return redirect(url_for('visit_view', visit_id=visit_id))

            except Exception as e:
                db.session.rollback()
                current_app.logger.error(f"❌ Erro ao alterar visita {visit_id}: {str(e)}")
                flash(f'Erro ao alterar visita: {str(e)}', 'error')

        return render_template('visits/form.html', form=form, visit=visit, action='edit')

    except Exception as e:
        current_app.logger.exception(f"❌ ERRO CRÍTICO na edição da visita {visit_id}: {str(e)}")
        flash('Erro interno ao carregar a visita para edição.', 'error')
        return redirect(url_for('visits_list'))

# Report management routes - movido para routes_reports.py

# Rota removida - usando nova implementação em routes_reports.py

@app.route('/reports/<int:report_id>')
@login_required
def view_report(report_id):
    """Visualizar relatório - versão robusta conforme especificação do documento"""
    try:
        # Logging detalhado conforme especificação
        current_app.logger.info(f"📖 /reports/{report_id}: Tentativa de acesso por usuário {current_user.id}")

        report = Relatorio.query.get_or_404(report_id)
        current_app.logger.info(f"✅ Relatório {report_id} encontrado: Status={report.status}")

        # Check basic permissions
        user_can_view = False
        user_can_edit = False

        if current_user.is_master:
            user_can_view = True
            user_can_edit = report.status not in ['Aprovado', 'Finalizado']
        elif report.autor_id == current_user.id:
            user_can_view = True
            user_can_edit = report.status in ['Rascunho', 'preenchimento', 'Rejeitado', 'Em edição', 'Aguardando Aprovação']
        else:
            # Allow project team members to view
            if hasattr(report, 'projeto') and report.projeto:
                try:
                    user_has_access = FuncionarioProjeto.query.filter_by(
                        projeto_id=report.projeto.id,
                        user_id=current_user.id,
                        ativo=True
                    ).first()

                    if user_has_access or (hasattr(report.projeto, 'responsavel_id') and report.projeto.responsavel_id == current_user.id):
                        user_can_view = True
                        user_can_edit = False  # Membros da equipe só visualizam
                except Exception as e:
                    current_app.logger.warning(f"Erro ao verificar acesso ao projeto para relatório {report_id}: {str(e)}")
                    pass

        if not user_can_view:
            flash('Acesso negado ao relatório.', 'error')
            return redirect(url_for('reports'))

        # Desserializar checklist com try/except conforme especificação
        try:
            import json
            checklist = {}
            if hasattr(report, 'checklist_data') and report.checklist_data:
                try:
                    checklist = json.loads(report.checklist_data)
                    if not isinstance(checklist, dict):
                        checklist = {}
                except (json.JSONDecodeError, TypeError, AttributeError) as json_error:
                    current_app.logger.error(f"❌ JSON inválido no checklist do relatório {report_id}: {str(json_error)}")
                    checklist = {}
            current_app.logger.info(f"✅ Checklist carregado: {len(checklist)} itens")
        except Exception as e:
            current_app.logger.exception(f"ERRO GERAL CHECKLIST relatório {report_id}: {str(e)}")
            checklist = {}

        try:
            fotos = FotoRelatorio.query.filter_by(relatorio_id=report_id).order_by(FotoRelatorio.ordem).all()
            fotos = fotos or []  # Garantir que não seja None
            current_app.logger.info(f"✅ Fotos carregadas: {len(fotos)} arquivos")
        except Exception as e:
            current_app.logger.error(f"❌ Erro ao buscar fotos do relatório {report_id}: {str(e)}")
            fotos = []

        return render_template('reports/view.html', 
                             report=report,  # Padronizado para 'report' conforme especificação
                             fotos=fotos,
                             checklist=checklist,
                             user_can_edit=user_can_edit,
                             user_can_view=user_can_view)

    except Exception as e:
        current_app.logger.exception(f"ERRO GERAL em /reports/{report_id}: {str(e)}")
        abort(500, description="Erro interno ao carregar relatório para edição.")

@app.route('/reports/<int:report_id>/view')
@login_required  
def report_view(report_id):
    """Visualizar relatório - rota compatível com redirecionamentos existentes"""
    return view_report(report_id)

@app.route('/reports/<int:report_id>/edit', methods=['GET', 'POST'])
@login_required
def report_edit(report_id):
    """Editar relatório - versão corrigida para relatórios rejeitados"""
    try:
        current_app.logger.info(f"✏️ report_edit chamado para report_id={report_id}")

        # Buscar relatório com tratamento de erro
        try:
            relatorio = Relatorio.query.get_or_404(report_id)
        except Exception as e:
            current_app.logger.error(f"❌ Erro ao buscar relatório {report_id}: {str(e)}")
            flash('Relatório não encontrado.', 'error')
            return redirect(url_for('reports'))

        # Verificar permissões básicas - CORRIGIDO PARA RELATÓRIOS REJEITADOS
        user_can_edit = False

        # Master pode editar tudo exceto aprovados
        if current_user.is_master:
            user_can_edit = relatorio.status not in ['Aprovado', 'Finalizado']

        # Autor pode editar se não aprovado - INCLUINDO REJEITADOS
        elif relatorio.autor_id == current_user.id:
            user_can_edit = relatorio.status in ['Rascunho', 'preenchimento', 'Rejeitado', 'Em edição', 'Aguardando Aprovação']

        if not user_can_edit:
            flash('Você não tem permissão para editar este relatório ou ele já foi finalizado.', 'error')
            return redirect(url_for('view_report', report_id=report_id))

        # Para relatórios rejeitados, mudar status para "Em edição" automaticamente
        if relatorio.status == 'Rejeitado' and request.method == 'GET':
            try:
                relatorio.status = 'Em edição'
                relatorio.updated_at = datetime.utcnow()
                db.session.commit()
                current_app.logger.info(f"📝 Status alterado de 'Rejeitado' para 'Em edição' para relatório {report_id}")
            except Exception as e:
                current_app.logger.error(f"❌ Erro ao alterar status: {str(e)}")
                db.session.rollback()

        # Buscar dados auxiliares com tratamento de erro
        projetos = []
        fotos = []

        try:
            projetos = Projeto.query.filter_by(status='Ativo').all()
            current_app.logger.info(f"📋 Projetos carregados: {len(projetos)}")
        except Exception as e:
            current_app.logger.error(f"⚠️ Erro ao buscar projetos: {str(e)}")
            projetos = []

        try:
            fotos = FotoRelatorio.query.filter_by(relatorio_id=relatorio.id).order_by(FotoRelatorio.ordem).all()
            current_app.logger.info(f"📸 Fotos carregadas: {len(fotos)}")
        except Exception as e:
            current_app.logger.error(f"⚠️ Erro ao buscar fotos: {str(e)}")
            fotos = []

        # Processamento de formulário POST
        if request.method == 'POST':
            try:
                action = request.form.get('action', 'update')

                if action == 'update':
                    # Atualizar campos básicos de forma segura
                    if 'titulo' in request.form:
                        relatorio.titulo = request.form.get('titulo', '').strip()

                    if 'conteudo' in request.form:
                        relatorio.conteudo = request.form.get('conteudo', '').strip()

                    if 'projeto_id' in request.form:
                        projeto_id = request.form.get('projeto_id')
                        if projeto_id and projeto_id.isdigit():
                            relatorio.projeto_id = int(projeto_id)

                    if 'observacoes' in request.form:
                        relatorio.observacoes = request.form.get('observacoes', '').strip()

                    # Process acompanhantes (visit attendees) - same logic as create_report
                    acompanhantes_data = request.form.get('acompanhantes')
                    if acompanhantes_data:
                        try:
                            import json
                            acompanhantes_list = json.loads(acompanhantes_data)
                            if isinstance(acompanhantes_list, list):
                                relatorio.acompanhantes = acompanhantes_list
                                current_app.logger.info(f"✅ Acompanhantes atualizados: {len(acompanhantes_list)} registros")
                        except Exception as e:
                            current_app.logger.error(f"❌ Erro ao processar acompanhantes: {e}")
                            # Keep existing acompanhantes if parsing fails

                    # Para relatórios em edição (que eram rejeitados), manter status
                    if relatorio.status == 'Em edição':
                        # Manter status Em edição até que seja enviado para aprovação novamente
                        pass

                    # Atualizar timestamp
                    relatorio.updated_at = datetime.utcnow()

                    db.session.commit()
                    flash('Relatório atualizado com sucesso!', 'success')

                elif action == 'submit_approval':
                    # Permitir envio para aprovação - INCLUINDO RELATÓRIOS EM EDIÇÃO
                    status_permitidos = ['preenchimento', 'Rascunho', 'Rejeitado', 'Em edição', 'Aguardando Aprovação']
                    if relatorio.status in status_permitidos:
                        relatorio.status = 'Aguardando Aprovação'
                        relatorio.updated_at = datetime.utcnow()
                        # Limpar comentário de reprovação anterior
                        relatorio.comentario_aprovacao = None
                        db.session.commit()
                        flash('Relatório reenviado para aprovação!', 'success')
                    else:
                        flash('Relatório não pode ser enviado para aprovação no status atual.', 'warning')

                return redirect(url_for('view_report', report_id=report_id))

            except Exception as e:
                db.session.rollback()
                current_app.logger.error(f"❌ Erro ao atualizar relatório {report_id}: {str(e)}")
                flash('Erro ao atualizar relatório. Tente novamente.', 'error')

        # Carregar checklist de forma segura
        checklist = {}
        try:
            if hasattr(relatorio, 'checklist_data') and relatorio.checklist_data:
                import json
                checklist = json.loads(relatorio.checklist_data)
        except Exception as e:
            current_app.logger.warning(f"⚠️ Checklist inválido no relatório {report_id}: {str(e)}")
            checklist = {}

        # Garantir valores padrão para evitar erros no template
        if not hasattr(relatorio, 'observacoes') or relatorio.observacoes is None:
            relatorio.observacoes = ''
        if not hasattr(relatorio, 'conteudo') or relatorio.conteudo is None:
            relatorio.conteudo = ''
        if not hasattr(relatorio, 'titulo') or relatorio.titulo is None:
            relatorio.titulo = 'Relatório de visita'

        # Garantir que o relacionamento com projeto existe
        try:
            projeto = relatorio.projeto
            if not projeto:
                current_app.logger.warning(f"⚠️ Projeto não encontrado para relatório {report_id}")
        except Exception as e:
            current_app.logger.warning(f"⚠️ Erro ao carregar projeto do relatório {report_id}: {str(e)}")

        # Garantir que o relacionamento com autor existe
        try:
            autor = relatorio.autor
            if not autor:
                current_app.logger.warning(f"⚠️ Autor não encontrado para relatório {report_id}")
        except Exception as e:
            current_app.logger.warning(f"⚠️ Erro ao carregar autor do relatório {report_id}: {str(e)}")

        current_app.logger.info(f"📖 Usuário {current_user.username} editando relatório {relatorio.numero} (status: {relatorio.status})")

        # Renderizar template com todas as variáveis necessárias
        return render_template('reports/edit.html', 
                             relatorio=relatorio, 
                             projetos=projetos, 
                             fotos=fotos,
                             checklist=checklist,
                             is_readonly=False,
                             user_can_edit=True,
                             user_can_view=True)

    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        current_app.logger.exception(f"❌ ERRO CRÍTICO na edição do relatório {report_id}: {str(e)}")
        current_app.logger.error(f"❌ TRACEBACK: {error_trace}")
        flash('Erro interno ao carregar relatório para edição.', 'error')
        return redirect(url_for('reports'))

@app.route('/reports/<int:report_id>/editarrel', methods=['GET'])
@login_required
def report_edit_complete(report_id):
    """
    Rota corrigida - Carrega relatório completo com todos os dados para edição
    Garante serialização JSON 100% sem objetos ORM
    """
    try:
        app.logger.info(f"📝 Iniciando carregamento do relatório completo ID={report_id}")

        relatorio = (
            db.session.query(Relatorio)
            .options(joinedload(Relatorio.projeto))
            .filter(Relatorio.id == report_id)
            .first()
        )

        if not relatorio:
            return render_template("reports/error_report.html", message="Relatório não encontrado"), 404

        # --- FUNÇÃO DE SERIALIZAÇÃO SEGURA ---
        def safe_attr(obj, attr, default=""):
            """Extrai atributo de forma segura, retornando string vazia se não existir"""
            value = getattr(obj, attr, default) if obj else default
            # Se o valor for um objeto complexo (User, etc), converte para string
            if hasattr(value, '__tablename__'):  # É um objeto ORM
                return str(value) if value else default
            return value if value is not None else default

        projeto = Projeto.query.get(relatorio.projeto_id)
        fotos = FotoRelatorio.query.filter_by(relatorio_id=report_id).all()
        
        # --- ACOMPANHANTES: CARREGAR APENAS OS SELECIONADOS DO RELATÓRIO ---
        acompanhantes_selecionados = []
        todos_acompanhantes_projeto = []
        
        if relatorio.acompanhantes:
            import json
            try:
                acomp_raw = relatorio.acompanhantes
                if isinstance(acomp_raw, str):
                    acomp_raw = json.loads(acomp_raw)
                
                if isinstance(acomp_raw, list):
                    for item in acomp_raw:
                        if isinstance(item, dict):
                            acompanhantes_selecionados.append({
                                "id": item.get("id"),
                                "nome": item.get("nome") or item.get("nome_funcionario", ""),
                                "funcao": item.get("funcao") or item.get("cargo", "")
                            })
                
                app.logger.info(f"✅ {len(acompanhantes_selecionados)} acompanhantes selecionados carregados")
            except Exception as e:
                app.logger.warning(f"⚠️ Erro ao parsear acompanhantes: {e}")
                acompanhantes_selecionados = []
        
        # Carregar todos os acompanhantes do projeto para permitir adição
        if relatorio.projeto_id:
            todos_funcionarios = FuncionarioProjeto.query.filter_by(projeto_id=relatorio.projeto_id).all()
            todos_acompanhantes_projeto = [
                {
                    "id": a.id,
                    "nome": safe_attr(a, "nome_funcionario") or safe_attr(a, "nome"),
                    "funcao": safe_attr(a, "funcao") or safe_attr(a, "cargo") or "Não informado"
                }
                for a in todos_funcionarios
            ]

        # --- CHECKLIST: Lidar com lista OU dicionário + CARREGAR DADOS COMPLETOS ---
        checklist = []
        if relatorio.checklist_data:
            import json
            try:
                checklist_raw = relatorio.checklist_data
                
                # Se for string JSON, parsear
                if isinstance(checklist_raw, str):
                    checklist_raw = json.loads(checklist_raw)
                
                # Se for lista (array de objetos)
                if isinstance(checklist_raw, list):
                    for item in checklist_raw:
                        if isinstance(item, dict):
                            # Buscar item do banco se tiver ID para pegar imagem e dados completos
                            item_completo = None
                            if item.get("id"):
                                try:
                                    from models import ChecklistObra
                                    item_completo = ChecklistObra.query.get(item.get("id"))
                                except:
                                    pass
                            
                            checklist.append({
                                "id": item.get("id"),
                                "texto": item.get("descricao") or item.get("texto", ""),
                                "checked": bool(item.get("concluido") or item.get("completado", False)),
                                "imagem_url": safe_attr(item_completo, "imagem_url") if item_completo else item.get("imagem_url", ""),
                                "observacao": item.get("observacao", "")
                            })
                        else:
                            # Item é string simples
                            checklist.append({
                                "id": None,
                                "texto": str(item),
                                "checked": False,
                                "imagem_url": "",
                                "observacao": ""
                            })
                
                # Se for dicionário (chave: valor)
                elif isinstance(checklist_raw, dict):
                    for descricao, concluido in checklist_raw.items():
                        checklist.append({
                            "id": None,
                            "texto": str(descricao),
                            "checked": bool(concluido),
                            "imagem_url": "",
                            "observacao": ""
                        })
                
                app.logger.info(f"✅ Checklist parseado: {len(checklist)} itens")
                
            except Exception as e:
                app.logger.warning(f"⚠️ Erro ao parsear checklist: {e}")
                checklist = []

        # --- CONVERTER OBJETOS ORM EM DICIONÁRIOS PLANOS ---
        report_data = {
            "relatorio": {
                "id": relatorio.id,
                "data_relatorio": relatorio.data_relatorio.strftime("%Y-%m-%d") if relatorio.data_relatorio else "",
                "titulo": safe_attr(relatorio, "titulo"),
                "numero": safe_attr(relatorio, "numero") or safe_attr(relatorio, "numero_relatorio"),
                "observacoes_finais": safe_attr(relatorio, "observacoes_finais") or safe_attr(relatorio, "observacoes") or safe_attr(relatorio, "conteudo"),
                "lembrete_proxima_visita": safe_attr(relatorio, "lembrete_proxima_visita") or safe_attr(relatorio, "lembrete"),
                "autor": safe_attr(safe_attr(relatorio, "autor"), "nome") if hasattr(relatorio, "autor") and relatorio.autor else "",
            },
            "projeto": {
                "id": safe_attr(projeto, "id"),
                "codigo": safe_attr(projeto, "codigo") or safe_attr(projeto, "numero"),
                "nome": safe_attr(projeto, "nome"),
                "endereco": safe_attr(projeto, "endereco"),
                "responsavel": safe_attr(safe_attr(projeto, "responsavel"), "nome") if hasattr(projeto, "responsavel") and projeto.responsavel else ""
            } if projeto else None,
            "fotos": [
                {
                    "id": f.id,
                    "url": url_for('api_get_photo', foto_id=f.id),
                    "filename": safe_attr(f, "filename"),
                    "legenda": safe_attr(f, "legenda") or safe_attr(f, "titulo"),
                    "categoria": safe_attr(f, "tipo_servico") or safe_attr(f, "categoria"),
                    "tipo_servico": safe_attr(f, "tipo_servico"),
                    "local": safe_attr(f, "local"),
                    "ordem": f.ordem if hasattr(f, "ordem") else 0,
                    "savedId": f.id
                } for f in fotos
            ],
            "checklist": checklist,
            "acompanhantes": acompanhantes_selecionados,
            "todos_acompanhantes": todos_acompanhantes_projeto
        }

        app.logger.info(f"✅ Dados prontos para template: {len(fotos)} fotos, {len(checklist)} checklist, {len(acompanhantes_selecionados)} acompanhantes selecionados")

        return render_template(
            "reports/form_complete.html",
            report_data=report_data,
            edit_mode=True
        )

    except Exception as e:
        app.logger.error(f"❌ Erro ao carregar relatório {report_id}: {e}", exc_info=True)
        return render_template("reports/error_report.html", message=str(e)), 500

@app.route('/api/reports/<int:report_id>/update', methods=['POST'])
@login_required
def update_report(report_id):
    """
    Rota de atualização de relatório em modo de edição.
    Atualiza o relatório existente sem criar um novo.
    Retorna apenas JSON (nunca HTML).
    """
    try:
        relatorio = db.session.get(Relatorio, report_id)
        if not relatorio:
            return jsonify({"success": False, "message": "Relatório não encontrado"}), 404

        app.logger.info(f"📝 Atualizando relatório ID={report_id}")

        # Processar dados do formulário
        data = request.form if request.form else {}
        
        # Atualizar campos básicos do relatório
        if "titulo" in data:
            relatorio.titulo = data.get("titulo", relatorio.titulo)
        if "observacoes_finais" in data:
            relatorio.observacoes_finais = data.get("observacoes_finais", relatorio.observacoes_finais)
        if "data_relatorio" in data:
            data_str = data.get("data_relatorio")
            if data_str:
                try:
                    relatorio.data_relatorio = datetime.strptime(data_str, '%Y-%m-%d')
                except ValueError:
                    app.logger.warning(f"⚠️ Formato de data inválido: {data_str}")
        if "lembrete_proxima_visita" in data:
            relatorio.lembrete_proxima_visita = data.get("lembrete_proxima_visita")

        # Atualizar acompanhantes
        if "acompanhantes" in data:
            import json
            try:
                acompanhantes_data = data.get("acompanhantes")
                app.logger.info(f"📥 Acompanhantes recebidos (raw): {acompanhantes_data}")
                
                if isinstance(acompanhantes_data, str):
                    acompanhantes_data = json.loads(acompanhantes_data)
                
                app.logger.info(f"👥 Acompanhantes parseados: {acompanhantes_data}")
                
                # CORREÇÃO: Salvar diretamente a lista (SQLAlchemy já serializa JSONB automaticamente)
                relatorio.acompanhantes = acompanhantes_data
                app.logger.info(f"✅ Acompanhantes atualizados: {len(acompanhantes_data) if isinstance(acompanhantes_data, list) else 0}")
            except Exception as e:
                app.logger.error(f"❌ Erro ao atualizar acompanhantes: {e}")
                import traceback
                traceback.print_exc()

        # Atualizar checklist
        if "checklist" in data:
            import json
            try:
                checklist_data = data.get("checklist")
                if isinstance(checklist_data, str):
                    checklist_data = json.loads(checklist_data)
                relatorio.checklist_data = json.dumps(checklist_data)
                app.logger.info(f"✅ Checklist atualizado")
            except Exception as e:
                app.logger.error(f"❌ Erro ao atualizar checklist: {e}")

        # Processar imagens
        # Manter apenas as imagens listadas em imagens_existentes
        imagens_existentes = request.form.getlist("imagens_existentes[]")
        if imagens_existentes:
            # Converter para lista de IDs inteiros
            ids_existentes = []
            for img_id in imagens_existentes:
                try:
                    ids_existentes.append(int(img_id))
                except ValueError:
                    pass
            
            # Remover imagens que não estão na lista
            if ids_existentes:
                FotoRelatorio.query.filter(
                    FotoRelatorio.relatorio_id == report_id,
                    ~FotoRelatorio.id.in_(ids_existentes)
                ).delete(synchronize_session=False)
                app.logger.info(f"✅ Mantidas {len(ids_existentes)} imagens existentes")
        
        # CORREÇÃO: Atualizar legendas, categorias e locais das imagens existentes
        import json
        legendas_imagens = request.form.get("legendas_imagens")
        if legendas_imagens:
            try:
                legendas_data = json.loads(legendas_imagens)
                app.logger.info(f"📝 Atualizando legendas de {len(legendas_data)} imagens")
                
                for img_data in legendas_data:
                    foto_id = img_data.get("id")
                    if foto_id:
                        foto = FotoRelatorio.query.filter_by(
                            id=foto_id, 
                            relatorio_id=report_id
                        ).first()
                        
                        if foto:
                            # Atualizar legenda
                            if "legenda" in img_data:
                                foto.legenda = img_data["legenda"]
                                app.logger.info(f"✅ Legenda atualizada para foto {foto_id}: {img_data['legenda'][:50]}...")
                            
                            # Atualizar categoria
                            if "categoria" in img_data:
                                foto.categoria = img_data["categoria"]
                            
                            # Atualizar local
                            if "local" in img_data:
                                foto.local = img_data["local"]
                                
                app.logger.info(f"✅ Legendas atualizadas com sucesso")
            except Exception as e:
                app.logger.error(f"❌ Erro ao atualizar legendas: {e}")
                import traceback
                traceback.print_exc()

        # Adicionar novas imagens (COM VERIFICAÇÃO DE DUPLICAÇÃO)
        novas_imagens = request.files.getlist("imagens")
        app.logger.info(f"📥 Novas imagens recebidas: {len(novas_imagens)}")
        
        if novas_imagens:
            ordem_atual = FotoRelatorio.query.filter_by(relatorio_id=report_id).count()
            app.logger.info(f"📊 Ordem atual das fotos: {ordem_atual}")
            
            for index, arquivo in enumerate(novas_imagens):
                if arquivo and arquivo.filename:
                    try:
                        app.logger.info(f"📤 Processando imagem {index + 1}/{len(novas_imagens)}: {arquivo.filename}")
                        
                        # Preparar nome do arquivo
                        nome_arquivo_original = secure_filename(arquivo.filename)
                        
                        # Ler dados do arquivo primeiro
                        file_data = arquivo.read()
                        file_size = len(file_data)
                        app.logger.info(f"📦 Tamanho do arquivo: {file_size} bytes")
                        
                        # 🔧 CORREÇÃO CRÍTICA: Verificar duplicação por HASH SHA-256 (mais seguro que filename)
                        import hashlib
                        imagem_hash = hashlib.sha256(file_data).hexdigest()
                        
                        foto_existente = FotoRelatorio.query.filter_by(
                            relatorio_id=report_id,
                            imagem_hash=imagem_hash
                        ).first()
                        
                        if foto_existente:
                            # Imagem já existe (por hash) - apenas atualizar metadados se necessário
                            app.logger.info(f"🔄 Imagem já existe (hash={imagem_hash[:12]}...) - ID: {foto_existente.id}. Atualizando apenas metadados.")
                            
                            # Buscar metadados do form se disponíveis
                            legenda = request.form.get(f"legenda_{index}")
                            categoria = request.form.get(f"categoria_{index}")
                            local = request.form.get(f"local_{index}")
                            
                            # Atualizar apenas metadados (sem duplicar arquivo)
                            if legenda is not None:
                                foto_existente.legenda = legenda
                                app.logger.info(f"✏️ Legenda atualizada: {legenda[:50]}...")
                            if categoria is not None:
                                foto_existente.categoria = categoria
                                foto_existente.tipo_servico = categoria
                                app.logger.info(f"✏️ Categoria atualizada: {categoria}")
                            if local is not None:
                                foto_existente.local = local
                                app.logger.info(f"✏️ Local atualizado: {local}")
                            
                            db.session.add(foto_existente)
                            app.logger.info(f"🖼️ Imagem existente atualizada (sem duplicação por hash)")
                        else:
                            # Nova imagem - processar normalmente
                            unique_filename = f"{uuid.uuid4()}_{nome_arquivo_original}"
                            caminho = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
                            
                            arquivo.seek(0)
                            arquivo.save(caminho)
                            app.logger.info(f"💾 Arquivo salvo em: {caminho}")
                            
                            # Buscar metadados do form se disponíveis
                            legenda = request.form.get(f"legenda_{index}", "")
                            categoria = request.form.get(f"categoria_{index}", "")
                            local = request.form.get(f"local_{index}", "")
                            
                            # Criar registro da foto
                            nova_foto = FotoRelatorio()
                            nova_foto.relatorio_id = report_id
                            nova_foto.filename = unique_filename
                            nova_foto.filename_original = nome_arquivo_original
                            nova_foto.imagem = file_data
                            nova_foto.imagem_hash = imagem_hash
                            nova_foto.imagem_size = file_size
                            nova_foto.legenda = legenda
                            nova_foto.categoria = categoria
                            nova_foto.tipo_servico = categoria
                            nova_foto.local = local
                            nova_foto.ordem = ordem_atual + 1
                            ordem_atual += 1
                            
                            db.session.add(nova_foto)
                            app.logger.info(f"🆕 Nova imagem adicionada: {unique_filename} (ordem: {nova_foto.ordem}, hash: {imagem_hash[:12]}...)")
                    except Exception as e:
                        app.logger.error(f"❌ Erro ao processar imagem {arquivo.filename}: {e}")
                        import traceback
                        traceback.print_exc()
                else:
                    app.logger.warning(f"⚠️ Arquivo vazio ou sem nome recebido no índice {index}")

        # SEMPRE mantém status como "preenchimento" (comportamento igual ao autosave)
        should_finalize = request.form.get('should_finalize') == 'true'
        if should_finalize:
            app.logger.info(f"✅ Relatório {relatorio.numero} salvo em preenchimento (should_finalize ignorado)")

        # Salvar alterações no banco
        db.session.commit()
        app.logger.info(f"✅ Relatório {report_id} atualizado com sucesso")
        
        # Criar notificação se relatório foi editado e está aguardando aprovação
        if not should_finalize:  # Não criar notificação se acabou de finalizar (já foi criada acima)
            try:
                from notification_service import notification_service
                notification_service.criar_notificacao_relatorio_editado(relatorio.id, current_user.id)
                app.logger.info(f"✅ Notificação de edição criada (se aplicável)")
            except Exception as notif_error:
                app.logger.error(f"⚠️ Erro ao criar notificação de edição: {notif_error}")
        
        return jsonify({
            "success": True,
            "message": "Relatório atualizado com sucesso",
            "relatorio_id": relatorio.id
        })

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"❌ Erro ao atualizar relatório {report_id}: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/reports/<int:report_id>/photos/add', methods=['GET', 'POST'])
@login_required
def report_add_photo(report_id):
    report = Relatorio.query.get_or_404(report_id)
    form = FotoRelatorioForm()

    if form.validate_on_submit():
        foto = form.foto.data
        if foto:
            filename = secure_filename(foto.filename)
            unique_filename = f"{uuid.uuid4()}_{filename}"
            foto_path = os.path.join(current_app.config['UPLOAD_FOLDER'], unique_filename)

            # Read file data before saving
            file_data = foto.read()
            foto.seek(0)  # Reset for saving to disk
            foto.save(foto_path)

            # Create photo record with binary data
            foto_relatorio = FotoRelatorio()
            foto_relatorio.relatorio_id = report_id
            foto_relatorio.filename = unique_filename
            foto_relatorio.titulo = form.titulo.data if hasattr(form, 'titulo') else ""
            foto_relatorio.descricao = form.descricao.data if hasattr(form, 'descricao') else ""
            foto_relatorio.tipo_servico = form.tipo_servico.data if hasattr(form, 'tipo_servico') else "Geral"
            foto_relatorio.ordem = FotoRelatorio.query.filter_by(relatorio_id=report_id).count() + 1
            foto_relatorio.imagem = file_data  # Salvar dados binários da imagem

            db.session.add(foto_relatorio)
            db.session.commit()
            flash('Foto adicionada com sucesso!', 'success')
            return redirect(url_for('report_view', report_id=report_id))

    return render_template('reports/form.html', form=form, report=report, action='add_photo')

@app.route('/reports/<int:report_id>/send', methods=['POST'])
@login_required
def report_send(report_id):
    report = Relatorio.query.get_or_404(report_id)

    if report.status != 'Finalizado':
        flash('Apenas relatórios finalizados podem ser enviados.', 'error')
        return redirect(url_for('report_view', report_id=report_id))

    # Get contacts who should receive reports
    contatos_projeto = ContatoProjeto.query.filter_by(
        projeto_id=report.projeto_id,
        receber_relatorios=True
    ).all()

    if not contatos_projeto:
        flash('Nenhum contato configurado para receber relatórios nesta obra.', 'error')
        return redirect(url_for('report_view', report_id=report_id))

    emails_enviados = 0
    for contato_projeto in contatos_projeto:
        if contato_projeto.contato.email:
            try:
                send_report_email(report, contato_projeto.contato.email, contato_projeto.contato.nome)

                # Log the email sending
                envio = EnvioRelatorio()
                envio.relatorio_id = report_id
                envio.email_destinatario = contato_projeto.contato.email
                envio.nome_destinatario = contato_projeto.contato.nome
                db.session.add(envio)
                emails_enviados += 1
            except Exception as e:
                flash(f'Erro ao enviar email para {contato_projeto.contato.email}: {str(e)}', 'error')

    if emails_enviados > 0:
        report.status = 'Enviado'
        report.data_envio = datetime.utcnow()
        db.session.commit()
        flash(f'Relatório enviado para {emails_enviados} destinatário(s)!', 'success')
    else:
        flash('Nenhum email foi enviado com sucesso.', 'error')

    return redirect(url_for('report_view', report_id=report_id))

# Reimbursement routes
@app.route('/reimbursements')
@login_required
def reimbursements_list():
    # Get search query parameter
    q = request.args.get('q')

    # Start with base query for current user's reimbursements
    query = Reembolso.query.filter_by(usuario_id=current_user.id)

    # Apply intelligent search if query provided
    if q and q.strip():
        from sqlalchemy import or_
        search_term = f"%{q.strip()}%"
        # Left join with project table for searching (since projeto_id can be null)
        query = query.outerjoin(Projeto, Reembolso.projeto_id == Projeto.id)
        query = query.filter(or_(
            Reembolso.descricao_outros.ilike(search_term),
            Reembolso.observacoes.ilike(search_term),
            Projeto.nome.ilike(search_term),
            Projeto.numero.ilike(search_term)
        ))

    reembolsos = query.order_by(Reembolso.created_at.desc()).all()
    return render_template('reimbursements/list.html', reembolsos=reembolsos)

@app.route('/reimbursements/request', methods=['GET', 'POST'])
@login_required
def request_reimbursement():
    """Solicitar novo reembolso"""
    if request.method == 'POST':
        try:
            # Create reimbursement record
            reembolso = Reembolso()
            reembolso.usuario_id = current_user.id
            reembolso.projeto_id = int(request.form.get('projeto_id'))
            # Handle period dates
            periodo_inicio = request.form.get('periodo_inicio')
            periodo_fim = request.form.get('periodo_fim')
            if periodo_inicio:
                reembolso.periodo_inicio = datetime.strptime(periodo_inicio, '%Y-%m-%d').date()
            if periodo_fim:
                reembolso.periodo_fim = datetime.strptime(periodo_fim, '%Y-%m-%d').date()
            reembolso.observacoes = request.form.get('motivo', '')

            # Parse numeric values
            distancia = float(request.form.get('distancia_km', 0))
            valor_km = float(request.form.get('valor_km', 0.75))
            alimentacao = float(request.form.get('alimentacao', 0))
            hospedagem = float(request.form.get('hospedagem', 0))
            outros_gastos = float(request.form.get('outros_gastos', 0))

            reembolso.quilometragem = distancia
            reembolso.valor_km = valor_km
            reembolso.alimentacao = alimentacao
            reembolso.hospedagem = hospedagem
            reembolso.outros_gastos = outros_gastos
            reembolso.descricao_outros = request.form.get('descricao_outros', '')

            # Calculate total
            total_combustivel = distancia * valor_km
            reembolso.total = total_combustivel + alimentacao + hospedagem + outros_gastos

            reembolso.status = 'Aguardando Aprovação'
            reembolso.created_at = datetime.utcnow()

            db.session.add(reembolso)
            db.session.flush()  # Get the ID

            # Handle file uploads (comprovantes)
            upload_folder = app.config.get('UPLOAD_FOLDER', 'uploads')
            if not os.path.exists(upload_folder):
                os.makedirs(upload_folder)

            comprovantes_count = 0
            comprovantes_info = []

            for i in range(4):  # Support up to 4 receipts
                comprovante_key = f'comprovante_{i}'
                desc_key = f'desc_comprovante_{i}'

                if comprovante_key in request.files:
                    file = request.files[comprovante_key]
                    if file and file.filename:
                        try:
                            filename = secure_filename(f"reembolso_{reembolso.id}_{uuid.uuid4().hex}_{file.filename}")
                            filepath = os.path.join(upload_folder, filename)
                            file.save(filepath)

                            desc = request.form.get(desc_key, f'Comprovante {i+1}')
                            comprovantes_info.append({
                                'filename': filename,
                                'description': desc
                            })
                            comprovantes_count += 1
                        except Exception as e:
                            print(f"Erro ao salvar comprovante {i}: {e}")

            # Store comprovantes info as JSON in observacoes
            if comprovantes_info:
                import json
                reembolso.observacoes = json.dumps(comprovantes_info)

            db.session.commit()

            flash(f'Solicitação de reembolso criada com sucesso! {comprovantes_count} comprovantes anexados. Status: Aguardando Aprovação.', 'success')
            return redirect(url_for('reimbursements_list'))

        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao criar solicitação: {str(e)}', 'error')

    projetos = Projeto.query.filter_by(status='Ativo').all()
    return render_template('reimbursements/request.html', projetos=projetos)

@app.route('/reimbursements/<int:id>/approve')
@login_required
def approve_reimbursement(id):
    """Aprovar solicitação de reembolso - apenas usuários master"""
    if not current_user.is_master:
        flash('Acesso negado. Apenas usuários master podem aprovar reembolsos.', 'error')
        return redirect(url_for('reimbursements_list'))

    reembolso = Reembolso.query.get_or_404(id)
    reembolso.status = 'Aprovado'
    reembolso.aprovado_por = current_user.id
    reembolso.data_aprovacao = datetime.utcnow()

    db.session.commit()
    flash(f'Reembolso aprovado com sucesso! PDF disponível para download.', 'success')
    return redirect(url_for('reimbursements_admin'))

@app.route('/reimbursements/<int:id>/reject')
@login_required
def reject_reimbursement(id):
    """Rejeitar solicitação de reembolso - apenas usuários master"""
    if not current_user.is_master:
        flash('Acesso negado. Apenas usuários master podem rejeitar reembolsos.', 'error')
        return redirect(url_for('reimbursements_list'))

    reembolso = Reembolso.query.get_or_404(id)
    reembolso.status = 'Rejeitado'
    reembolso.aprovado_por = current_user.id
    reembolso.data_aprovacao = datetime.utcnow()

    db.session.commit()
    flash(f'Reembolso rejeitado.', 'warning')
    return redirect(url_for('reimbursements_admin'))

@app.route('/install-guide')
def install_guide():
    """Página com instruções de instalação do PWA"""
    return render_template('pwa_install_guide.html')

@app.route('/reimbursements/admin')
@login_required
def reimbursements_admin():
    """Painel administrativo de reembolsos - apenas usuários master"""
    if not current_user.is_master:
        flash('Acesso negado.', 'error')
        return redirect(url_for('reimbursements_list'))

    reembolsos = Reembolso.query.order_by(Reembolso.created_at.desc()).all()
    return render_template('reimbursements/admin.html', reembolsos=reembolsos)

@app.route('/reimbursements/<int:id>/pdf')
@login_required
def generate_reimbursement_pdf(id):
    """Gerar PDF do reembolso aprovado"""
    reembolso = Reembolso.query.get_or_404(id)

    # Check permissions
    if not current_user.is_master and reembolso.usuario_id != current_user.id:
        flash('Acesso negado.', 'error')
        return redirect(url_for('reimbursements_list'))

    if reembolso.status != 'Aprovado':
        flash('PDF só pode ser gerado para reembolsos aprovados.', 'error')
        return redirect(url_for('reimbursements_list'))

    try:
        from pdf_generator import ReportPDFGenerator

        pdf_generator = ReportPDFGenerator()
        output_path = os.path.join('static', 'reimbursements', f'reembolso_{reembolso.id}.pdf')

        # Ensure directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        pdf_path = pdf_generator.generate_reimbursement_pdf(reembolso, output_path)

        return send_from_directory(
            os.path.dirname(output_path),
            os.path.basename(output_path),
            as_attachment=True,
            download_name=f'reembolso_{reembolso.id}_aprovado.pdf'
        )

    except Exception as e:
        flash(f'Erro ao gerar PDF: {str(e)}', 'error')
        return redirect(url_for('reimbursements_list'))

@app.route('/reimbursements/new', methods=['GET', 'POST'])
@login_required
def reimbursement_new():
    form = ReembolsoForm()

    if form.validate_on_submit():
        reembolso = Reembolso(
            usuario_id=current_user.id,
            projeto_id=form.projeto_id.data or None,
            periodo_inicio=form.periodo_inicio.data,
            periodo_fim=form.periodo_fim.data,
            quilometragem=form.quilometragem.data or 0,
            valor_km=form.valor_km.data or 0,
            alimentacao=form.alimentacao.data or 0,
            hospedagem=form.hospedagem.data or 0,
            outros_gastos=form.outros_gastos.data or 0,
            descricao_outros=form.descricao_outros.data,
            observacoes=form.observacoes.data
        )

        reembolso.total = calculate_reimbursement_total(reembolso)

        db.session.add(reembolso)
        db.session.commit()
        flash('Solicitação de reembolso criada com sucesso!', 'success')
        return redirect(url_for('reimbursements_list'))

    return render_template('reimbursement/form.html', form=form)

# File serving (unique function)
@app.route('/uploads/<filename>')
def uploaded_file(filename):
    """Servir imagens com verificação manual de autenticação - VERSÃO RAILWAY OTIMIZADA"""
    try:
        # Verificação manual de autenticação (sem decorator para evitar 302)
        from flask_login import current_user
        import logging

        # Log detalhado para debugging
        current_app.logger.info(f"🔍 SOLICITAÇÃO IMAGEM: {filename}")
        current_app.logger.info(f"🔐 STATUS AUTH: authenticated={current_user.is_authenticated if current_user else False}")

        # Se não autenticado, servir placeholder sem redirecionar
        if not current_user or not current_user.is_authenticated:
            current_app.logger.warning(f"⚠️ USUÁRIO NÃO AUTENTICADO para imagem: {filename}")
            return serve_placeholder_image(filename, "Login necessário para visualizar imagens")

        # Validar filename
        if not filename or filename in ['undefined', 'null', '', 'None']:
            current_app.logger.error(f"❌ FILENAME INVÁLIDO: {repr(filename)}")
            return serve_placeholder_image('arquivo_invalido', "Nome de arquivo inválido")

        current_app.logger.info(f"🔍 INICIANDO BUSCA: {filename}")

        # Buscar no banco PostgreSQL primeiro
        from models import FotoRelatorio, FotoRelatorioExpress

        # Definir diretórios de busca (priorizar uploads)
        search_directories = [
            ('uploads', app.config.get('UPLOAD_FOLDER', 'uploads')),
            ('attached_assets', 'attached_assets'),
            ('static_uploads', os.path.join('static', 'uploads'))
        ]

        # Tentar encontrar nos relatórios normais
        try:
            foto_normal = FotoRelatorio.query.filter_by(filename=filename).first()
            if foto_normal:
                current_app.logger.info(f"✅ ENCONTRADA NO BANCO (Relatório {foto_normal.relatorio_id}): {filename}")

                # Verificar se tem dados binários salvos no banco
                if hasattr(foto_normal, 'imagem') and foto_normal.imagem:
                    current_app.logger.info(f"📱 SERVINDO IMAGEM DIRETAMENTE DO BANCO: {filename}")
                    try:
                        content_type = get_content_type(filename)
                        response = Response(foto_normal.imagem, mimetype=content_type)
                        response.headers['Content-Type'] = content_type
                        response.headers['Cache-Control'] = 'public, max-age=3600'
                        response.headers['X-Image-Source'] = 'database_binary'
                        return response
                    except Exception as binary_error:
                        current_app.logger.error(f"❌ Erro ao servir imagem do banco: {binary_error}")

                # Buscar arquivo físico
                for dir_name, dir_path in search_directories:
                    filepath = os.path.join(dir_path, filename)
                    if os.path.exists(filepath):
                        current_app.logger.info(f"✅ ARQUIVO FÍSICO ENCONTRADO EM {dir_name}: {filepath}")
                        try:
                            content_type = get_content_type(filename)
                            response = send_from_directory(dir_path, filename)
                            response.headers['Content-Type'] = content_type
                            response.headers['Cache-Control'] = 'public, max-age=3600'
                            response.headers['X-Image-Source'] = f'normal_report_{dir_name}'
                            return response
                        except Exception as send_error:
                            current_app.logger.error(f"❌ Erro ao enviar arquivo de {dir_path}: {send_error}")
                            continue

                # Se chegou aqui, arquivo existe no banco mas não no filesystem
                current_app.logger.warning(f"⚠️ ARQUIVO NO BANCO MAS NÃO ENCONTRADO FISICAMENTE: {filename}")
                return serve_placeholder_image(filename, "Imagem registrada no banco mas arquivo físico perdido")
        except Exception as db_error:
            current_app.logger.error(f"❌ Erro ao buscar foto normal no banco: {db_error}")

        # Tentar encontrar nos relatórios express
        try:
            foto_express = FotoRelatorioExpress.query.filter_by(filename=filename).first()
            if foto_express:
                current_app.logger.info(f"✅ ENCONTRADA NO BANCO (Express {foto_express.relatorio_express_id}): {filename}")

                # Verificar se tem dados binários salvos no banco
                if hasattr(foto_express, 'imagem') and foto_express.imagem:
                    current_app.logger.info(f"📱 SERVINDO IMAGEM EXPRESS DIRETAMENTE DO BANCO: {filename}")
                    try:
                        content_type = get_content_type(filename)
                        response = Response(foto_express.imagem, mimetype=content_type)
                        response.headers['Content-Type'] = content_type
                        response.headers['Cache-Control'] = 'public, max-age=3600'
                        response.headers['X-Image-Source'] = 'database_binary_express'
                        return response
                    except Exception as binary_error:
                        current_app.logger.error(f"❌ Erro ao servir imagem express do banco: {binary_error}")

                # Buscar arquivo físico
                for dir_name, dir_path in search_directories:
                    filepath = os.path.join(dir_path, filename)
                    if os.path.exists(filepath):
                        current_app.logger.info(f"✅ ARQUIVO EXPRESS FÍSICO ENCONTRADO EM {dir_name}: {filepath}")
                        try:
                            content_type = get_content_type(filename)
                            response = send_from_directory(dir_path, filename)
                            response.headers['Content-Type'] = content_type
                            response.headers['Cache-Control'] = 'public, max-age=3600'
                            response.headers['X-Image-Source'] = f'express_report_{dir_name}'
                            return response
                        except Exception as send_error:
                            current_app.logger.error(f"❌ Erro ao enviar arquivo express de {dir_path}: {send_error}")
                            continue

                # Se chegou aqui, arquivo existe no banco mas não no filesystem
                current_app.logger.warning(f"⚠️ ARQUIVO EXPRESS NO BANCO MAS NÃO ENCONTRADO FISICAMENTE: {filename}")
                return serve_placeholder_image(filename, "Imagem express registrada no banco mas arquivo físico perdido")
        except Exception as db_error:
            current_app.logger.error(f"❌ Erro ao buscar foto express no banco: {db_error}")

        # Não encontrado no banco - tentar busca física direta
        current_app.logger.warning(f"❌ IMAGEM NÃO ENCONTRADA NO BANCO: {filename}")

        # Busca física direta como fallback
        for dir_name, dir_path in search_directories:
            if os.path.exists(dir_path):
                filepath = os.path.join(dir_path, filename)
                if os.path.exists(filepath):
                    current_app.logger.info(f"🔄 ARQUIVO ENCONTRADO SEM REGISTRO NO BANCO em {dir_name}: {filepath}")
                    try:
                        content_type = get_content_type(filename)
                        response = send_from_directory(dir_path, filename)
                        response.headers['Content-Type'] = content_type
                        response.headers['Cache-Control'] = 'public, max-age=1800'  # Cache menor para arquivos órfãos
                        response.headers['X-Image-Source'] = f'orphan_{dir_name}'
                        return response
                    except Exception as send_error:
                        current_app.logger.error(f"❌ Erro ao enviar arquivo órfão: {send_error}")

        # Arquivo completamente não encontrado
        current_app.logger.error(f"❌ ARQUIVO COMPLETAMENTE NÃO ENCONTRADO: {filename}")
        return serve_placeholder_image(filename, "Imagem não encontrada em nenhum local")

    except Exception as e:
        current_app.logger.exception(f"❌ ERRO CRÍTICO ao servir imagem {filename}")
        return serve_placeholder_image(filename, f"Erro interno do servidor: {str(e)}")

def get_content_type(filename):
    """Determinar content type baseado na extensão"""
    if filename.lower().endswith('.png'):
        return 'image/png'
    elif filename.lower().endswith('.gif'):
        return 'image/gif'
    elif filename.lower().endswith('.webp'):
        return 'image/webp'
    elif filename.lower().endswith(('.jpg', '.jpeg')):
        return 'image/jpeg'
    else:
        return 'image/jpeg'  # default

# Rotas de compatibilidade removidas - sistema simplificado

# Funções de busca complexa removidas - sistema simplificado

def serve_placeholder_image(filename=None, message=None):
    """Serve uma imagem placeholder quando o arquivo não é encontrado"""
    try:
        # Primeiro, tentar servir o placeholder estático se existir
        static_placeholder = os.path.join(os.getcwd(), 'static', 'img', 'no-image.png')
        if os.path.exists(static_placeholder):
            current_app.logger.info(f"📷 Servindo placeholder estático para: {filename}")
            return send_from_directory(os.path.join(os.getcwd(), 'static', 'img'), 'no-image.png')

        # Se não existir, criar SVG placeholder dinâmico
        display_filename = (filename or "arquivo")[:30]
        display_message = message or "não encontrada"

        # Sanitizar strings para SVG
        display_filename = str(display_filename).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        display_message = str(display_message).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

        svg_content = f'''<?xml version="1.0" encoding="UTF-8"?>
<svg width="200" height="150" xmlns="http://www.w3.org/2000/svg" viewBox="0 0 200 150">
  <rect width="100%" height="100%" fill="#f8f9fa" stroke="#dee2e6" stroke-width="1"/>

  <!-- Header -->
  <rect x="5" y="5" width="190" height="25" fill="#e9ecef" rx="3"/>
  <text x="100" y="20" font-family="Arial, sans-serif" font-size="12" 
        fill="#495057" text-anchor="middle" font-weight="bold">📷 ELP Sistema</text>

  <!-- Icon -->
  <circle cx="100" cy="60" r="15" fill="#6c757d" opacity="0.3"/>
  <rect x="92" y="55" width="16" height="10" fill="#fff" opacity="0.8"/>

  <!-- Message -->
  <text x="100" y="85" font-family="Arial, sans-serif" font-size="10" 
        fill="#6c757d" text-anchor="middle">{display_message}</text>

  <!-- Filename -->
  <text x="100" y="105" font-family="monospace" font-size="8" 
        fill="#868e96" text-anchor="middle">{display_filename}</text>

  <!-- Help text -->
  <text x="100" y="125" font-family="Arial, sans-serif" font-size="8" 
        fill="#adb5bd" text-anchor="middle">Recarregue a página</text>

  <!-- Footer -->
  <text x="100" y="140" font-family="Arial, sans-serif" font-size="7" 
        fill="#ced4da" text-anchor="middle">Sistema de Gestão de Obras</text>
</svg>'''

        from flask import Response
        response = Response(svg_content, mimetype='image/svg+xml')
        response.headers['Content-Type'] = 'image/svg+xml'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'

        current_app.logger.info(f"📷 Servindo SVG placeholder para: {filename}")
        return response

    except Exception as e:
        current_app.logger.error(f"❌ Erro ao servir placeholder: {str(e)}")
        # Fallback absoluto - retornar uma resposta de erro simples
        from flask import Response
        return Response("Imagem não encontrada", status=404, mimetype='text/plain')

@app.route("/imagens/<int:id>")
@login_required
def get_imagem(id):
    """Servir imagem diretamente do banco de dados para FotoRelatorio"""
    try:
        from models import FotoRelatorio
        foto = FotoRelatorio.query.get_or_404(id)

        # Se tem imagem no banco, usar ela
        if foto.imagem:
            # Determinar mimetype baseado no filename
            mimetype = get_content_type(foto.filename)
            return Response(foto.imagem, mimetype=mimetype)

        # Fallback: tentar carregar do arquivo se não tem no banco (compatibilidade)
        if foto.filename:
            upload_folder = app.config.get('UPLOAD_FOLDER', 'uploads')
            filepath = os.path.join(upload_folder, foto.filename)
            if os.path.exists(filepath):
                try:
                    with open(filepath, 'rb') as f:
                        file_data = f.read()
                    mimetype = get_content_type(foto.filename)
                    return Response(file_data, mimetype=mimetype)
                except Exception as e:
                    current_app.logger.error(f"Erro ao ler arquivo {filepath}: {e}")

        # Se chegou aqui, não encontrou a imagem
        return serve_placeholder_image(foto.filename, "Imagem não encontrada no banco ou disco")

    except Exception as e:
        current_app.logger.error(f"Erro ao servir imagem ID {id}: {e}")
        return serve_placeholder_image(f"foto_{id}", f"Erro: {str(e)}")

@app.route("/imagens_express/<int:id>")
@login_required
def get_imagem_express(id):
    """Servir imagem diretamente do banco de dados para FotoRelatorioExpress"""
    try:
        from models import FotoRelatorioExpress
        foto = FotoRelatorioExpress.query.get_or_404(id)

        # Se tem imagem no banco, usar ela
        if foto.imagem:
            # Determinar mimetype baseado no filename
            mimetype = get_content_type(foto.filename)
            return Response(foto.imagem, mimetype=mimetype)

        # Fallback: tentar carregar do arquivo se não tem no banco (compatibilidade)
        if foto.filename:
            upload_folder = app.config.get('UPLOAD_FOLDER', 'uploads')
            filepath = os.path.join(upload_folder, foto.filename)
            if os.path.exists(filepath):
                try:
                    with open(filepath, 'rb') as f:
                        file_data = f.read()
                    mimetype = get_content_type(foto.filename)
                    return Response(file_data, mimetype=mimetype)
                except Exception as e:
                    current_app.logger.error(f"Erro ao ler arquivo {filepath}: {e}")

        # Se chegou aqui, não encontrou a imagem
        return serve_placeholder_image(foto.filename, "Imagem não encontrada no banco ou disco")

    except Exception as e:
        current_app.logger.error(f"Erro ao servir imagem express ID {id}: {e}")
        return serve_placeholder_image(f"foto_express_{id}", f"Erro: {str(e)}")

# GPS location endpoint
@app.route('/get_location', methods=['POST', 'GET'])
@csrf.exempt
def get_location():
    import requests
    from urllib.parse import quote

    data = request.get_json()
    latitude = data.get('latitude')
    longitude = data.get('longitude')

    if latitude and longitude:
        try:
            # Use Nominatim reverse geocoding to get formatted address
            url = f"https://nominatim.openstreetmap.org/reverse?format=json&lat={latitude}&lon={longitude}&addressdetails=1&language=pt-BR"
            headers = {'User-Agent': 'SistemaObras/1.0'}
            response = requests.get(url, headers=headers, timeout=10)

            if response.status_code == 200:
                data = response.json()

                # Extract address components
                addr = data.get('address', {})

                # Build formatted address
                address_parts = []

                # Street name and number
                if addr.get('road'):
                    if addr.get('house_number'):
                        address_parts.append(f"{addr.get('road')}, {addr['house_number']}")
                    else:
                        address_parts.append(addr.get('road'))

                # Neighborhood
                if addr.get('suburb') or addr.get('neighbourhood'):
                    address_parts.append(addr.get('suburb') or addr.get('neighbourhood'))

                # City and state
                city = addr.get('city') or addr.get('town') or addr.get('village')
                if city:
                    state = addr.get('state')
                    if state:
                        address_parts.append(f"{city} - {state}")
                    else:
                        address_parts.append(city)

                formatted_address = ', '.join(filter(None, address_parts))

                return jsonify({
                    'success': True,
                    'endereco': formatted_address or data.get('display_name', f"Lat: {latitude}, Lng: {longitude}")
                })

        except Exception as e:
            print(f"Erro ao obter endereço: {e}")

        # Fallback to coordinates if geocoding fails
        return jsonify({
            'success': True,
            'endereco': f"Lat: {latitude}, Lng: {longitude}"
        })

    return jsonify({'success': False})

# Enhanced reporting features

@app.route('/reports/approval-dashboard')
@login_required
def reports_approval_dashboard():
    """Dashboard for report approvals - only for master users"""
    if not current_user.is_master:
        flash('Acesso negado. Apenas usuários master podem acessar o painel de aprovação.', 'error')
        return redirect(url_for('index'))

    # Get reports awaiting approval
    relatorios = Relatorio.query.filter_by(status='Aguardando Aprovação').order_by(Relatorio.created_at.desc()).all()

    return render_template('reports/approval_dashboard.html', relatorios=relatorios)

@app.route('/reports/<int:report_id>/approve', methods=['POST'])
@login_required
def report_approve(report_id):
    """Approve or reject a report"""
    if not current_user.is_master:
        return jsonify({'success': False, 'message': 'Acesso negado.'})

    relatorio = Relatorio.query.get_or_404(report_id)
    data = request.get_json()
    action = data.get('action')
    comment = data.get('comment', '')

    if action == 'approve':
        relatorio.status = 'Aprovado'
        flash_message = 'Relatório aprovado com sucesso.'

        # Fazer backup automático no Google Drive quando aprovado
        try:
            # Preparar dados do relatório para backup
            fotos_paths = []
            fotos = FotoRelatorio.query.filter_by(relatorio_id=report_id).all()
            upload_folder = app.config.get('UPLOAD_FOLDER', 'uploads')

            for foto in fotos:
                foto_path = os.path.join(upload_folder, foto.filename)
                if os.path.exists(foto_path):
                    fotos_paths.append(foto_path)

            report_data = {
                'id': relatorio.id,
                'numero': relatorio.numero,
                'pdf_path': None,  # PDF será gerado se necessário
                'images': fotos_paths
            }

            project_name = f"{relatorio.projeto.numero}_{relatorio.projeto.nome}"
            backup_result = backup_to_drive(report_data, project_name)

            if backup_result.get('success'):
                flash_message += f' Backup realizado: {backup_result.get("successful_uploads", 0)} arquivo(s) enviado(s).'
            else:
                flash_message += f' Aviso de backup: {backup_result.get("message", "Erro no backup")}'

        except Exception as e:
            flash_message += f' Erro no backup: {str(e)}'

    elif action == 'reject':
        relatorio.status = 'Rejeitado'
        flash_message = 'Relatório rejeitado.'
    else:
        return jsonify({'success': False, 'message': 'Ação inválida.'})

    relatorio.aprovador_id = current_user.id
    relatorio.data_aprovacao = datetime.now()
    relatorio.comentario_aprovacao = comment

    db.session.commit()
    
    # Enviar e-mail de aprovação para todos os envolvidos (após commit)
    if action == 'approve':
        try:
            from email_service import EmailService
            email_service = EmailService()
            resultado_email = email_service.send_report_approval_email(report_id)
            
            if resultado_email.get('success'):
                enviados = resultado_email.get('enviados', 0)
                flash_message += f" E-mails enviados: {enviados}"
                current_app.logger.info(f"✅ {enviados} e-mail(s) de aprovação enviados para relatório {relatorio.numero}")
            else:
                erro_msg = resultado_email.get('error', 'Erro desconhecido')
                flash_message += f" Não foi possível enviar os e-mails de notificação."
                current_app.logger.warning(f"⚠️ Erro ao enviar e-mails para relatório {relatorio.numero}: {erro_msg}")
        except Exception as e:
            flash_message += f" Não foi possível enviar os e-mails de notificação."
            current_app.logger.error(f"❌ Exceção ao enviar e-mails para relatório {relatorio.numero}: {str(e)}")

    return jsonify({'success': True, 'message': flash_message})

@app.route('/reports/<int:report_id>/generate-pdf')
@login_required
def report_generate_pdf(report_id):
    """Generate PDF for a report"""
    relatorio = Relatorio.query.get_or_404(report_id)

    # Check permissions
    if relatorio.autor_id != current_user.id and not current_user.is_master:
        flash('Acesso negado.', 'error')
        return redirect(url_for('reports'))

    try:
        pdf_path, filename = generate_visit_report_pdf(relatorio)
        return send_from_directory(
            app.config['UPLOAD_FOLDER'],
            filename,
            as_attachment=True,
            download_name=f"relatorio_{relatorio.numero}.pdf"
        )
    except Exception as e:
        current_app.logger.error(f"Error generating PDF: {str(e)}")
        flash('Erro ao gerar PDF do relatório.', 'error')
        return redirect(url_for('report_view', report_id=report_id))

@app.route('/reports/<int:report_id>/photo-editor')
@login_required
def report_photo_editor(report_id):
    """Editor de fotos professional com Fabric.js para relatórios"""
    relatorio = Relatorio.query.get_or_404(report_id)

    # Check permissions
    if relatorio.autor_id != current_user.id and not current_user.is_master:
        flash('Acesso negado.', 'error')
        return redirect(url_for('reports'))

    photo_id = request.args.get('photo_id', 'temp')
    image_url = request.args.get('image_url', '')

    return render_template('reports/fabric_photo_editor.html', 
                         relatorio=relatorio,
                         photo_id=photo_id,
                         image_url=image_url)

@app.route('/reports/<int:report_id>/photos/annotate', methods=['POST'])
@login_required
@csrf.exempt
def report_photo_annotate(report_id):
    """Save annotated photo"""
    relatorio = Relatorio.query.get_or_404(report_id)

    # Check permissions
    if relatorio.autor_id != current_user.id and not current_user.is_master:
        return jsonify({'success': False, 'message': 'Acesso negado.'})

    photo_id = request.form.get('photo_id')
    annotated_image = request.files.get('annotated_image')

    if not photo_id or not annotated_image:
        return jsonify({'success': False, 'message': 'Dados incompletos.'})

    foto = FotoRelatorio.query.get_or_404(photo_id)

    if foto.relatorio_id != relatorio.id:
        return jsonify({'success': False, 'message': 'Foto não pertence a este relatório.'})

    try:
        # Read file data before saving
        file_data = annotated_image.read()
        annotated_image.seek(0)  # Reset for saving to disk

        # Save annotated image
        filename = secure_filename(f"annotated_{foto.id}_{uuid.uuid4().hex}.png")
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        annotated_image.save(file_path)

        # Update photo record with binary data
        foto.filename_anotada = filename
        foto.imagem = file_data  # Salvar dados binários da imagem anotada
        db.session.commit()

        return jsonify({'success': True, 'message': 'Anotações salvas com sucesso.'})

    except Exception as e:
        db.session.rollback()
        current_app.logger.exception(f"Error saving annotated photo: {str(e)}")
        return jsonify({'success': False, 'message': 'Erro ao salvar anotações.'})

@app.route('/reports/<int:report_id>/submit-for-approval', methods=['POST'])
@login_required
def report_submit_for_approval(report_id):
    """Submit report for approval"""
    relatorio = Relatorio.query.get_or_404(report_id)

    # Check permissions - author or master can submit
    if relatorio.autor_id != current_user.id and not current_user.is_master:
        flash('Acesso negado.', 'error')
        return redirect(url_for('reports'))

    # Allow submission from multiple editable statuses
    status_permitidos = ['preenchimento', 'Rascunho', 'Rejeitado', 'Em edição', 'Aguardando Aprovação']
    if relatorio.status not in status_permitidos:
        flash('Relatório não pode ser enviado para aprovação no status atual.', 'error')
        return redirect(url_for('report_view', report_id=report_id))

    relatorio.status = 'Aguardando Aprovação'
    relatorio.updated_at = datetime.utcnow()
    # Clear any previous rejection comments
    relatorio.comentario_aprovacao = None
    db.session.commit()

    flash('Relatório enviado para aprovação com sucesso!', 'success')
    return redirect(url_for('report_view', report_id=report_id))

@app.route('/visits/<int:visit_id>/communication', methods=['GET', 'POST'])
@login_required
def visit_communication(visit_id):
    """Visit communication system"""
    visita = Visita.query.get_or_404(visit_id)

    if request.method == 'POST':
        mensagem = request.form.get('mensagem')
        tipo = request.form.get('tipo', 'Comunicacao')

        if mensagem:
            comunicacao = ComunicacaoVisita(
                visita_id=visit_id,
                usuario_id=current_user.id,
                mensagem=mensagem,
                tipo=tipo
            )
            db.session.add(comunicacao)
            db.session.commit()

            flash('Comunicação adicionada com sucesso.', 'success')

        return redirect(url_for('visit_communication', visit_id=visit_id))

    # Get all communications for this visit
    comunicacoes = ComunicacaoVisita.query.filter_by(visita_id=visit_id).order_by(ComunicacaoVisita.created_at.desc()).all()

    return render_template('visits/communication.html', visita=visita, comunicacoes=comunicacoes)

# Calendar API routes
@app.route('/api/visits/calendar')
def api_visits_calendar():
    """API endpoint for calendar data - Item 29: Incluir participantes com cores"""
    # Check authentication for API - return JSON 401 instead of HTML redirect
    if not current_user.is_authenticated:
        return jsonify({
            'success': False,
            'error': 'Authentication required'
        }), 401

    try:
        # Enhanced logging for diagnostics
        current_app.logger.info("📅 Carregando dados do calendário...")

        # Buscar todas as visitas com joins corretos - corrigido join problemático
        visits = db.session.query(Visita).outerjoin(
            Projeto, Visita.projeto_id == Projeto.id
        ).join(
            User, Visita.responsavel_id == User.id
        ).all()

        current_app.logger.info(f"📅 {len(visits)} visitas encontradas")

        visits_data = []
        for visit in visits:
            # Buscar participantes da visita com suas cores - Item 29
            participantes = []
            try:
                # Buscar participantes usando query separada para evitar problemas de relacionamento
                participantes_query = db.session.query(VisitaParticipante).filter_by(
                    visita_id=visit.id
                ).join(User, VisitaParticipante.user_id == User.id).all()

                for participante in participantes_query:
                    if participante.user:
                        participantes.append({
                            'id': participante.user.id,
                            'nome': participante.user.nome_completo,
                            'cor_agenda': participante.user.cor_agenda or '#0EA5E9',
                            'confirmado': participante.confirmado
                        })
            except Exception as part_error:
                current_app.logger.warning(f"⚠️ Erro ao carregar participantes da visita {visit.id}: {part_error}")
                participantes = []

            # Buscar dados do responsável separadamente para evitar problemas
            responsavel_nome = ''
            responsavel_cor = '#0EA5E9'
            try:
                responsavel = db.session.get(User, visit.responsavel_id)
                if responsavel:
                    responsavel_nome = responsavel.nome_completo
                    responsavel_cor = responsavel.cor_agenda or '#0EA5E9'

                    # Incluir responsável na lista se não estiver nos participantes
                    responsavel_incluido = any(p['id'] == visit.responsavel_id for p in participantes)
                    if not responsavel_incluido:
                        participantes.insert(0, {
                            'id': responsavel.id,
                            'nome': responsavel.nome_completo,
                            'cor_agenda': responsavel.cor_agenda or '#0EA5E9',
                            'confirmado': True,
                            'is_responsavel': True
                        })
            except Exception as resp_error:
                current_app.logger.warning(f"⚠️ Erro ao carregar responsável da visita {visit.id}: {resp_error}")

            # Buscar dados do projeto separadamente para evitar problemas
            projeto_nome = "Sem projeto"
            projeto_numero = None
            try:
                if visit.projeto_id:
                    projeto = db.session.get(Projeto, visit.projeto_id)
                    if projeto:
                        projeto_nome = f"{projeto.numero} - {projeto.nome}"
                        projeto_numero = projeto.numero
                elif visit.projeto_outros:
                    projeto_nome = visit.projeto_outros
                elif visit.is_pessoal:
                    projeto_nome = "Compromisso Pessoal"
            except Exception as proj_error:
                current_app.logger.warning(f"⚠️ Erro ao carregar projeto da visita {visit.id}: {proj_error}")

            # Item 31: Verificar se é compromisso pessoal
            title = visit.numero or f"Visita {visit.id}"
            if visit.is_pessoal:
                if visit.criado_por != current_user.id:
                    title = "Confidencial"  # Mostrar como confidencial para outros usuários
                    participantes = []  # Não mostrar participantes para compromissos confidenciais
                else:
                    title = f"{visit.numero or f'Visita {visit.id}'} (Pessoal)"

            # PARTE 3: Criar um card para cada participante
            total_participantes = len(participantes)
            has_multiple_participants = total_participantes > 1
            
            # Se não houver participantes (visita antiga), criar ao menos um evento com o responsável
            if not participantes:
                visits_data.append({
                    'id': f"{visit.id}-{visit.responsavel_id}",  # ID único: visit_id-participant_id
                    'visit_id': visit.id,  # ID real da visita para navegação
                    'numero': visit.numero or f"V{visit.id}",
                    'title': title,
                    'data_inicio': visit.data_inicio.isoformat() if visit.data_inicio else None,
                    'data_fim': visit.data_fim.isoformat() if visit.data_fim else None,
                    'data_realizada': visit.data_realizada.isoformat() if visit.data_realizada else None,
                    'status': visit.status or 'Agendada',
                    'projeto_nome': projeto_nome,
                    'projeto_numero': projeto_numero,
                    'responsavel_nome': responsavel_nome,
                    'responsavel_cor': responsavel_cor,
                    'observacoes': visit.observacoes or '',
                    'atividades_realizadas': visit.atividades_realizadas or '',
                    'participantes': [],
                    'is_pessoal': visit.is_pessoal or False,
                    'criado_por': visit.criado_por,
                    'has_multiple_participants': False,
                    'total_participants': 0,
                    'participant_name': responsavel_nome,
                    'participant_id': visit.responsavel_id
                })
            else:
                # Criar um card para cada participante
                for participante in participantes:
                    # Criar título com nome do participante se houver múltiplos
                    participant_title = title
                    if has_multiple_participants:
                        participant_title = f"{title} - {participante['nome']}"
                    
                    visits_data.append({
                        'id': f"{visit.id}-{participante['id']}",  # ID único: visit_id-participant_id
                        'visit_id': visit.id,  # ID real da visita para navegação
                        'numero': visit.numero or f"V{visit.id}",
                        'title': participant_title,
                        'data_inicio': visit.data_inicio.isoformat() if visit.data_inicio else None,
                        'data_fim': visit.data_fim.isoformat() if visit.data_fim else None,
                        'data_realizada': visit.data_realizada.isoformat() if visit.data_realizada else None,
                        'status': visit.status or 'Agendada',
                        'projeto_nome': projeto_nome,
                        'projeto_numero': projeto_numero,
                        'responsavel_nome': responsavel_nome,
                        'responsavel_cor': participante['cor_agenda'],  # Cor do participante específico
                        'observacoes': visit.observacoes or '',
                        'atividades_realizadas': visit.atividades_realizadas or '',
                        'participantes': participantes,  # Manter lista completa para referência
                        'is_pessoal': visit.is_pessoal or False,
                        'criado_por': visit.criado_por,
                        'has_multiple_participants': has_multiple_participants,  # Flag para mostrar ícone
                        'total_participants': total_participantes,
                        'participant_name': participante['nome'],  # Nome específico deste card
                        'participant_id': participante['id']  # ID do participante deste card
                    })

        current_app.logger.info(f"✅ Calendário carregado com {len(visits_data)} eventos")

        # Ensure proper JSON response structure for FullCalendar compatibility
        response_data = visits_data  # Return direct array for frontend compatibility

        return jsonify(response_data)

    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        current_app.logger.exception(f"❌ Erro no calendário API: {str(e)}")
        current_app.logger.error(f"❌ Full traceback: {error_trace}")

        # Always return JSON for API endpoints, never HTML
        return jsonify({
            'success': False,
            'error': 'Erro ao carregar calendário'
        }), 500

@app.route('/api/visits')
def api_visits_list():
    """API endpoint for visits list - simplified version"""
    # Check authentication for API - return JSON 401 instead of HTML redirect
    if not current_user.is_authenticated:
        return jsonify({
            'success': False,
            'error': 'Authentication required'
        }), 401

    try:
        current_app.logger.info("📋 Carregando lista de visitas...")

        # Get all visits with proper joins to avoid lazy loading issues
        visits = db.session.query(Visita).join(
            User, Visita.responsavel_id == User.id
        ).outerjoin(
            Projeto, Visita.projeto_id == Projeto.id
        ).all()

        current_app.logger.info(f"📋 {len(visits)} visitas encontradas")

        visits_data = []
        for visit in visits:
            # Get project name safely
            projeto_nome = "Sem projeto"
            try:
                if visit.projeto_id and visit.projeto:
                    projeto_nome = f"{visit.projeto.numero} - {visit.projeto.nome}"
                elif visit.projeto_outros:
                    projeto_nome = visit.projeto_outros
            except Exception as proj_error:
                current_app.logger.warning(f"⚠️ Erro ao carregar projeto da visita {visit.id}: {proj_error}")

            # Get responsible user safely  
            responsavel_nome = ''
            try:
                if visit.responsavel:
                    responsavel_nome = visit.responsavel.nome_completo
            except Exception as resp_error:
                current_app.logger.warning(f"⚠️ Erro ao carregar responsável da visita {visit.id}: {resp_error}")

            visits_data.append({
                'id': visit.id,
                'numero': visit.numero or f"V{visit.id}",
                'title': visit.numero or f"Visita {visit.id}",
                'start': visit.data_inicio.isoformat() if visit.data_inicio else None,
                'end': visit.data_fim.isoformat() if visit.data_fim else None,
                'status': visit.status or 'Agendada',
                'projeto_nome': projeto_nome,
                'responsavel_nome': responsavel_nome,
                'observacoes': visit.observacoes or '',
                'atividades_realizadas': visit.atividades_realizadas or '',
                'is_pessoal': visit.is_pessoal or False,
                'created_at': visit.created_at.isoformat() if visit.created_at else None
            })

        current_app.logger.info(f"✅ Lista de visitas carregada com {len(visits_data)} itens")

        # Always return array for frontend compatibility
        return jsonify(visits_data)

    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        current_app.logger.exception(f"❌ Erro na API de visitas: {str(e)}")
        current_app.logger.error(f"❌ Full traceback: {error_trace}")

        # Always return JSON for API endpoints, never HTML
        return jsonify({
            'success': False,
            'error': 'Erro ao carregar lista de visitas'
        }), 500

@app.route('/api/visits/<int:visit_id>/details')
@login_required  
def api_visit_details(visit_id):
    """API endpoint for visit details"""
    try:
        visit = Visita.query.get_or_404(visit_id)

        visit_data = {
            'id': visit.id,
            'numero': visit.numero,
            'data_inicio': visit.data_inicio.isoformat() if visit.data_inicio else None,
            'data_fim': visit.data_fim.isoformat() if visit.data_fim else None,
            'data_realizada': visit.data_realizada.isoformat() if visit.data_realizada else None,
            'status': visit.status,
            'projeto_nome': visit.projeto.nome,
            'projeto_numero': visit.projeto.numero,
            'responsavel_nome': visit.responsavel.nome_completo,
            'observacoes_objetivo': visit.observacoes or '',
            'atividades_realizadas': visit.atividades_realizadas or '',
            'observacoes': visit.observacoes or ''
        }

        return jsonify({
            'success': True,
            'visit': visit_data
        })

    except Exception as e:
        print(f"Visit details API error: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/visits/export/google')
@login_required
def api_export_google_calendar():
    """
    PARTE 4: Export all user visits to Google Calendar with duplicate prevention
    Exporta TODAS as visitas do usuário (filtradas por próximos 3 meses)
    """
    try:
        from datetime import datetime, timedelta
        from urllib.parse import urlencode, quote
        
        # Buscar TODAS as visitas do usuário atual (como participante ou responsável)
        # Sem filtro de data para permitir exportação de visitas passadas também
        
        # Buscar visitas onde o usuário é participante OU responsável
        visitas_participante = db.session.query(Visita).join(
            VisitaParticipante, Visita.id == VisitaParticipante.visita_id
        ).filter(
            VisitaParticipante.user_id == current_user.id
        ).all()
        
        visitas_responsavel = Visita.query.filter(
            Visita.responsavel_id == current_user.id
        ).all()
        
        # Combinar e remover duplicatas
        visitas_ids = set([v.id for v in visitas_participante] + [v.id for v in visitas_responsavel])
        visits = Visita.query.filter(Visita.id.in_(visitas_ids)).order_by(Visita.data_inicio).all()

        if not visits:
            return jsonify({
                'success': False,
                'error': 'Nenhuma visita encontrada para exportar'
            })

        # Gerar URLs do Google Calendar para cada visita
        base_url = "https://calendar.google.com/calendar/render?action=TEMPLATE"
        export_urls = []
        novas_exportacoes = 0
        duplicatas = 0

        for visit in visits:
            # PARTE 4: Verificar se já foi exportada (tem google_event_id)
            if visit.google_event_id:
                duplicatas += 1
                continue  # Pular visitas já exportadas
            
            # Formatar para Google Calendar
            start_time = visit.data_inicio.strftime('%Y%m%dT%H%M%S')
            end_time = (visit.data_fim or visit.data_inicio + timedelta(hours=2)).strftime('%Y%m%dT%H%M%S')
            
            # Preparar título e descrição
            projeto_nome = visit.projeto_nome or "Sem projeto"
            title = f"Visita {visit.numero} - {projeto_nome}"
            
            # Buscar participantes para incluir na descrição
            participantes_nomes = []
            try:
                participantes_query = db.session.query(VisitaParticipante).filter_by(
                    visita_id=visit.id
                ).join(User, VisitaParticipante.user_id == User.id).all()
                participantes_nomes = [p.user.nome_completo for p in participantes_query if p.user]
            except:
                pass
            
            participantes_str = ", ".join(participantes_nomes) if participantes_nomes else visit.responsavel.nome_completo
            
            details = f"Observações: {visit.observacoes or 'N/A'}\\n"
            details += f"Projeto: {projeto_nome}\\n"
            details += f"Responsável: {visit.responsavel.nome_completo}\\n"
            details += f"Participantes: {participantes_str}"
            
            # Localização
            location = ""
            if visit.projeto and hasattr(visit.projeto, 'endereco'):
                location = visit.projeto.endereco or ''
            
            # Montar parâmetros
            params = {
                'text': title,
                'dates': f'{start_time}/{end_time}',
                'details': details,
                'location': location
            }
            
            google_url = base_url + '&' + urlencode(params, quote_via=quote)
            
            # Gerar ID único para controle (baseado no ID da visita + timestamp)
            # Quando houver integração API real, este ID será substituído pelo do Google
            event_id = f"visit_{visit.id}_{int(datetime.now().timestamp())}"
            visit.google_event_id = event_id
            
            export_urls.append({
                'visit_id': visit.id,
                'visit_numero': visit.numero,
                'url': google_url,
                'title': title
            })
            
            novas_exportacoes += 1
        
        # Salvar os google_event_id no banco
        if novas_exportacoes > 0:
            db.session.commit()

        return jsonify({
            'success': True,
            'message': f'{novas_exportacoes} visitas prontas para exportação. {duplicatas} já exportadas anteriormente.',
            'total_visits': len(visits),
            'novas_exportacoes': novas_exportacoes,
            'duplicatas': duplicatas,
            'export_urls': export_urls,
            'note': 'Clique em cada URL para adicionar ao Google Calendar. Para sincronização automática bidirecional, configure a integração do Google Calendar.'
        })

    except Exception as e:
        current_app.logger.exception(f"❌ Erro na exportação para Google Calendar: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/visits/<int:visit_id>/export/google')
@login_required
def api_export_single_visit_google(visit_id):
    """Export single visit to Google Calendar"""
    try:
        visit = Visita.query.get_or_404(visit_id)

        # Format for Google Calendar
        start_time = visit.data_inicio.strftime('%Y%m%dT%H%M%S')
        end_time = (visit.data_fim or visit.data_inicio + timedelta(hours=2)).strftime('%Y%m%dT%H%M%S')

        base_url = "https://calendar.google.com/calendar/render?action=TEMPLATE"
        params = {
            'text': f'Visita {visit.numero} - {visit.projeto.nome}',
            'dates': f'{start_time}/{end_time}',
            'details': f'Objetivo: {visit.objetivo}\\nProjeto: {visit.projeto.nome}\\nResponsável: {visit.responsavel.nome_completo}',
            'location': visit.projeto.endereco or 'Localização do projeto'
        }

        google_url = base_url + '&' + '&'.join([f'{k}={v}' for k, v in params.items()])

        return jsonify({
            'success': True,
            'url': google_url
        })

    except Exception as e:
        print(f"Single visit Google export error: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/visits/export/outlook')
@login_required
def api_export_outlook():
    """Export all visits to Google Calendar"""
    return api_export_ics()

@app.route('/api/visits/export/ics')
@login_required  
def api_export_ics():
    """Export visits as ICS file"""
    try:
        from datetime import datetime, timedelta

        visits = Visita.query.filter_by(status='Agendada').all()

        # Generate ICS content
        ics_content = "BEGIN:VCALENDAR\nVERSION:2.0\nPRODID:ELP-Sistema\nCALSCALE:GREGORIAN\n"

        for visit in visits:
            start_time = visit.data_agendada.strftime('%Y%m%dT%H%M%S')
            end_time = (visit.data_agendada + timedelta(hours=2)).strftime('%Y%m%dT%H%M%S')

            ics_content += f"""BEGIN:VEVENT
DTSTART:{start_time}
DTEND:{end_time}
SUMMARY:Visita {visit.numero} - {visit.projeto.nome}
DESCRIPTION:Objetivo: {visit.objetivo}\nProjeto: {visit.projeto.nome}\nResponsável: {visit.responsavel.nome_completo}
LOCATION:{visit.projeto.endereco or 'Localização do projeto'}
UID:{visit.id}@elp-sistema.com
END:VEVENT
"""

        ics_content += "END:VCALENDAR"

        # Create response with ICS file
        response = make_response(ics_content)
        response.headers['Content-Type'] = 'text/calendar'
        response.headers['Content-Disposition'] = 'attachment; filename=visitas_elp.ics'

        return response

    except Exception as e:
        print(f"ICS export error: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/visits/<int:visit_id>/export/outlook')
@login_required
def visit_export_outlook(visit_id):
    """Export single visit to Outlook (.ics file)"""
    try:
        visit = Visita.query.get_or_404(visit_id)

        # Calculate end time
        end_time = visit.data_fim or (visit.data_inicio + timedelta(hours=2))

        # Create ICS content with proper line breaks for Outlook compatibility
        ics_lines = [
            "BEGIN:VCALENDAR",
            "VERSION:2.0",
            "PRODID:-//ELP Consultoria//Visit Scheduler//EN",
            "METHOD:PUBLISH",
            "BEGIN:VEVENT",
            f"UID:visit-{visit.id}@elp.com.br",
            f"DTSTART:{visit.data_inicio.strftime('%Y%m%dT%H%M%S')}",
            f"DTEND:{end_time.strftime('%Y%m%dT%H%M%S')}",
            f"SUMMARY:Visita Tecnica - {visit.projeto.nome}",
            f"DESCRIPTION:Observacoes: {visit.observacoes or 'N/A'}. Projeto: {visit.projeto_nome}. Responsavel: {visit.responsavel.nome_completo}",
            f"LOCATION:{visit.endereco_gps or visit.projeto.endereco or ''}",
            "STATUS:CONFIRMED",
            "SEQUENCE:0",
            "PRIORITY:5",
            "END:VEVENT",
            "END:VCALENDAR"
        ]

        ics_content = "\r\n".join(ics_lines)

        response = make_response(ics_content)
        response.headers["Content-Disposition"] = f"attachment; filename=visita_{visit.numero}.ics"
        response.headers["Content-Type"] = "text/calendar"

        return response

    except Exception as e:
        print(f"Outlook export error: {e}")
        flash('Erro ao exportar para Outlook. Tente novamente.', 'error')
        return redirect(url_for('visits_list'))

# =====================
# ROTAS DE GERENCIAMENTO DE E-MAILS DE CLIENTES
# =====================

@app.route('/projetos/<int:projeto_id>/emails')
@login_required
def projeto_emails(projeto_id):
    """Lista todos os e-mails de clientes de um projeto"""
    projeto = Projeto.query.get_or_404(projeto_id)
    emails = EmailCliente.query.filter_by(projeto_id=projeto_id, ativo=True).order_by(
        EmailCliente.nome_contato
    ).all()

    return render_template('emails/list.html', projeto=projeto, emails=emails)

@app.route('/projetos/<int:projeto_id>/emails/novo', methods=['GET', 'POST'])
@login_required
def novo_email_cliente(projeto_id):
    """Adiciona novo e-mail de cliente ao projeto"""
    projeto = Projeto.query.get_or_404(projeto_id)
    form = EmailClienteForm()

    if form.validate_on_submit():
        # Verificar se o e-mail já existe para este projeto (apenas e-mails ativos)
        email_existente = EmailCliente.query.filter_by(
            projeto_id=projeto_id,
            email=form.email.data.lower().strip(),
            ativo=True
        ).first()

        if email_existente:
            flash('Este e-mail já está cadastrado para esta obra.', 'error')
            return render_template('emails/form.html', form=form, projeto=projeto, titulo='Novo E-mail de Cliente')

        # Criar novo e-mail
        email_cliente = EmailCliente(
            projeto_id=projeto_id,
            email=form.email.data.lower().strip(),
            nome_contato=form.nome_contato.data,
            cargo=form.cargo.data,
            empresa=form.empresa.data,
            telefone=form.telefone.data,
            receber_notificacoes=form.receber_notificacoes.data,
            receber_relatorios=form.receber_relatorios.data,
            ativo=form.ativo.data
        )

        try:
            db.session.add(email_cliente)
            db.session.commit()
            flash(f'E-mail {email_cliente.email} adicionado com sucesso!', 'success')
            return redirect(url_for('projeto_emails', projeto_id=projeto_id))
        except Exception as e:
            db.session.rollback()
            flash('Erro ao adicionar e-mail. Tente novamente.', 'error')

    return render_template('emails/form.html', form=form, projeto=projeto, titulo='Novo E-mail de Cliente')

@app.route('/emails/<int:email_id>/editar', methods=['GET', 'POST'])
@login_required
def editar_email_cliente(email_id):
    """Edita e-mail de cliente"""
    email_cliente = EmailCliente.query.get_or_404(email_id)
    projeto = email_cliente.projeto
    form = EmailClienteForm(obj=email_cliente)

    if form.validate_on_submit():
        # Verificar se mudou o e-mail e se já existe outro com o mesmo e-mail (apenas e-mails ativos)
        if form.email.data.lower().strip() != email_cliente.email:
            email_existente = EmailCliente.query.filter_by(
                projeto_id=projeto.id,
                email=form.email.data.lower().strip(),
                ativo=True
            ).filter(EmailCliente.id != email_id).first()

            if email_existente:
                flash('Este e-mail já está cadastrado para esta obra.', 'error')
                return render_template('emails/form.html', form=form, projeto=projeto, titulo='Editar E-mail de Cliente')

        # Atualizar dados
        email_cliente.email = form.email.data.lower().strip()
        email_cliente.nome_contato = form.nome_contato.data
        email_cliente.cargo = form.cargo.data
        email_cliente.empresa = form.empresa.data
        email_cliente.telefone = form.telefone.data
        email_cliente.receber_notificacoes = form.receber_notificacoes.data
        email_cliente.receber_relatorios = form.receber_relatorios.data
        email_cliente.ativo = form.ativo.data
        email_cliente.updated_at = datetime.utcnow()

        try:
            db.session.commit()
            flash('E-mail atualizado com sucesso!', 'success')
            return redirect(url_for('projeto_emails', projeto_id=projeto.id))
        except Exception as e:
            db.session.rollback()
            flash('Erro ao atualizar e-mail. Tente novamente.', 'error')

    return render_template('emails/form.html', form=form, projeto=projeto, titulo='Editar E-mail de Cliente')

@app.route('/emails/<int:email_id>/remover', methods=['POST'])
@login_required
def remover_email_cliente(email_id):
    """Remove (desativa) e-mail de cliente"""
    email_cliente = EmailCliente.query.get_or_404(email_id)
    projeto_id = email_cliente.projeto_id

    try:
        email_cliente.ativo = False
        email_cliente.updated_at = datetime.utcnow()
        db.session.commit()
        flash('E-mail removido com sucesso!', 'success')
    except Exception as e:
        db.session.rollback()
        flash('Erro ao remover e-mail. Tente novamente.', 'error')

    return redirect(url_for('projeto_emails', projeto_id=projeto_id))

@app.route('/admin/emails')
@login_required
def admin_emails():
    """Painel administrativo para gerenciar todos os e-mails de clientes"""
    if not current_user.is_master:
        flash('Acesso negado. Apenas usuários master podem acessar esta área.', 'error')
        return redirect(url_for('index'))

    # Buscar todos os projetos com seus e-mails
    projetos = Projeto.query.join(EmailCliente).filter(EmailCliente.ativo == True).distinct().all()

    # Contar estatísticas
    total_emails = EmailCliente.query.filter_by(ativo=True).count()
    projetos_sem_email = Projeto.query.filter(~Projeto.id.in_(
        db.session.query(EmailCliente.projeto_id).filter_by(ativo=True).distinct()
    )).count()

    return render_template('emails/admin.html', 
                         projetos=projetos, 
                         total_emails=total_emails,
                         projetos_sem_email=projetos_sem_email)

# Rotas para Legendas Pré-definidas (exclusivo para Administradores)
@app.route('/admin/legendas')
@login_required
def admin_legendas():
    """Painel de administração de legendas - Railway PostgreSQL"""
    if not current_user.is_master:
        flash('Acesso negado. Apenas usuários master podem gerenciar legendas.', 'error')
        return redirect(url_for('index'))

    try:
        # Forçar rollback de transações pendentes
        db.session.rollback()

        from models import LegendaPredefinida

        # Busca
        q = request.args.get("q", "")
        query = LegendaPredefinida.query.filter_by(ativo=True)

        if q and q.strip():
            search_term = f"%{q.strip()}%"
            query = query.filter(LegendaPredefinida.texto.ilike(search_term))

        # Ordenação usando apenas campos que existem
        legendas = query.order_by(
            LegendaPredefinida.categoria.asc(),
            LegendaPredefinida.id.asc()
        ).all()

        current_app.logger.info(f"✅ Admin legendas: {len(legendas)} legendas carregadas")

        return render_template('admin/legendas.html', legendas=legendas, q=q)

    except Exception as e:
        current_app.logger.exception(f"❌ Erro crítico admin legendas: {str(e)}")
        db.session.rollback()
        flash('Erro ao carregar legendas. Tente novamente.', 'error')
        return redirect(url_for('index'))

# Rotas para Configuração de E-mail por Usuário (exclusivo para Administradores)
@app.route('/admin/user-email-configs')
@login_required
def admin_user_email_configs():
    """Painel de administração de configurações de e-mail por usuário"""
    if not current_user.is_master:
        flash('Acesso negado. Apenas usuários master podem gerenciar configurações de e-mail.', 'error')
        return redirect(url_for('index'))

    try:
        from models import UserEmailConfig, User
        
        # Buscar todas as configurações com informações do usuário
        configs = db.session.query(UserEmailConfig, User).join(User).all()
        
        # Usuários que ainda não têm configuração
        users_with_config = [config.UserEmailConfig.user_id for config in configs]
        users_without_config = User.query.filter(~User.id.in_(users_with_config), User.ativo == True).all()
        
        return render_template('admin/user_email_configs.html', 
                             configs=configs,
                             users_without_config=users_without_config)
                             
    except Exception as e:
        current_app.logger.exception(f"❌ Erro ao carregar configurações de e-mail: {str(e)}")
        flash('Erro ao carregar configurações de e-mail. Tente novamente.', 'error')
        return redirect(url_for('index'))

@app.route('/admin/user-email-configs/new', methods=['GET', 'POST'])
@login_required  
def admin_user_email_config_new():
    """Criar nova configuração de e-mail por usuário"""
    if not current_user.is_master:
        flash('Acesso negado. Apenas usuários master podem gerenciar configurações de e-mail.', 'error')
        return redirect(url_for('index'))

    from forms_email import UserEmailConfigForm
    from models import UserEmailConfig, User
    
    form = UserEmailConfigForm()
    
    # Carregar usuários que ainda não têm configuração
    users_with_config = [config.user_id for config in UserEmailConfig.query.all()]
    available_users = User.query.filter(~User.id.in_(users_with_config), User.ativo == True).all()
    form.user_id.choices = [(user.id, f"{user.nome_completo} ({user.email})") for user in available_users]
    
    if form.validate_on_submit():
        try:
            config = UserEmailConfig(
                user_id=form.user_id.data,
                smtp_server=form.smtp_server.data,
                smtp_port=form.smtp_port.data,
                email_address=form.email_address.data,
                use_tls=form.use_tls.data,
                use_ssl=form.use_ssl.data,
                is_active=True
            )
            config.set_password(form.email_password.data)
            
            db.session.add(config)
            db.session.commit()
            
            flash('Configuração de e-mail criada com sucesso!', 'success')
            return redirect(url_for('admin_user_email_configs'))
            
        except Exception as e:
            db.session.rollback()
            current_app.logger.exception(f"❌ Erro ao criar configuração de e-mail: {str(e)}")
            flash('Erro ao criar configuração de e-mail. Tente novamente.', 'error')
    
    return render_template('admin/user_email_config_form.html', form=form, is_edit=False)

@app.route('/admin/user-email-configs/<int:config_id>/edit', methods=['GET', 'POST'])
@login_required
def admin_user_email_config_edit(config_id):
    """Editar configuração de e-mail por usuário"""
    if not current_user.is_master:
        flash('Acesso negado. Apenas usuários master podem gerenciar configurações de e-mail.', 'error')
        return redirect(url_for('index'))

    from forms_email import UserEmailConfigForm
    from models import UserEmailConfig, User
    
    config = UserEmailConfig.query.get_or_404(config_id)
    form = UserEmailConfigForm(obj=config)
    
    # Para edição, apenas mostrar o usuário atual
    form.user_id.choices = [(config.user.id, f"{config.user.nome_completo} ({config.user.email})")]
    form.user_id.data = config.user_id
    
    if form.validate_on_submit():
        try:
            config.smtp_server = form.smtp_server.data
            config.smtp_port = form.smtp_port.data
            config.email_address = form.email_address.data
            config.use_tls = form.use_tls.data
            config.use_ssl = form.use_ssl.data
            
            # Só atualizar senha se uma nova foi fornecida
            if form.email_password.data:
                config.set_password(form.email_password.data)
            
            db.session.commit()
            
            flash('Configuração de e-mail atualizada com sucesso!', 'success')
            return redirect(url_for('admin_user_email_configs'))
            
        except Exception as e:
            db.session.rollback()
            current_app.logger.exception(f"❌ Erro ao atualizar configuração de e-mail: {str(e)}")
            flash('Erro ao atualizar configuração de e-mail. Tente novamente.', 'error')
    
    return render_template('admin/user_email_config_form.html', form=form, config=config, is_edit=True)

@app.route('/admin/user-email-configs/<int:config_id>/delete', methods=['POST'])
@login_required
def admin_user_email_config_delete(config_id):
    """Excluir configuração de e-mail por usuário"""
    if not current_user.is_master:
        flash('Acesso negado. Apenas usuários master podem gerenciar configurações de e-mail.', 'error')
        return redirect(url_for('index'))

    from models import UserEmailConfig
    
    try:
        config = UserEmailConfig.query.get_or_404(config_id)
        user_name = config.user.nome_completo
        
        db.session.delete(config)
        db.session.commit()
        
        flash(f'Configuração de e-mail de {user_name} removida com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception(f"❌ Erro ao excluir configuração de e-mail: {str(e)}")
        flash('Erro ao excluir configuração de e-mail. Tente novamente.', 'error')
    
    return redirect(url_for('admin_user_email_configs'))

@app.route('/admin/user-email-configs/<int:config_id>/toggle', methods=['POST'])
@login_required
def admin_user_email_config_toggle(config_id):
    """Ativar/desativar configuração de e-mail por usuário"""
    if not current_user.is_master:
        flash('Acesso negado. Apenas usuários master podem gerenciar configurações de e-mail.', 'error')
        return redirect(url_for('index'))

    from models import UserEmailConfig
    
    try:
        config = UserEmailConfig.query.get_or_404(config_id)
        user_name = config.user.nome_completo
        
        # Toggle status
        config.is_active = not config.is_active
        status_text = "ativada" if config.is_active else "desativada"
        
        db.session.commit()
        
        flash(f'Configuração de e-mail de {user_name} {status_text} com sucesso!', 'success')
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception(f"❌ Erro ao alterar status da configuração de e-mail: {str(e)}")
        flash('Erro ao alterar status da configuração de e-mail. Tente novamente.', 'error')
    
    return redirect(url_for('admin_user_email_configs'))

@app.route('/admin/user-email-configs/<int:config_id>/test', methods=['POST'])
@login_required
def admin_user_email_config_test(config_id):
    """Testar configuração SMTP de e-mail por usuário"""
    if not current_user.is_master:
        flash('Acesso negado. Apenas usuários master podem gerenciar configurações de e-mail.', 'error')
        return redirect(url_for('index'))

    from models import UserEmailConfig
    import smtplib
    
    try:
        config = UserEmailConfig.query.get_or_404(config_id)
        user_name = config.user.nome_completo
        
        # Testar conexão SMTP
        try:
            # Configurar conexão SMTP
            if config.use_ssl:
                server = smtplib.SMTP_SSL(config.smtp_server, config.smtp_port)
            else:
                server = smtplib.SMTP(config.smtp_server, config.smtp_port)
                if config.use_tls:
                    server.starttls()
            
            # Tentar login
            server.login(config.email_address, config.get_password())
            server.quit()
            
            # Sucesso - atualizar status
            config.last_test_status = 'success'
            config.last_test_at = datetime.utcnow()
            db.session.commit()
            
            flash(f'Teste SMTP para {user_name} realizado com sucesso!', 'success')
            
        except Exception as smtp_error:
            # Erro SMTP - atualizar status
            config.last_test_status = 'error'
            config.last_test_at = datetime.utcnow()
            db.session.commit()
            
            flash(f'Erro no teste SMTP para {user_name}: {str(smtp_error)}', 'error')
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception(f"❌ Erro ao testar configuração SMTP: {str(e)}")
        flash('Erro ao testar configuração SMTP. Tente novamente.', 'error')
    
    return redirect(url_for('admin_user_email_configs'))

@app.route('/admin/legendas/nova', methods=['GET', 'POST'])
@login_required
def admin_legenda_nova():
    """Criar nova legenda predefinida - apenas Usuários Master"""
    if not current_user.is_master:
        flash('Acesso negado. Apenas usuários master podem gerenciar legendas.', 'error')
        return redirect(url_for('index'))

    from forms import LegendaPredefinidaForm
    from models import LegendaPredefinida
    form = LegendaPredefinidaForm()

    if form.validate_on_submit():
        try:
            legenda = LegendaPredefinida()
            legenda.texto = form.texto.data
            legenda.categoria = form.categoria.data
            legenda.ativo = form.ativo.data
            legenda.criado_por = current_user.id

            db.session.add(legenda)
            db.session.commit()

            flash('Legenda criada com sucesso!', 'success')
            return redirect(url_for('admin_legendas'))

        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao criar legenda: {str(e)}', 'error')

    return render_template('admin/legenda_form.html', form=form, title='Nova Legenda')

@app.route('/admin/legendas/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def admin_legenda_editar(id):
    """Editar legenda predefinida - apenas Usuários Master"""
    if not current_user.is_master:
        flash('Acesso negado. Apenas usuários master podem gerenciar legendas.', 'error')
        return redirect(url_for('index'))

    from models import LegendaPredefinida
    from forms import LegendaPredefinidaForm

    legenda = LegendaPredefinida.query.get_or_404(id)
    form = LegendaPredefinidaForm(obj=legenda)

    if form.validate_on_submit():
        try:
            legenda.texto = form.texto.data
            legenda.categoria = form.categoria.data
            legenda.ativo = form.ativo.data

            db.session.commit()

            flash('Legenda atualizada com sucesso!', 'success')
            return redirect(url_for('admin_legendas'))

        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar legenda: {str(e)}', 'error')

    return render_template('admin/legenda_form.html', form=form, legenda=legenda, title='Editar Legenda')

@app.route('/admin/legendas/<int:id>/excluir', methods=['POST'])
@login_required
def admin_legenda_excluir(id):
    """Excluir legenda predefinida - apenas Usuários Master"""
    if not current_user.is_master:
        flash('Acesso negado. Apenas usuários master podem gerenciar legendas.', 'error')
        return redirect(url_for('index'))

    try:
        from models import LegendaPredefinida
        legenda = LegendaPredefinida.query.get_or_404(id)

        db.session.delete(legenda)
        db.session.commit()

        flash('Legenda excluída com sucesso!', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir legenda: {str(e)}', 'error')

    return redirect(url_for('admin_legendas'))

# Rota de diagnóstico para Railway PostgreSQL
@app.route('/api/legendas/diagnostico')
def api_legendas_diagnostico():
    """Diagnóstico completo do sistema de legendas para Railway"""
    try:
        diagnostico = {
            'timestamp': datetime.utcnow().isoformat(),
            'database': {
                'engine': str(db.engine.dialect.name),
                'url_host': db.engine.url.host if db.engine.url else 'N/A'
            },
            'legendas': {},
            'errors': []
        }

        # Forçar rollback
        try:
            db.session.rollback()
        except Exception as rollback_error:
            diagnostico['errors'].append(f"Rollback error: {str(rollback_error)}")

        # Test basic query
        try:
            from models import LegendaPredefinida
            total = LegendaPredefinida.query.count()
            ativas = LegendaPredefinida.query.filter_by(ativo=True).count()

            diagnostico['legendas'] = {
                'total': total,
                'ativas': ativas,
                'inativas': total - ativas
            }

            # Test categorias
            try:
                categorias = db.session.query(LegendaPredefinida.categoria).filter_by(ativo=True).distinct().all()
                diagnostico['legendas']['categorias'] = [cat[0] for cat in categorias]
            except Exception as cat_error:
                diagnostico['errors'].append(f"Categorias error: {str(cat_error)}")

        except Exception as query_error:
            diagnostico['errors'].append(f"Query error: {str(query_error)}")
            diagnostico['legendas'] = {'error': str(query_error)}

        # Test API endpoint
        try:
            with app.test_client() as client:
                api_response = client.get('/api/legendas?categoria=all')
                diagnostico['api_test'] = {
                    'status_code': api_response.status_code,
                    'success': api_response.status_code == 200
                }
        except Exception as api_error:
            diagnostico['errors'].append(f"API test error: {str(api_error)}")

        return jsonify(diagnostico)

    except Exception as e:
        return jsonify({
            'error': str(e),
            'timestamp': datetime.utcnow().isoformat()
        }), 500




@app.route('/api/legendas', methods=['OPTIONS'])
def api_legendas_options():
    """Suporte para requisições OPTIONS (CORS preflight)"""
    response = jsonify({'success': True})
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
    return response

# API para salvar dados de relatórios (mobile/desktop)
@app.route('/api/relatorios', methods=['POST'])
def api_salvar_relatorio():
    """API para salvar relatórios - compatível mobile/desktop"""
    try:
        data = request.get_json()

        if not data:
            return jsonify({
                'success': False,
                'error': 'Dados não recebidos'
            }), 400

        # Aqui você pode implementar a lógica de salvamento
        # Por enquanto, só retorna sucesso para confirmar que a API funciona

        return jsonify({
            'success': True,
            'message': 'Dados recebidos com sucesso',
            'data_received': len(str(data))
        })

    except Exception as e:
        print(f"ERRO API SALVAR: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/test')
def api_test():
    """API de teste para verificar conectividade"""
    try:
        from models import LegendaPredefinida
        count = LegendaPredefinida.query.filter_by(ativo=True).count()

        return jsonify({
            'success': True,
            'message': 'API funcionando',
            'database_connection': True,
            'legendas_count': count,
            'timestamp': datetime.utcnow().isoformat()
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'database_connection': False
        }), 500

@app.route('/clear-pwa-cache')
def clear_pwa_cache():
    """Endpoint para forçar limpeza de cache PWA mobile"""
    return """
    <!DOCTYPE html>
    <html>
    <head>
        <title>Limpeza PWA Cache</title>
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <script src="{{ url_for('static', filename='js/force-online-mode.js') }}"></script>
    </head>
    <body style="font-family: Arial, sans-serif; padding: 20px; background: #f8f9fa;">
        <div style="max-width: 600px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
            <h1 style="color: #20c1e8; text-align: center;">🧹 Limpeza PWA Cache</h1>
            <p style="text-align: center; color: #666; margin-bottom: 30px;">
                Este endpoint força a limpeza completa do cache PWA mobile para garantir dados idênticos do PostgreSQL.
            </p>

            <div style="background: #e8f5e8; padding: 15px; border-radius: 8px; margin: 20px 0;">
                <h3 style="color: #2d5f2d; margin-top: 0;">✅ Ações Executadas:</h3>
                <ul style="color: #2d5f2d;">
                    <li>localStorage completamente limpo</li>
                    <li>sessionStorage removido</li> 
                    <li>Service Workers desregistrados</li>
                    <li>Cache do navegador removido</li>
                    <li>Reload forçado sem cache</li>
                </ul>
            </div>

            <div style="text-align: center; margin-top: 30px;">
                <button onclick="window.clearPWACache()" style="background: #20c1e8; color: white; border: none; padding: 15px 30px; border-radius: 8px; font-size: 16px; cursor: pointer;">
                    🔄 Limpar Cache Agora
                </button>
            </div>

            <div style="background: #fff3cd; padding: 15px; border-radius: 8px; margin: 20px 0;">
                <h4 style="color: #856404; margin-top: 0;">📱 Como usar no Mobile PWA:</h4>
                <ol style="color: #856404;">
                    <li>Acesse esta URL no app mobile instalado</li>
                    <li>Clique em "Limpar Cache Agora"</li>
                    <li>Aguarde o reload automático</li>
                    <li>Verifique se os dados estão idênticos ao desktop</li>
                </ol>
            </div>
        </div>

        <script>
            // Executar limpeza automática ao carregar
            console.log('🧹 Iniciando limpeza automática PWA cache...');
            setTimeout(() => {
                new ForceOnlineMode();
            }, 1000);
        </script>
    </body>
    </html>
    """

# Rotas administrativas para Checklist Padrão
@app.route('/admin/checklist-padrao')
@login_required
def admin_checklist_padrao():
    """Página para administradores gerenciarem checklist padrão"""
    if not current_user.is_master:
        flash('Acesso negado. Apenas administradores podem acessar esta página.', 'error')
        return redirect(url_for('index'))

    checklist_items = ChecklistPadrao.query.filter_by(ativo=True).order_by(ChecklistPadrao.ordem).all()
    return render_template('admin/checklist_padrao.html', checklist_items=checklist_items)

@app.route('/developer/checklist-padrao')
@login_required
def developer_checklist_padrao():
    """Página para desenvolvedores gerenciarem checklist padrão"""
    if not current_user.is_developer:
        flash('Acesso negado. Apenas desenvolvedores podem acessar esta página.', 'error')
        return redirect(url_for('index'))

    checklist_items = ChecklistPadrao.query.filter_by(ativo=True).order_by(ChecklistPadrao.ordem).all()
    return render_template('developer/checklist_padrao.html', checklist_items=checklist_items)

@app.route('/developer/api/checklist/default')
@login_required
def api_checklist_default():
    """API para carregar itens de checklist padrão para Express Reports"""
    try:
        # Buscar itens de checklist padrão criados no perfil desenvolvedor
        checklist_items = ChecklistPadrao.query.filter_by(ativo=True).order_by(ChecklistPadrao.ordem, ChecklistPadrao.id).all()

        items_data = []
        for item in checklist_items:
            items_data.append({
                'id': item.id,
                'texto': item.texto,  # ChecklistPadrao usa 'texto' não 'titulo'
                'descricao': getattr(item, 'descricao', '') or '',  # Campo opcional
                'categoria': getattr(item, 'categoria', 'Geral') or 'Geral',  # Campo opcional
                'ordem': item.ordem or 0,
                'obrigatorio': getattr(item, 'obrigatorio', False)
            })

        return jsonify({
            'success': True,
            'items': items_data,
            'total': len(items_data)
        })

    except Exception as e:
        current_app.logger.error(f"Erro ao carregar checklist padrão: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Erro ao carregar checklist padrão',
            'details': str(e)
        })

@app.route('/api/projeto/<int:project_id>/checklist')
@login_required
def api_get_project_checklist(project_id):
    """API para obter checklist do projeto (personalizado ou padrão)"""
    try:
        projeto = Projeto.query.get_or_404(project_id)
        
        # Verificar configuração de checklist do projeto
        config = ProjetoChecklistConfig.query.filter_by(projeto_id=project_id).first()
        
        checklist_items = []
        checklist_type = 'padrao'
        
        if config and config.tipo_checklist == 'personalizado':
            # Buscar checklist personalizado da obra
            items = ChecklistObra.query.filter_by(
                projeto_id=project_id, 
                ativo=True
            ).order_by(ChecklistObra.ordem).all()
            
            if items:
                checklist_type = 'personalizado'
                for item in items:
                    checklist_items.append({
                        'id': item.id,
                        'texto': item.texto,
                        'ordem': item.ordem or 0
                    })
        
        # Se não tiver personalizado ou não tiver itens, usar padrão
        if not checklist_items:
            items = ChecklistPadrao.query.filter_by(ativo=True).order_by(ChecklistPadrao.ordem).all()
            checklist_type = 'padrao'
            
            for item in items:
                checklist_items.append({
                    'id': item.id,
                    'texto': item.texto,
                    'ordem': item.ordem or 0,
                    'descricao': getattr(item, 'descricao', '') or '',
                    'categoria': getattr(item, 'categoria', 'Geral') or 'Geral'
                })
        
        return jsonify({
            'success': True,
            'checklist': checklist_items,
            'tipo': checklist_type,
            'project_id': project_id,
            'project_name': projeto.nome,
            'total': len(checklist_items)
        })
    
    except Exception as e:
        current_app.logger.error(f"Erro ao carregar checklist do projeto {project_id}: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Erro ao carregar checklist',
            'details': str(e)
        }), 500

# Google Drive Backup Routes
@app.route('/admin/drive/test')
@login_required
def admin_drive_test():
    """Testar conexão com Google Drive - apenas administradores"""
    if not current_user.is_master:
        flash('Acesso negado. Apenas administradores podem testar a conexão.', 'error')
        return redirect(url_for('index'))

    try:
        result = test_drive_connection()
        if result['success']:
            flash(f'Conexão OK: {result["message"]}', 'success')
        else:
            flash(f'Erro na conexão: {result["message"]}', 'error')
    except Exception as e:
        flash(f'Erro ao testar conexão: {str(e)}', 'error')

    return redirect(url_for('index'))

@app.route('/admin/drive/force-backup/<int:report_id>')
@login_required
def admin_force_backup(report_id):
    """Forçar backup de relatório específico - apenas administradores"""
    if not current_user.is_master:
        return jsonify({'success': False, 'message': 'Acesso negado'})

    try:
        relatorio = Relatorio.query.get_or_404(report_id)

        # Preparar dados do relatório para backup
        fotos_paths = []
        fotos = FotoRelatorio.query.filter_by(relatorio_id=report_id).all()
        upload_folder = app.config.get('UPLOAD_FOLDER', 'uploads')

        for foto in fotos:
            foto_path = os.path.join(upload_folder, foto.filename)
            if os.path.exists(foto_path):
                fotos_paths.append(foto_path)

        report_data = {
            'id': relatorio.id,
            'numero': relatorio.numero,
            'pdf_path': None,  # PDF será gerado se necessário
            'images': fotos_paths
        }

        project_name = f"{relatorio.projeto.numero}_{relatorio.projeto.nome}"
        backup_result = backup_to_drive(report_data, project_name)

        if backup_result.get('success'):
            return jsonify({
                'success': True,
                'message': f'Backup realizado: {backup_result.get("successful_uploads", 0)} arquivo(s) enviado(s)',
                'details': backup_result
            })
        else:
            return jsonify({
                'success': False,
                'message': backup_result.get('message', 'Erro no backup'),
                'error': backup_result.get('error')
            })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro ao forçar backup: {str(e)}'
        })

@app.route('/developer/checklist-padrao/add', methods=['POST'])
@login_required
@csrf.exempt
def developer_checklist_add():
    """Adicionar novo item ao checklist padrão"""
    if not current_user.is_developer:
        return jsonify({'error': 'Acesso negado'}), 403

    try:
        data = request.get_json()
        texto = data.get('texto', '').strip()

        if not texto:
            return jsonify({'error': 'Texto é obrigatório'}), 400

        if len(texto) > 500:
            return jsonify({'error': 'Texto deve ter no máximo 500 caracteres'}), 400

        # Verificar se já existe
        existing = ChecklistPadrao.query.filter_by(texto=texto, ativo=True).first()
        if existing:
            return jsonify({'error': 'Item já existe no checklist'}), 400

        # Obter próximo número de ordem
        max_ordem = db.session.query(db.func.max(ChecklistPadrao.ordem)).scalar() or 0

        # Criar novo item
        novo_item = ChecklistPadrao(
            texto=texto,
            ordem=max_ordem + 1
        )

        db.session.add(novo_item)
        db.session.commit()

        return jsonify({'success': True, 'item_id': novo_item.id})

    except Exception as e:
        db.session.rollback()
        print(f"Erro ao adicionar item checklist: {e}")
        return jsonify({'error': f'Erro interno: {str(e)}'}), 500

@app.route('/developer/checklist-padrao/edit/<int:item_id>', methods=['PUT'])
@login_required
@csrf.exempt
def developer_checklist_edit(item_id):
    """Editar item do checklist padrão"""
    if not current_user.is_developer:
        return jsonify({'error': 'Acesso negado'}), 403

    try:
        item = ChecklistPadrao.query.get_or_404(item_id)
        data = request.get_json()
        novo_texto = data.get('texto', '').strip()

        if not novo_texto:
            return jsonify({'error': 'Texto é obrigatório'}), 400

        if len(novo_texto) > 500:
            return jsonify({'error': 'Texto deve ter no máximo 500 caracteres'}), 400

        # Verificar duplicatas (exceto o item atual)
        existing = ChecklistPadrao.query.filter(
            ChecklistPadrao.texto == novo_texto,
            ChecklistPadrao.ativo == True,
            ChecklistPadrao.id != item_id
        ).first()

        if existing:
            return jsonify({'error': 'Já existe um item com este texto'}), 400

        item.texto = novo_texto
        item.updated_at = datetime.utcnow()

        db.session.commit()

        return jsonify({'success': True})

    except Exception as e:
        db.session.rollback()
        print(f"Erro ao editar item checklist: {e}")
        return jsonify({'error': f'Erro interno: {str(e)}'}), 500

@app.route('/developer/checklist-padrao/delete/<int:item_id>', methods=['DELETE'])
@login_required
@csrf.exempt
def developer_checklist_delete(item_id):
    """Remover item do checklist padrão"""
    if not current_user.is_developer:
        return jsonify({'error': 'Acesso negado'}), 403

    try:
        item = ChecklistPadrao.query.get_or_404(item_id)

        # Marcar como inativo em vez de deletar fisicamente
        item.ativo = False
        item.updated_at = datetime.utcnow()

        db.session.commit()

        return jsonify({'success': True})

    except Exception as e:
        db.session.rollback()
        print(f"Erro ao remover item checklist: {e}")
        return jsonify({'error': f'Erro interno: {str(e)}'}), 500

@app.route('/developer/checklist-padrao/reorder', methods=['POST'])
@login_required
@csrf.exempt
def developer_checklist_reorder():
    """Reordenar itens do checklist padrão"""
    if not current_user.is_developer:
        return jsonify({'error': 'Acesso negado'}), 403

    try:
        data = request.get_json()
        items = data.get('items', [])

        for item_data in items:
            item_id = item_data.get('id')
            nova_ordem = item_data.get('ordem')

            if item_id and nova_ordem:
                item = ChecklistPadrao.query.get(item_id)
                if item:
                    item.ordem = nova_ordem
                    item.updated_at = datetime.utcnow()

        db.session.commit()

        return jsonify({'success': True})

    except Exception as e:
        db.session.rollback()
        print(f"Erro ao reordenar checklist: {e}")
        return jsonify({'error': f'Erro interno: {str(e)}'}), 500

# ====== ADMIN ROUTES FOR CHECKLIST (equivalentes às developer) ======

@app.route('/admin/checklist-padrao/add', methods=['POST'])
@login_required
@csrf.exempt
def admin_checklist_add():
    """Adicionar novo item ao checklist padrão (admin)"""
    if not current_user.is_master:
        return jsonify({'error': 'Acesso negado'}), 403

    try:
        data = request.get_json()
        texto = data.get('texto', '').strip()

        if not texto:
            return jsonify({'error': 'Texto é obrigatório'}), 400

        # Get next order
        ultimo_item = ChecklistPadrao.query.order_by(ChecklistPadrao.ordem.desc()).first()
        nova_ordem = (ultimo_item.ordem + 1) if ultimo_item else 1

        novo_item = ChecklistPadrao(
            texto=texto,
            ordem=nova_ordem,
            ativo=True
        )

        db.session.add(novo_item)
        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Item adicionado com sucesso',
            'item': {
                'id': novo_item.id,
                'texto': novo_item.texto,
                'ordem': novo_item.ordem
            }
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Erro interno: {str(e)}'}), 500

@app.route('/admin/checklist-padrao/edit/<int:item_id>', methods=['PUT'])
@login_required
@csrf.exempt
def admin_checklist_edit(item_id):
    """Editar item do checklist padrão (admin)"""
    if not current_user.is_master:
        return jsonify({'error': 'Acesso negado'}), 403

    try:
        item = ChecklistPadrao.query.get_or_404(item_id)

        data = request.get_json()
        novo_texto = data.get('texto', '').strip()

        if not novo_texto:
            return jsonify({'error': 'Texto é obrigatório'}), 400

        item.texto = novo_texto
        item.updated_at = datetime.utcnow()
        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Item atualizado com sucesso',
            'item': {
                'id': item.id,
                'texto': item.texto,
                'ordem': item.ordem
            }
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Erro interno: {str(e)}'}), 500

@app.route('/admin/checklist-padrao/delete/<int:item_id>', methods=['DELETE'])
@login_required
@csrf.exempt
def admin_checklist_delete(item_id):
    """Remover item do checklist padrão (admin)"""
    if not current_user.is_master:
        return jsonify({'error': 'Acesso negado'}), 403

    try:
        item = ChecklistPadrao.query.get_or_404(item_id)

        # Marcar como inativo em vez de deletar fisicamente
        item.ativo = False
        item.updated_at = datetime.utcnow()

        db.session.commit()

        return jsonify({'success': True})

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Erro interno: {str(e)}'}), 500

@app.route('/admin/checklist-padrao/reorder', methods=['POST'])
@login_required
@csrf.exempt
def admin_checklist_reorder():
    """Reordenar itens do checklist padrão (admin)"""
    if not current_user.is_master:
        return jsonify({'error': 'Acesso negado'}), 403

    try:
        data = request.get_json()
        items = data.get('items', [])

        for item_data in items:
            item_id = item_data.get('id')
            nova_ordem = item_data.get('ordem')

            if item_id and nova_ordem:
                item = ChecklistPadrao.query.get(item_id)
                if item:
                    item.ordem = nova_ordem
                    item.updated_at = datetime.utcnow()

        db.session.commit()

        return jsonify({'success': True})

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Erro interno: {str(e)}'}), 500

# =============================================================================
# SISTEMA DE E-MAIL - ROTAS PARA ENVIO DE RELATÓRIOS POR E-MAIL
# =============================================================================

@app.route('/admin/configuracao-email')
@login_required
def configuracao_email_list():
    """Lista as configurações de e-mail"""
    if not (current_user.is_master or current_user.is_developer):
        flash('Acesso negado. Apenas administradores podem configurar e-mails.', 'error')
        return redirect(url_for('index'))

    configs = ConfiguracaoEmail.query.order_by(ConfiguracaoEmail.nome_configuracao).all()
    return render_template('admin/configuracao_email_list.html', configs=configs)

@app.route('/admin/configuracao-email/nova', methods=['GET', 'POST'])
@login_required
def configuracao_email_nova():
    """Criar nova configuração de e-mail"""
    if not (current_user.is_master or current_user.is_developer):
        flash('Acesso negado.', 'error')
        return redirect(url_for('index'))

    form = ConfiguracaoEmailForm()
    if form.validate_on_submit():
        try:
            # Se marcar como ativo, desativar outras configurações
            if form.ativo.data:
                ConfiguracaoEmail.query.filter_by(ativo=True).update({'ativo': False})

            config = ConfiguracaoEmail(
                nome_configuracao=form.nome_configuracao.data,
                servidor_smtp=form.servidor_smtp.data,
                porta_smtp=form.porta_smtp.data,
                use_tls=form.use_tls.data,
                use_ssl=form.use_ssl.data,
                email_remetente=form.email_remetente.data,
                nome_remetente=form.nome_remetente.data,
                template_assunto=form.template_assunto.data or "Relatório do Projeto {projeto_nome} - {data}",
                template_corpo=form.template_corpo.data or """<p>Prezado(a) {nome_cliente},</p><p>Segue em anexo o relatório da obra/projeto conforme visita realizada em {data_visita}.</p><p>Em caso de dúvidas, favor entrar em contato conosco.</p><p>Atenciosamente,<br>Equipe ELP Consultoria e Engenharia<br>Engenharia Civil & Fachadas</p>""",
                ativo=form.ativo.data
            )

            db.session.add(config)
            db.session.commit()
            flash('Configuração de e-mail criada com sucesso!', 'success')
            return redirect(url_for('configuracao_email_list'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao criar configuração: {str(e)}', 'error')

    return render_template('admin/configuracao_email_form.html', form=form, title='Nova Configuração de E-mail')

@app.route('/relatorio/<int:relatorio_id>/enviar-email', methods=['GET', 'POST'])
@login_required
def relatorio_enviar_email(relatorio_id):
    """Enviar relatório por e-mail"""
    relatorio = Relatorio.query.get_or_404(relatorio_id)

    # Verificar se o usuário tem acesso ao projeto
    if not current_user.is_master and relatorio.projeto.responsavel_id != current_user.id:
        flash('Acesso negado.', 'error')
        return redirect(url_for('index'))

    # Buscar e-mails do projeto
    emails_projeto = email_service.buscar_emails_projeto(relatorio.projeto.id)

    if not emails_projeto:
        flash('Nenhuma e-mail cadastrado para esta obra. Cadastre e-mails na seção de clientes.', 'warning')
        return redirect(url_for('report_view', report_id=relatorio_id))

    # Configurar escolhas do formulário
    form = EnvioEmailForm()
    form.destinatarios.choices = [
        (email.email, f"{email.nome_contato} ({email.email}) - {email.cargo or 'N/A'}")
        for email in emails_projeto
    ]

    # Buscar configuração ativa
    config_ativa = email_service.get_configuracao_ativa()
    if not config_ativa:
        flash('Nenhuma configuração de e-mail ativa. Configure o sistema de e-mail primeiro.', 'error')
        return redirect(url_for('report_view', report_id=relatorio_id))

    if request.method == 'POST' and form.validate_on_submit():
        try:
            # Processar e-mails (mesmo processo do relatório normal)
            def processar_emails(texto_emails):
                if not texto_emails:
                    return []
                emails = []
                for linha in texto_emails.split('\n'):
                    for email in linha.split(','):
                        email = email.strip()
                        if email:
                            emails.append(email)
                return emails

            cc_emails = processar_emails(form.cc_emails.data)
            bcc_emails = processar_emails(form.bcc_emails.data)

            # Validar e-mails
            todos_emails = form.destinatarios.data + cc_emails + bcc_emails
            emails_validos, emails_invalidos = email_service.validar_emails(todos_emails)

            if emails_invalidos:
                flash(f'E-mails inválidos encontrados: {", ".join(emails_invalidos)}', 'error')
                return render_template('email/enviar_relatorio.html', 
                                     form=form, relatorio=relatorio, config=config_ativa)

            # Preparar dados para envio
            destinatarios_data = {
                'destinatarios': form.destinatarios.data,
                'cc': cc_emails,
                'bcc': bcc_emails,
                'assunto_custom': form.assunto_personalizado.data,
                'corpo_custom': form.corpo_personalizado.data
            }

            # Enviar e-mails
            resultado = email_service.enviar_relatorio_por_email(
                relatorio, destinatarios_data, current_user.id
            )

            if resultado['success']:
                if resultado['falhas'] > 0:
                    flash(f'E-mails enviados parcialmente: {resultado["sucessos"]} sucessos, {resultado["falhas"]} falhas.', 'warning')
                else:
                    flash(f'E-mails enviados com sucesso para {resultado["sucessos"]} destinatários!', 'success')
            else:
                flash(f'Erro ao enviar e-mails: {resultado.get("error", "Erro desconhecido")}', 'error')

            return redirect(url_for('report_view', report_id=relatorio_id))

        except Exception as e:
            flash(f'Erro ao enviar e-mails: {str(e)}', 'error')

    return render_template('email/enviar_relatorio.html', 
                         form=form, relatorio=relatorio, config=config_ativa)

@app.route('/relatorio/<int:relatorio_id>/preview-email')
@login_required
def relatorio_preview_email(relatorio_id):
    """Preview do e-mail antes de enviar"""
    relatorio = Relatorio.query.get_or_404(relatorio_id)

    # Verificar acesso
    if not current_user.is_master and relatorio.projeto.responsavel_id != current_user.id:
        return jsonify({'error': 'Acesso negado'}), 403

    # Buscar configuração ativa
    config = email_service.get_configuracao_ativa()
    if not config:
        return jsonify({'error': 'Nenhuma configuração de e-mail ativa'}), 400

    # Preparar dados do preview
    projeto = relatorio.projeto
    data_visita = relatorio.data_visita.strftime('%d/%m/%Y') if relatorio.data_visita else 'N/A'
    data_atual = datetime.now().strftime('%d/%m/%Y')

    assunto = config.template_assunto.format(
        projeto_nome=projeto.nome,
        data=data_atual
    )

    corpo_html = config.template_corpo.format(
        nome_cliente="[Nome do Cliente]",
        data_visita=data_visita,
        projeto_nome=projeto.nome
    )

    return jsonify({
        'assunto': assunto,
        'corpo_html': corpo_html
    })

# =============================================================================
# SISTEMA DE RELATÓRIO EXPRESS STANDALONE
# =============================================================================

@app.route('/express')
@login_required
def express_list():
    """Listar relatórios express"""
    relatorios = RelatorioExpress.query.order_by(RelatorioExpress.created_at.desc()).all()

    # Estatísticas
    relatorios_rascunho = len([r for r in relatorios if r.status == 'rascunho'])
    relatorios_finalizados = len([r for r in relatorios if r.status == 'finalizado'])

    return render_template('express/list.html', 
                         relatorios=relatorios,
                         relatorios_rascunho=relatorios_rascunho,
                         relatorios_finalizados=relatorios_finalizados)

@app.route('/express/novo', methods=['GET', 'POST'])
@login_required
def express_new():
    # Detectar se é mobile
    user_agent = request.headers.get('User-Agent', '').lower()
    is_mobile = any(device in user_agent for device in ['mobile', 'android', 'iphone', 'ipad']) or request.args.get('mobile') == '1'

    form = RelatorioExpressForm()
    
    # Carregar categorias dinâmicas se projeto_id for fornecido (Item 16)
    # Pegar de args (GET) ou form (POST)
    projeto_id = request.args.get('projeto_id', type=int) or request.form.get('projeto_id', type=int)
    categorias_lista = []
    
    if projeto_id:
        from models import CategoriaObra, Projeto
        projeto = Projeto.query.get(projeto_id)
        categorias = CategoriaObra.query.filter_by(projeto_id=projeto_id).order_by(CategoriaObra.ordem).all()
        
        current_app.logger.info(f"📂 EXPRESS: projeto_id={projeto_id}, categorias encontradas={len(categorias)}")
        
        if categorias:
            # Usar categorias customizadas do projeto
            categorias_lista = [{'nome': cat.nome_categoria, 'icon': 'fa-tag'} for cat in categorias]
            categoria_choices = [('', 'Selecione...')] + [(cat.nome_categoria, cat.nome_categoria) for cat in categorias]
            # RelatorioExpressForm não tem foto_forms (usa MultipleFileField simples)
            if hasattr(form, 'foto_forms'):
                for foto_form in form.foto_forms:
                    foto_form.tipo_servico.choices = categoria_choices
            current_app.logger.info(f"✅ EXPRESS: Usando {len(categorias_lista)} categorias customizadas: {[c['nome'] for c in categorias_lista]}")
        else:
            # Projeto sem categorias cadastradas
            categorias_lista = []
            if hasattr(form, 'foto_forms'):
                for foto_form in form.foto_forms:
                    foto_form.tipo_servico.choices = [('', 'Nenhuma categoria cadastrada para este projeto')]
            current_app.logger.warning(f"⚠️ EXPRESS: Projeto {projeto_id} não tem categorias cadastradas")
    else:
        # Sem projeto_id
        current_app.logger.warning("⚠️ EXPRESS: Sem projeto_id fornecido")
        categorias_lista = []
        if hasattr(form, 'foto_forms'):
            for foto_form in form.foto_forms:
                foto_form.tipo_servico.choices = [('', 'Projeto não selecionado')]

    if form.validate_on_submit():
        try:
            # Determinar ação
            action = request.form.get('action', 'save_draft')

            # Gerar número único
            numero = gerar_numero_relatorio_express()

            # Criar relatório express
            relatorio_express = RelatorioExpress()
            relatorio_express.numero = numero
            relatorio_express.autor_id = current_user.id

            # Dados da empresa
            relatorio_express.empresa_nome = form.empresa_nome.data
            relatorio_express.empresa_endereco = form.empresa_endereco.data
            relatorio_express.empresa_telefone = form.empresa_telefone.data
            relatorio_express.empresa_email = form.empresa_email.data
            relatorio_express.empresa_responsavel = form.empresa_responsavel.data

            # Dados da visita
            relatorio_express.data_visita = form.data_visita.data
            relatorio_express.periodo_inicio = form.periodo_inicio.data
            relatorio_express.periodo_fim = form.periodo_fim.data
            relatorio_express.condicoes_climaticas = form.condicoes_climaticas.data
            relatorio_express.temperatura = form.temperatura.data

            # Localização - processar coordenadas primeiro
            endereco_final = form.endereco_visita.data  # Valor padrão do formulário
            
            if form.latitude.data:
                relatorio_express.latitude = float(form.latitude.data)
            if form.longitude.data:
                relatorio_express.longitude = float(form.longitude.data)
            
            # Conversão automática de coordenadas para endereço (ANTES de salvar endereco_visita)
            if relatorio_express.latitude and relatorio_express.longitude:
                try:
                    from utils import get_address_from_coordinates
                    endereco_convertido = get_address_from_coordinates(relatorio_express.latitude, relatorio_express.longitude)
                    if endereco_convertido:
                        # Usar o endereço convertido em vez do valor do formulário
                        endereco_final = endereco_convertido
                        current_app.logger.info(f"📍 Endereço convertido automaticamente: {endereco_convertido}")
                    else:
                        current_app.logger.warning("⚠️ Não foi possível converter as coordenadas para endereço")
                except Exception as e:
                    current_app.logger.error(f"❌ Erro na conversão de coordenadas para endereço: {str(e)}")
                    # Se a conversão falhar, manter o endereço digitado pelo usuário
                    pass
            
            # Salvar o endereço final (convertido ou original)
            relatorio_express.endereco_visita = endereco_final

            # Observações
            relatorio_express.observacoes_gerais = form.observacoes_gerais.data
            relatorio_express.pendencias = form.pendencias.data
            relatorio_express.recomendacoes = form.recomendacoes.data

            # Checklist (salvar dados JSON)
            if form.checklist_completo.data:
                relatorio_express.checklist_dados = form.checklist_completo.data

            # Status baseado na ação
            if action == 'finalize':
                relatorio_express.status = 'finalizado'
                relatorio_express.finalizado_at = datetime.utcnow()
            else:
                relatorio_express.status = 'rascunho'

            db.session.add(relatorio_express)
            db.session.flush()  # Para obter o ID

            # Salvar participantes selecionados
            if form.participantes.data:
                from models import RelatorioExpressParticipante
                for funcionario_id in form.participantes.data:
                    participante = RelatorioExpressParticipante()
                    participante.relatorio_express_id = relatorio_express.id
                    participante.funcionario_id = funcionario_id
                    db.session.add(participante)

            # Processar fotos configuradas do modal
            foto_configs_str = request.form.get('foto_configuracoes')
            if foto_configs_str:
                try:
                    foto_configs = json.loads(foto_configs_str)
                    upload_folder = app.config.get('UPLOAD_FOLDER', 'uploads')
                    if not os.path.exists(upload_folder):
                        os.makedirs(upload_folder)

                    ordem = 1
                    for config in foto_configs:
                        if config.get('data') and config.get('legenda'):
                            # Salvar imagem do base64
                            import base64
                            image_data = config['data'].split(',')[1]  # Remover prefixo data:image/...
                            image_bytes = base64.b64decode(image_data)

                            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                            filename = f"express_{relatorio_express.id}_{timestamp}_{ordem}.png"
                            foto_path = os.path.join(upload_folder, filename)

                            with open(foto_path, 'wb') as f:
                                f.write(image_bytes)

                            # Criar registro da foto
                            foto_express = FotoRelatorioExpress()
                            foto_express.relatorio_express_id = relatorio_express.id
                            foto_express.filename = filename
                            foto_express.filename_original = config.get('originalName', filename)
                            foto_express.ordem = ordem
                            foto_express.legenda = config['legenda']
                            foto_express.tipo_servico = config.get('categoria', 'Geral')
                            foto_express.imagem = image_bytes  # Salvar dados binários da imagem express

                            db.session.add(foto_express)
                            ordem += 1

                except Exception as e:
                    current_app.logger.error(f"Erro ao processar fotos: {str(e)}")

            # Fallback: processar fotos básicas se não houver configurações
            elif form.fotos.data:
                upload_folder = app.config.get('UPLOAD_FOLDER', 'uploads')
                if not os.path.exists(upload_folder):
                    os.makedirs(upload_folder)

                ordem = 1
                for foto_file in form.fotos.data:
                    if foto_file:
                        # Salvar arquivo
                        filename = secure_filename(foto_file.filename)
                        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                        filename = f"express_{relatorio_express.id}_{timestamp}_{filename}"

                        foto_path = os.path.join(upload_folder, filename)

                        # Ler dados do arquivo antes de salvar
                        file_data = foto_file.read()
                        foto_file.seek(0)  # Reset para salvar o arquivo também
                        foto_file.save(foto_path)

                        # Criar registro da foto
                        foto_express = FotoRelatorioExpress()
                        foto_express.relatorio_express_id = relatorio_express.id
                        foto_express.filename = filename
                        foto_express.filename_original = filename
                        foto_express.ordem = ordem
                        foto_express.legenda = f'Foto {ordem}'
                        foto_express.imagem = file_data  # Salvar dados binários da imagem express básica

                        db.session.add(foto_express)
                        ordem += 1

            db.session.commit()

            if action == 'finalize':
                flash('Relatório Express finalizado com sucesso!', 'success')
                return redirect(url_for('express_detail', id=relatorio_express.id))
            else:
                flash('Rascunho de Relatório Express salvo com sucesso!', 'info')
                return redirect(url_for('express_edit', id=relatorio_express.id))

        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao criar relatório express: {str(e)}', 'error')
            current_app.logger.error(f'Erro ao criar relatório express: {str(e)}')

    # SEMPRE usar template desktop para garantir estilização adequada
    template = 'express/novo.html'
    return render_template(template, form=form, is_mobile=is_mobile, 
                         categorias=categorias_lista, projeto_id=projeto_id)

@app.route('/express/<int:id>')
@login_required
def express_detail(id):
    """Ver detalhes do relatório express"""
    relatorio = RelatorioExpress.query.get_or_404(id)

    # Verificar acesso
    if not current_user.is_master and relatorio.autor_id != current_user.id:
        flash('Acesso negado.', 'error')
        return redirect(url_for('express_list'))

    return render_template('express/detalhes.html', relatorio=relatorio)

@app.route('/express/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def express_edit(id):
    """Editar relatório express (apenas rascunhos)"""
    relatorio = RelatorioExpress.query.get_or_404(id)

    # Verificar acesso
    if not current_user.is_master and relatorio.autor_id != current_user.id:
        flash('Acesso negado.', 'error')
        return redirect(url_for('express_list'))

    # Apenas rascunhos podem ser editados
    if relatorio.status != 'rascunho':
        flash('Apenas rascunhos podem ser editados.', 'warning')
        return redirect(url_for('express_detail', id=id))

    form = RelatorioExpressForm()

    if form.validate_on_submit():
        try:
            action = request.form.get('action', 'save_draft')

            # Atualizar dados da empresa
            relatorio.empresa_nome = form.empresa_nome.data
            relatorio.empresa_endereco = form.empresa_endereco.data
            relatorio.empresa_telefone = form.empresa_telefone.data
            relatorio.empresa_email = form.empresa_email.data
            relatorio.empresa_responsavel = form.empresa_responsavel.data

            # Atualizar dados da visita
            relatorio.data_visita = form.data_visita.data
            relatorio.periodo_inicio = form.periodo_inicio.data
            relatorio.periodo_fim = form.periodo_fim.data
            relatorio.condicoes_climaticas = form.condicoes_climaticas.data
            relatorio.temperatura = form.temperatura.data

            # Localização - processar coordenadas primeiro na edição
            endereco_final = form.endereco_visita.data  # Valor padrão do formulário
            
            if form.latitude.data:
                relatorio.latitude = float(form.latitude.data)
            if form.longitude.data:
                relatorio.longitude = float(form.longitude.data)
            
            # Conversão automática de coordenadas para endereço (ANTES de salvar endereco_visita)
            if relatorio.latitude and relatorio.longitude:
                try:
                    from utils import get_address_from_coordinates
                    endereco_convertido = get_address_from_coordinates(relatorio.latitude, relatorio.longitude)
                    if endereco_convertido:
                        # Usar o endereço convertido em vez do valor do formulário
                        endereco_final = endereco_convertido
                        current_app.logger.info(f"📍 Endereço convertido automaticamente na edição: {endereco_convertido}")
                    else:
                        current_app.logger.warning("⚠️ Não foi possível converter as coordenadas para endereço na edição")
                except Exception as e:
                    current_app.logger.error(f"❌ Erro na conversão de coordenadas para endereço na edição: {str(e)}")
                    # Se a conversão falhar, manter o endereço existente
                    pass
            
            # Salvar o endereço final (convertido ou original)
            relatorio.endereco_visita = endereco_final

            # Atualizar observações
            relatorio.observacoes_gerais = form.observacoes_gerais.data
            relatorio.pendencias = form.pendencias.data
            relatorio.recomendacoes = form.recomendacoes.data

            # Atualizar checklist
            if form.checklist_completo.data:
                relatorio.checklist_dados = form.checklist_completo.data

            # Atualizar participantes
            from models import RelatorioExpressParticipante
            # Remover participantes existentes
            RelatorioExpressParticipante.query.filter_by(relatorio_express_id=relatorio.id).delete()
            # Adicionar novos participantes
            if form.participantes.data:
                for funcionario_id in form.participantes.data:
                    participante = RelatorioExpressParticipante()
                    participante.relatorio_express_id = relatorio.id
                    participante.funcionario_id = funcionario_id
                    db.session.add(participante)

            # Atualizar status se finalizar
            if action == 'finalize':
                relatorio.status = 'finalizado'
                relatorio.finalizado_at = datetime.utcnow()

            relatorio.updated_at = datetime.utcnow()
            db.session.commit()

            if action == 'finalize':
                flash('Relatório Express finalizado com sucesso!', 'success')
                return redirect(url_for('express_detail', id=id))
            else:
                flash('Relatório Express atualizado com sucesso!', 'info')

        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar relatório express: {str(e)}', 'error')

    # Preencher form com dados existentes
    if request.method == 'GET':
        form.empresa_nome.data = relatorio.empresa_nome
        form.empresa_endereco.data = relatorio.empresa_endereco
        form.empresa_telefone.data = relatorio.empresa_telefone
        form.empresa_email.data = relatorio.empresa_email
        form.empresa_responsavel.data = relatorio.empresa_responsavel

        form.data_visita.data = relatorio.data_visita
        form.periodo_inicio.data = relatorio.periodo_inicio
        form.periodo_fim.data = relatorio.periodo_fim
        form.condicoes_climaticas.data = relatorio.condicoes_climaticas
        form.temperatura.data = relatorio.temperatura
        form.endereco_visita.data = relatorio.endereco_visita

        form.latitude.data = str(relatorio.latitude) if relatorio.latitude else ''
        form.longitude.data = str(relatorio.longitude) if relatorio.longitude else ''

        form.observacoes_gerais.data = relatorio.observacoes_gerais
        form.pendencias.data = relatorio.pendencias
        form.recomendacoes.data = relatorio.recomendacoes
        
        # Carregar participantes existentes
        participantes_ids = [p.funcionario_id for p in relatorio.participantes_lista]
        form.participantes.data = participantes_ids

    return render_template('express/novo.html', form=form, relatorio=relatorio, editing=True)

@app.route('/express/<int:id>/pdf')
@login_required
def express_pdf(id):
    """Gerar PDF do relatório express"""
    relatorio = RelatorioExpress.query.get_or_404(id)

    # Verificar acesso
    if not current_user.is_master and relatorio.autor_id != current_user.id:
        flash('Acesso negado.', 'error')
        return redirect(url_for('express_list'))

    # Apenas relatórios finalizados geram PDF
    if relatorio.status != 'finalizado':
        flash('Apenas relatórios finalizados podem gerar PDF.', 'warning')
        return redirect(url_for('express_detail', id=id))

    try:
        # Gerar PDF
        pdf_filename = f"relatorio_express_{relatorio.numero}.pdf"
        pdf_path = os.path.join('uploads', pdf_filename)

        result = gerar_pdf_relatorio_express(relatorio, pdf_path)

        if result.get('success'):
            # Se o PDF foi gerado e retornado como bytes
            pdf_content = result.get('pdf_content')
            if pdf_content:
                from flask import send_file
                return send_file(
                    io.BytesIO(pdf_content),
                    as_attachment=True,
                    download_name=pdf_filename,
                    mimetype='application/pdf'
                )
            else:
                flash('PDF gerado, mas sem conteúdo de arquivo.', 'error')
                return redirect(url_for('express_detail', id=id))
        else:
            flash(f'Erro ao gerar PDF: {result.get("error", "Erro desconhecido")}', 'error')
            return redirect(url_for('express_detail', id=id))

    except Exception as e:
        flash(f'Erro ao gerar PDF: {str(e)}', 'error')
        return redirect(url_for('express_detail', id=id))

@app.route('/express/<int:id>/delete', methods=['POST'])
@login_required
def express_delete(id):
    """Excluir relatório express (apenas master users)"""
    if not current_user.is_master:
        flash('Acesso negado.', 'error')
        return redirect(url_for('express_list'))

    relatorio = RelatorioExpress.query.get_or_404(id)

    try:
        # Excluir fotos associadas
        fotos = FotoRelatorioExpress.query.filter_by(relatorio_express_id=id).all()
        for foto in fotos:
            # Remover arquivo do disco
            upload_folder = app.config.get('UPLOAD_FOLDER', 'uploads')
            foto_path = os.path.join(upload_folder, foto.filename)
            if os.path.exists(foto_path):
                os.remove(foto_path)

            # Remover do banco
            db.session.delete(foto)

        # Excluir relatório
        db.session.delete(relatorio)
        db.session.commit()

        flash(f'Relatório Express {relatorio.numero} excluído com sucesso.', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao excluir relatório: {str(e)}', 'error')

    return redirect(url_for('express_list'))

@app.route('/relatorio-express/<int:relatorio_id>/remover-foto/<int:foto_id>', methods=['POST'])
@login_required
@csrf.exempt
def relatorio_express_remover_foto(relatorio_id, foto_id):
    """Remover foto do relatório express"""
    relatorio = RelatorioExpress.query.get_or_404(relatorio_id)

    # Verificar acesso
    if not current_user.is_master and relatorio.projeto.responsavel_id != current_user.id:
        return jsonify({'error': 'Acesso negado'}), 403

    try:
        foto = FotoRelatorioExpress.query.get_or_404(foto_id)

        # Verificar se a foto pertence ao relatório
        if foto.relatorio_express_id != relatorio_id:
            return jsonify({'error': 'Foto não pertence a este relatório'}), 400

        # Remover arquivo do disco
        foto_path = os.path.join(app.config.get('UPLOAD_FOLDER', 'uploads'), foto.filename)
        if os.path.exists(foto_path):
            os.remove(foto_path)

        # Remover registro
        db.session.delete(foto)
        db.session.commit()

        return jsonify({'success': True})

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Erro ao remover foto: {str(e)}'}), 500

@app.route('/relatorio-express/<int:relatorio_id>/gerar-pdf')
@login_required
def relatorio_express_gerar_pdf(relatorio_id):
    """Gerar PDF do relatório express"""
    relatorio = RelatorioExpress.query.get_or_404(relatorio_id)

    # Verificar acesso
    if not current_user.is_master and relatorio.autor_id != current_user.id:
        flash('Acesso negado.', 'error')
        return redirect(url_for('index'))

    try:
        # Gerar PDF
        pdf_path = gerar_pdf_relatorio_express(relatorio_id)

        # Retornar arquivo para download
        return send_file(pdf_path, as_attachment=True, 
                        download_name=f"relatorio_express_{relatorio.numero}.pdf")

    except Exception as e:
        flash(f'Erro ao gerar PDF: {str(e)}', 'error')
        return redirect(url_for('relatorio_express_detalhes', relatorio_id=relatorio_id))

@app.route('/relatorio-express/<int:relatorio_id>/enviar-email', methods=['GET', 'POST'])
@login_required  
def relatorio_express_enviar_email(relatorio_id):
    """Enviar relatório express por e-mail"""
    relatorio = RelatorioExpress.query.get_or_404(relatorio_id)

    # Verificar acesso
    if not current_user.is_master and relatorio.projeto.responsavel_id != current_user.id:
        flash('Acesso negado.', 'error')
        return redirect(url_for('index'))

    # Buscar e-mails do projeto
    emails_projeto = email_service.buscar_emails_projeto(relatorio.projeto.id)

    if not emails_projeto:
        flash('Nenhuma e-mail cadastrado para esta obra. Cadastre e-mails na seção de clientes.', 'warning')
        return redirect(url_for('relatorio_express_detalhes', relatorio_id=relatorio_id))

    # Configurar formulário
    form = EnvioEmailForm()
    form.destinatarios.choices = [
        (email.email, f"{email.nome_contato} ({email.email}) - {email.cargo or 'N/A'}")
        for email in emails_projeto
    ]

    # Buscar configuração ativa
    config_ativa = email_service.get_configuracao_ativa()
    if not config_ativa:
        flash('Nenhuma configuração de e-mail ativa. Configure o sistema primeiro.', 'error')
        return redirect(url_for('relatorio_express_detalhes', relatorio_id=relatorio_id))

    if request.method == 'POST' and form.validate_on_submit():
        try:
            # Processar e-mails (mesmo processo do relatório normal)
            def processar_emails(texto_emails):
                if not texto_emails:
                    return []
                emails = []
                for linha in texto_emails.split('\n'):
                    for email in linha.split(','):
                        email = email.strip()
                        if email:
                            emails.append(email)
                return emails

            cc_emails = processar_emails(form.cc_emails.data)
            bcc_emails = processar_emails(form.bcc_emails.data)

            # Validar e-mails
            todos_emails = form.destinatarios.data + cc_emails + bcc_emails
            emails_validos, emails_invalidos = email_service.validar_emails(todos_emails)

            if emails_invalidos:
                flash(f'E-mails inválidos: {", ".join(emails_invalidos)}', 'error')
                return render_template('express/enviar_email.html', 
                                     form=form, relatorio=relatorio, config=config_ativa)

            # Enviar e-mails (adaptado para relatório express)
            resultado = enviar_relatorio_express_por_email(
                relatorio, form.destinatarios.data, cc_emails, bcc_emails,
                form.assunto_personalizado.data, form.corpo_personalizado.data,
                current_user.id, config_ativa
            )

            if resultado['success']:
                flash(f'E-mails enviados com sucesso para {resultado["sucessos"]} destinatários!', 'success')
            else:
                flash(f'Erro ao enviar e-mails: {resultado.get("error")}', 'error')

            return redirect(url_for('relatorio_express_detalhes', relatorio_id=relatorio_id))

        except Exception as e:
            flash(f'Erro ao enviar e-mails: {str(e)}', 'error')

    return render_template('express/enviar_email.html', 
                         form=form, relatorio=relatorio, config=config_ativa)

def enviar_relatorio_express_por_email(relatorio, destinatarios, cc_emails, bcc_emails, 
                                     assunto_custom, corpo_custom, usuario_id, config):
    """Função auxiliar para enviar relatório express por e-mail"""
    try:
        # Gerar PDF
        pdf_bytes = gerar_pdf_relatorio_express(relatorio, pdf_path)

        projeto = relatorio.projeto
        data_atual = datetime.now().strftime('%d/%m/%Y')

        # Preparar assunto e corpo
        assunto = assunto_custom or f"Relatório Express - {projeto.nome} - {data_atual}"

        if corpo_custom:
            corpo_html = corpo_custom
        else:
            corpo_html = f"""
            <p>Prezado(a) Cliente,</p>

            <p>Segue em anexo o Relatório Express do projeto <strong>{projeto.nome}</strong>.</p>

            <p>Este relatório contém observações rápidas e fotos da visita realizada.</p>

            <p>Em caso de dúvidas, favor entrar em contato conosco.</p>

            <p>Atenciosamente,<br>
            Equipe ELP Consultoria e Engenharia<br>
            Engenharia Civil & Fachadas</p>
            """

        sucessos = 0
        falhas = 0

        # Configurar Flask-Mail
        email_service.configure_smtp(config)

        # Enviar para cada destinatário
        for email_dest in destinatarios:
            try:
                msg = Message(
                    subject=assunto,
                    recipients=[email_dest],
                    cc=cc_emails,
                    bcc=bcc_emails,
                    html=corpo_html
                )

                # Anexar PDF
                pdf_bytes.seek(0)
                msg.attach(
                    filename=f"relatorio_express_{relatorio.numero}.pdf",
                    content_type='application/pdf',
                    data=pdf_bytes.read()
                )

                email_service.mail.send(msg)
                sucessos += 1

            except Exception as e:
                current_app.logger.error(f"Erro ao enviar email para {email_dest}: {e}")
                falhas += 1

        return {
            'success': sucessos > 0,
            'sucessos': sucessos,
            'falhas': falhas
        }

    except Exception as e:
        return {
            'success': False,
            'error': str(e),
            'sucessos': 0,
            'falhas': len(destinatarios)
        }

# ==================== HELPER: Aprovador Padrão ====================

def get_aprovador_padrao_para_projeto(projeto_id=None):
    """
    Buscar aprovador padrão para um projeto específico ou global
    Nova lógica: prioriza Aprovador Temporário do projeto, caso contrário usa Aprovador Global

    Args:
        projeto_id: ID do projeto (None para buscar apenas global)

    Returns:
        User object do aprovador ou None se não encontrar
    """
    try:
        # Primeiro, tentar encontrar aprovador temporário específico do projeto
        if projeto_id:
            aprovador_temporario = AprovadorPadrao.query.filter_by(
                projeto_id=projeto_id,
                is_global=False,
                ativo=True
            ).order_by(AprovadorPadrao.prioridade.asc(), AprovadorPadrao.created_at.desc()).first()

            if aprovador_temporario and aprovador_temporario.aprovador:
                return aprovador_temporario.aprovador

        # Se não encontrou temporário, buscar aprovador global único
        aprovador_global = AprovadorPadrao.query.filter_by(
            is_global=True,
            ativo=True
        ).first()

        if aprovador_global and aprovador_global.aprovador:
            return aprovador_global.aprovador

        return None

    except Exception as e:
        current_app.logger.error(f"Erro ao buscar aprovador padrão: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def get_aprovador_global():
    """
    Retorna o Aprovador Global atual do sistema (único)
    
    Returns:
        AprovadorPadrao object ou None
    """
    try:
        return AprovadorPadrao.query.filter_by(
            is_global=True,
            ativo=True
        ).first()
    except Exception as e:
        current_app.logger.error(f"Erro ao buscar aprovador global: {str(e)}")
        return None

def current_user_is_aprovador_global():
    """
    Verifica se o usuário atual é o Aprovador Global
    
    IMPORTANTE: Apenas o Aprovador Global tem permissão, não usuários Master.
    Esta verificação é exclusiva para ações críticas de gerenciamento de aprovadores.
    
    Returns:
        Boolean - True se o usuário atual for o Aprovador Global, False caso contrário
    """
    if not current_user or not current_user.is_authenticated:
        return False
    
    aprovador_global = get_aprovador_global()
    return aprovador_global and aprovador_global.aprovador_id == current_user.id

# ==================== ADMIN: Aprovadores Padrão ====================

@app.route('/admin/aprovadores-padrao')
@login_required
def admin_aprovadores_padrao():
    """Gerenciar aprovadores - Aprovador Global (único) e Temporários por projeto"""
    if not current_user.is_master:
        flash('Acesso negado. Apenas usuários master podem acessar esta funcionalidade.', 'error')
        return redirect(url_for('index'))

    # Buscar o Aprovador Global único
    aprovador_global = AprovadorPadrao.query.filter_by(is_global=True, ativo=True).first()
    
    # Buscar Aprovadores Temporários por projeto
    aprovadores_temporarios = AprovadorPadrao.query.filter_by(
        is_global=False,
        ativo=True
    ).order_by(AprovadorPadrao.created_at.desc()).all()

    # Buscar projetos ativos para seleção
    projetos_ativos = Projeto.query.filter_by(status='Ativo').all()

    # Buscar usuários master para seleção como aprovadores
    usuarios_master = User.query.filter_by(is_master=True, ativo=True).all()
    
    # Verificar se o usuário atual é o aprovador global
    is_current_user_aprovador_global = current_user_is_aprovador_global()

    return render_template('admin/aprovadores_padrao.html',
                         aprovador_global=aprovador_global,
                         aprovadores_temporarios=aprovadores_temporarios,
                         projetos_ativos=projetos_ativos,
                         usuarios_master=usuarios_master,
                         is_current_user_aprovador_global=is_current_user_aprovador_global)

@app.route('/admin/aprovadores-padrao/temporario/novo', methods=['GET', 'POST'])
@login_required
def admin_aprovador_temporario_novo():
    """Adicionar novo Aprovador Temporário para um projeto - APENAS Aprovador Global"""
    # REGRA DE PERMISSÃO: Apenas o Aprovador Global pode adicionar aprovadores temporários
    # Mesmo usuários Master não têm permissão para esta ação
    if not current_user_is_aprovador_global():
        flash('Apenas o Aprovador Global pode executar esta ação.', 'error')
        return redirect(url_for('index'))

    if request.method == 'POST':
        try:
            projeto_id = request.form.get('projeto_id')
            aprovador_id = request.form.get('aprovador_id')
            observacoes = request.form.get('observacoes', '').strip()

            # Validações
            if not aprovador_id:
                flash('Aprovador é obrigatório.', 'error')
                return redirect(url_for('admin_aprovador_temporario_novo'))

            if not projeto_id:
                flash('Projeto é obrigatório para Aprovador Temporário.', 'error')
                return redirect(url_for('admin_aprovador_temporario_novo'))

            aprovador_id = int(aprovador_id)
            projeto_id = int(projeto_id)

            # Verificar se já existe aprovador temporário para este projeto
            existing = AprovadorPadrao.query.filter_by(
                projeto_id=projeto_id,
                is_global=False,
                ativo=True
            ).first()

            if existing:
                projeto = Projeto.query.get(projeto_id)
                flash(f'Já existe um Aprovador Temporário configurado para {projeto.nome}.', 'warning')
                return redirect(url_for('admin_aprovadores_padrao'))

            # Criar Aprovador Temporário
            novo_aprovador = AprovadorPadrao(
                is_global=False,
                projeto_id=projeto_id,
                aprovador_id=aprovador_id,
                observacoes=observacoes,
                criado_por=current_user.id
            )

            db.session.add(novo_aprovador)
            db.session.commit()

            projeto = Projeto.query.get(projeto_id)
            flash(f'Aprovador Temporário configurado com sucesso para {projeto.nome}!', 'success')
            return redirect(url_for('admin_aprovadores_padrao'))

        except ValueError:
            flash('Dados inválidos no formulário.', 'error')
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao criar Aprovador Temporário: {str(e)}', 'error')

    # GET request - mostrar formulário
    projetos_ativos = Projeto.query.filter_by(status='Ativo').all()
    usuarios_master = User.query.filter_by(is_master=True, ativo=True).all()

    return render_template('admin/aprovador_temporario_form.html',
                         projetos_ativos=projetos_ativos,
                         usuarios_master=usuarios_master)

@app.route('/admin/aprovadores-padrao/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def admin_aprovador_padrao_editar(id):
    """Editar aprovador padrão - APENAS Aprovador Global"""
    # REGRA DE PERMISSÃO: Apenas o Aprovador Global pode editar configurações de aprovadores
    # Mesmo usuários Master não têm permissão para esta ação
    if not current_user_is_aprovador_global():
        flash('Apenas o Aprovador Global pode executar esta ação.', 'error')
        return redirect(url_for('index'))

    aprovador_padrao = AprovadorPadrao.query.get_or_404(id)

    if request.method == 'POST':
        projeto_id = request.form.get('projeto_id')
        aprovador_id = request.form.get('aprovador_id')
        observacoes = request.form.get('observacoes', '').strip()

        if not aprovador_id:
            flash('Aprovador é obrigatório.', 'error')
            return redirect(url_for('admin_aprovador_padrao_editar', id=id))

        try:
            aprovador_id = int(aprovador_id)
            projeto_id = int(projeto_id) if projeto_id else None

            # Atualizar configuração
            aprovador_padrao.projeto_id = projeto_id
            aprovador_padrao.aprovador_id = aprovador_id
            aprovador_padrao.observacoes = observacoes
            aprovador_padrao.updated_at = datetime.utcnow()

            db.session.commit()

            projeto_nome = aprovador_padrao.projeto.nome if aprovador_padrao.projeto else "Global"
            flash(f'Aprovador padrão atualizado com sucesso para {projeto_nome}!', 'success')
            return redirect(url_for('admin_aprovadores_padrao'))

        except ValueError:
            flash('Dados inválidos no formulário.', 'error')
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar aprovador padrão: {str(e)}', 'error')

    # GET request - mostrar formulário preenchido
    projetos_ativos = Projeto.query.filter_by(status='Ativo').all()
    usuarios_master = User.query.filter_by(is_master=True, ativo=True).all()

    return render_template('admin/aprovador_padrao_form.html',
                         aprovador_padrao=aprovador_padrao,
                         projetos_ativos=projetos_ativos,
                         usuarios_master=usuarios_master,
                         is_edit=True)

@app.route('/admin/aprovadores-padrao/transferir-global', methods=['POST'])
@login_required
def admin_transferir_aprovador_global():
    """Transferir o título de Aprovador Global para outro usuário - APENAS Aprovador Global atual"""
    # REGRA DE PERMISSÃO: Apenas o Aprovador Global atual pode transferir sua função
    # EXCEÇÃO: Se não há Aprovador Global, usuários Master podem definir o primeiro
    aprovador_global_atual = get_aprovador_global()
    
    if aprovador_global_atual:
        # Já existe um Aprovador Global - apenas ele pode transferir
        if not current_user_is_aprovador_global():
            flash('Apenas o Aprovador Global pode executar esta ação.', 'error')
            return redirect(url_for('admin_aprovadores_padrao'))
    else:
        # Não existe Aprovador Global - usuários Master podem definir o primeiro
        if not current_user.is_master:
            flash('Apenas usuários Master podem definir o primeiro Aprovador Global.', 'error')
            return redirect(url_for('admin_aprovadores_padrao'))

    try:
        novo_aprovador_id = request.form.get('novo_aprovador_id')
        
        if not novo_aprovador_id:
            flash('Novo aprovador é obrigatório.', 'error')
            return redirect(url_for('admin_aprovadores_padrao'))
        
        novo_aprovador_id = int(novo_aprovador_id)
        
        # Verificar se o novo aprovador existe e está ativo
        novo_usuario = User.query.get(novo_aprovador_id)
        if not novo_usuario or not novo_usuario.ativo:
            flash('Usuário inválido ou inativo.', 'error')
            return redirect(url_for('admin_aprovadores_padrao'))
        
        # Desativar Aprovador Global atual
        aprovador_global_atual = get_aprovador_global()
        if aprovador_global_atual:
            aprovador_global_atual.ativo = False
            aprovador_global_atual.updated_at = datetime.utcnow()
        
        # Criar novo Aprovador Global
        novo_aprovador_global = AprovadorPadrao(
            is_global=True,
            projeto_id=None,
            aprovador_id=novo_aprovador_id,
            observacoes=f'Transferido de {aprovador_global_atual.aprovador.nome_completo if aprovador_global_atual else "N/A"}',
            criado_por=current_user.id
        )
        
        db.session.add(novo_aprovador_global)
        db.session.commit()
        
        flash(f'Aprovador Global transferido com sucesso para {novo_usuario.nome_completo}!', 'success')
        
    except ValueError:
        flash('Dados inválidos.', 'error')
    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao transferir Aprovador Global: {str(e)}', 'error')
    
    return redirect(url_for('admin_aprovadores_padrao'))

@app.route('/admin/aprovadores-padrao/<int:id>/desativar')
@login_required
def admin_aprovador_padrao_desativar(id):
    """Desativar Aprovador Temporário - APENAS Aprovador Global"""
    # REGRA DE PERMISSÃO: Apenas o Aprovador Global pode remover aprovadores temporários
    # Mesmo usuários Master não têm permissão para esta ação
    if not current_user_is_aprovador_global():
        flash('Apenas o Aprovador Global pode executar esta ação.', 'error')
        return redirect(url_for('admin_aprovadores_padrao'))

    try:
        aprovador_padrao = AprovadorPadrao.query.get_or_404(id)
        
        # Não permitir desativar o Aprovador Global por esta rota
        if aprovador_padrao.is_global:
            flash('Não é possível desativar o Aprovador Global por esta rota. Use a função de transferência.', 'error')
            return redirect(url_for('admin_aprovadores_padrao'))
        
        aprovador_padrao.ativo = False
        aprovador_padrao.updated_at = datetime.utcnow()

        db.session.commit()

        projeto_nome = aprovador_padrao.projeto.nome if aprovador_padrao.projeto else "Sem Projeto"
        flash(f'Aprovador Temporário removido para {projeto_nome}.', 'info')

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao remover Aprovador Temporário: {str(e)}', 'error')

    return redirect(url_for('admin_aprovadores_padrao'))

# ==================== API: Aprovador Padrão ====================

@app.route('/api/aprovador-padrao/<int:projeto_id>')
@login_required
def api_get_aprovador_padrao(projeto_id):
    """API para buscar aprovador padrão de um projeto - AJAX"""
    try:
        aprovador = get_aprovador_padrao_para_projeto(projeto_id)

        if aprovador:
            return jsonify({
                'success': True,
                'aprovador_id': aprovador.id,
                'aprovador_nome': aprovador.nome,
                'aprovador_username': aprovador.username
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Nenhum aprovador padrão configurado'
            })

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Erro ao buscar aprovador padrão: {str(e)}'
        })


# Error handlers
@app.errorhandler(404)
def not_found_error(error):
    return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return render_template('500.html'), 500

# Rota para primeiro login - troca de senha obrigatória
@app.route('/first-login', methods=['GET', 'POST'])
@login_required
def first_login():
    # Se não é primeiro login, redireciona para home
    if not hasattr(current_user, 'primeiro_login') or not current_user.primeiro_login:
        return redirect(url_for('index'))

    form = FirstLoginForm()

    if form.validate_on_submit():
        # Verificar senha atual
        if not check_password_hash(current_user.password_hash, form.current_password.data):
            flash('Senha atual incorreta.', 'error')
            return render_template('auth/first_login.html', form=form)

        # Atualizar senha e marcar como não sendo mais primeiro login
        try:
            current_user.password_hash = generate_password_hash(form.new_password.data)
            current_user.primeiro_login = False
            db.session.commit()

            flash('Senha alterada com sucesso! Bem-vindo ao sistema.', 'success')
            return redirect(url_for('index'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao alterar senha: {str(e)}', 'error')

    return render_template('auth/first_login.html', form=form)




# API endpoint for address geocoding (convert address to coordinates)
@app.route('/api/geocode-address', methods=['POST'])
@csrf.exempt
def geocode_address():
    """Convert address to GPS coordinates using address normalization"""
    try:
        data = request.get_json()
        address = data.get('address')

        if not address or not address.strip():
            return jsonify({
                'success': False, 
                'message': 'Endereço é obrigatório'
            })

        # Use the utility function which now includes address normalization
        latitude, longitude = get_coordinates_from_address(address.strip())

        if latitude and longitude:
            return jsonify({
                'success': True,
                'latitude': latitude,
                'longitude': longitude,
                'message': f'Coordenadas encontradas para o endereço'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Não foi possível encontrar coordenadas para este endereço'
            })

    except Exception as e:
        print(f'Erro no geocoding de endereço: {e}')
        return jsonify({
            'success': False,
            'message': 'Erro interno do servidor'
        })



# ====== PROJECT CHECKLIST ROUTES ======

@app.route("/projects/<int:project_id>/checklist")
@login_required 
def project_checklist_view(project_id):
    """View project checklist configuration"""
    project = Projeto.query.get_or_404(project_id)

    # Get or create checklist config for this project
    config = ProjetoChecklistConfig.query.filter_by(projeto_id=project_id).first()
    if not config:
        # Default to standard checklist
        config = ProjetoChecklistConfig(
            projeto_id=project_id,
            tipo_checklist="padrao",
            criado_por=current_user.id
        )
        db.session.add(config)
        db.session.commit()

    # Get appropriate checklist items
    if config.tipo_checklist == "personalizado":
        checklist_items = ChecklistObra.query.filter_by(
            projeto_id=project_id, 
            ativo=True
        ).order_by(ChecklistObra.ordem).all()
    else:
        checklist_items = ChecklistPadrao.query.filter_by(
            ativo=True
        ).order_by(ChecklistPadrao.ordem).all()

    return render_template("projects/checklist_view.html", 
                         project=project, 
                         config=config,
                         checklist_items=checklist_items)

@app.route("/projects/<int:project_id>/checklist/config", methods=["POST"])
@login_required
@csrf.exempt  
def project_checklist_config(project_id):
    """Configure checklist type for project"""
    project = Projeto.query.get_or_404(project_id)

    try:
        data = request.get_json()
        tipo_checklist = data.get("tipo_checklist")

        if tipo_checklist not in ["padrao", "personalizado"]:
            return jsonify({"error": "Tipo de checklist inválido"}), 400

        # Get or create config
        config = ProjetoChecklistConfig.query.filter_by(projeto_id=project_id).first()
        if not config:
            config = ProjetoChecklistConfig(
                projeto_id=project_id,
                tipo_checklist=tipo_checklist,
                criado_por=current_user.id
            )
            db.session.add(config)
        else:
            config.tipo_checklist = tipo_checklist
            config.updated_at = datetime.utcnow()

        # If switching to personalizado and no custom items exist, create from padrao
        if tipo_checklist == "personalizado":
            existing_custom = ChecklistObra.query.filter_by(projeto_id=project_id).count()
            if existing_custom == 0:
                # Copy from standard checklist
                padrao_items = ChecklistPadrao.query.filter_by(ativo=True).order_by(ChecklistPadrao.ordem).all()
                for item in padrao_items:
                    custom_item = ChecklistObra(
                        projeto_id=project_id,
                        texto=item.texto,
                        ordem=item.ordem,
                        criado_por=current_user.id
                    )
                    db.session.add(custom_item)

        db.session.commit()

        return jsonify({
            "success": True,
            "message": f"Checklist configurado para usar modelo {'personalizado' if tipo_checklist == 'personalizado' else 'padrão'}",
            "tipo_checklist": tipo_checklist,
            "redirect": url_for("project_checklist_edit", project_id=project_id) if tipo_checklist == "personalizado" else None
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Erro interno: {str(e)}"}), 500

@app.route("/projects/<int:project_id>/checklist/edit")
@login_required
def project_checklist_edit(project_id):
    """Edit custom checklist for project"""
    project = Projeto.query.get_or_404(project_id)

    # Check if project uses custom checklist
    config = ProjetoChecklistConfig.query.filter_by(projeto_id=project_id).first()
    if not config or config.tipo_checklist != "personalizado":
        flash("Este projeto não está configurado para usar checklist personalizado", "warning")
        return redirect(url_for("project_view", project_id=project_id))

    # Get custom checklist items
    checklist_items = ChecklistObra.query.filter_by(
        projeto_id=project_id,
        ativo=True
    ).order_by(ChecklistObra.ordem).all()

    return render_template("projects/checklist_edit.html", 
                         project=project,
                         checklist_items=checklist_items)

@app.route("/projects/<int:project_id>/checklist/items", methods=["POST"])
@login_required
@csrf.exempt  
def project_checklist_add_item(project_id):
    """Add new item to custom checklist"""
    project = Projeto.query.get_or_404(project_id)

    try:
        data = request.get_json()
        texto = data.get("texto", "").strip()

        if not texto:
            return jsonify({"error": "Texto é obrigatório"}), 400

        # Check if project uses custom checklist
        config = ProjetoChecklistConfig.query.filter_by(projeto_id=project_id).first()
        if not config or config.tipo_checklist != "personalizado":
            return jsonify({"error": "Projeto não configurado para checklist personalizado"}), 400

        # Get next order
        last_item = ChecklistObra.query.filter_by(
            projeto_id=project_id,
            ativo=True
        ).order_by(ChecklistObra.ordem.desc()).first()
        nova_ordem = (last_item.ordem + 1) if last_item else 1

        # Create new item
        new_item = ChecklistObra(
            projeto_id=project_id,
            texto=texto,
            ordem=nova_ordem,
            criado_por=current_user.id
        )

        db.session.add(new_item)
        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Item adicionado com sucesso",
            "item": {
                "id": new_item.id,
                "texto": new_item.texto,
                "ordem": new_item.ordem
            }
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Erro interno: {str(e)}"}), 500

@app.route("/projects/<int:project_id>/checklist/items/<int:item_id>", methods=["PUT"])
@login_required
@csrf.exempt
def project_checklist_edit_item(project_id, item_id):
    """Edit checklist item"""
    project = Projeto.query.get_or_404(project_id)
    item = ChecklistObra.query.filter_by(
        id=item_id, 
        projeto_id=project_id
    ).first_or_404()

    try:
        data = request.get_json()
        texto = data.get("texto", "").strip()

        if not texto:
            return jsonify({"error": "Texto é obrigatório"}), 400

        item.texto = texto
        item.updated_at = datetime.utcnow()

        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Item atualizado com sucesso",
            "item": {
                "id": item.id,
                "texto": item.texto,
                "ordem": item.ordem
            }
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Erro interno: {str(e)}"}), 500

@app.route("/projects/<int:project_id>/checklist/items/<int:item_id>", methods=["DELETE"]) 
@login_required
@csrf.exempt
def project_checklist_delete_item(project_id, item_id):
    """Delete checklist item"""
    project = Projeto.query.get_or_404(project_id)
    item = ChecklistObra.query.filter_by(
        id=item_id,
        projeto_id=project_id
    ).first_or_404()

    try:
        item.ativo = False
        item.updated_at = datetime.utcnow()

        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Item removido com sucesso"
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Erro interno: {str(e)}"}), 500


# ====== PARTE 5: RELATÓRIO MENSAL DE VISITAS POR USUÁRIO ======

@app.route('/reports/visits-monthly', methods=['GET', 'POST'])
@login_required
def reports_visits_monthly():
    """
    PARTE 5: Relatório Mensal de Visitas por Usuário
    - Permite filtrar por mês, ano e usuário
    - Gera PDF ou Excel conforme selecionado
    """
    from datetime import datetime, timedelta
    from calendar import monthrange
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from flask import make_response, send_file
    
    # Get all active users for filter dropdown
    usuarios = User.query.filter_by(ativo=True).order_by(User.nome_completo).all()
    
    if request.method == 'POST':
        # Process report generation
        try:
            mes = int(request.form.get('mes', datetime.now().month))
            ano = int(request.form.get('ano', datetime.now().year))
            user_id = request.form.get('user_id')
            formato = request.form.get('formato', 'pdf')  # pdf ou excel
            
            # Validar filtros
            if not user_id:
                flash('Selecione um usuário para gerar o relatório', 'warning')
                return redirect(url_for('reports_visits_monthly'))
            
            user = User.query.get_or_404(user_id)
            
            # Calcular período (primeiro e último dia do mês)
            primeiro_dia = datetime(ano, mes, 1)
            ultimo_dia_num = monthrange(ano, mes)[1]
            ultimo_dia = datetime(ano, mes, ultimo_dia_num, 23, 59, 59)
            
            # Buscar visitas do usuário no período (como participante OU responsável)
            visitas_participante = db.session.query(Visita).join(
                VisitaParticipante, Visita.id == VisitaParticipante.visita_id
            ).filter(
                VisitaParticipante.user_id == user_id,
                Visita.data_inicio >= primeiro_dia,
                Visita.data_inicio <= ultimo_dia
            ).all()
            
            visitas_responsavel = Visita.query.filter(
                Visita.responsavel_id == user_id,
                Visita.data_inicio >= primeiro_dia,
                Visita.data_inicio <= ultimo_dia
            ).all()
            
            # Combinar e remover duplicatas
            visitas_ids = set([v.id for v in visitas_participante] + [v.id for v in visitas_responsavel])
            visitas = Visita.query.filter(Visita.id.in_(visitas_ids)).order_by(Visita.data_inicio).all()
            
            # Calcular métricas
            total_visitas = len(visitas)
            visitas_por_status = {}
            visitas_por_projeto = {}
            
            for visita in visitas:
                # Contar por status
                status = visita.status or 'Sem status'
                visitas_por_status[status] = visitas_por_status.get(status, 0) + 1
                
                # Contar por projeto
                projeto_nome = visita.projeto_nome or 'Sem projeto'
                visitas_por_projeto[projeto_nome] = visitas_por_projeto.get(projeto_nome, 0) + 1
            
            # Nome do mês em português
            meses_pt = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                       'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
            mes_nome = meses_pt[mes]
            
            # Gerar relatório conforme formato
            if formato == 'excel':
                return _gerar_relatorio_excel(visitas, user, mes_nome, ano, 
                                             visitas_por_status, visitas_por_projeto, total_visitas)
            else:  # pdf
                return _gerar_relatorio_pdf(visitas, user, mes_nome, ano,
                                           visitas_por_status, visitas_por_projeto, total_visitas)
                                           
        except Exception as e:
            current_app.logger.exception(f"❌ Erro ao gerar relatório mensal: {str(e)}")
            flash(f'Erro ao gerar relatório: {str(e)}', 'danger')
            return redirect(url_for('reports_visits_monthly'))
    
    # GET: Show filter form
    return render_template('reports/visits_monthly.html', usuarios=usuarios)


def _gerar_relatorio_pdf(visitas, user, mes_nome, ano, visitas_por_status, visitas_por_projeto, total_visitas):
    """Gera relatório PDF de visitas mensais"""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from flask import send_file
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=0.5*cm,
        alignment=TA_CENTER
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.HexColor('#4b5563'),
        spaceAfter=1*cm,
        alignment=TA_CENTER
    )
    
    section_style = ParagraphStyle(
        'SectionHeader',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=0.3*cm,
        spaceBefore=0.5*cm
    )
    
    # Título
    elements.append(Paragraph(f"Relatório Mensal de Visitas", title_style))
    elements.append(Paragraph(f"{mes_nome} de {ano} - {user.nome_completo}", subtitle_style))
    
    # Resumo
    elements.append(Paragraph("Resumo do Período", section_style))
    resumo_data = [
        ['Métrica', 'Valor'],
        ['Total de Visitas', str(total_visitas)],
    ]
    
    # Adicionar contagens por status
    for status, count in visitas_por_status.items():
        resumo_data.append([f'Visitas {status}', str(count)])
    
    resumo_table = Table(resumo_data, colWidths=[12*cm, 4*cm])
    resumo_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(resumo_table)
    elements.append(Spacer(1, 0.5*cm))
    
    # Visitas por Projeto
    if visitas_por_projeto:
        elements.append(Paragraph("Visitas por Projeto", section_style))
        projeto_data = [['Projeto', 'Quantidade']]
        for projeto, count in sorted(visitas_por_projeto.items(), key=lambda x: x[1], reverse=True):
            projeto_data.append([projeto, str(count)])
        
        projeto_table = Table(projeto_data, colWidths=[12*cm, 4*cm])
        projeto_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f0fdf4')),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(projeto_table)
        elements.append(Spacer(1, 0.5*cm))
    
    # Detalhamento das Visitas
    if visitas:
        elements.append(Paragraph("Detalhamento das Visitas", section_style))
        visitas_data = [['Data', 'Nº Visita', 'Projeto', 'Status']]
        
        for visita in visitas:
            data_str = visita.data_inicio.strftime('%d/%m/%Y %H:%M')
            numero = visita.numero or f'V{visita.id}'
            projeto = visita.projeto_nome or 'Sem projeto'
            status = visita.status or 'Sem status'
            visitas_data.append([data_str, numero, projeto, status])
        
        visitas_table = Table(visitas_data, colWidths=[4*cm, 3*cm, 6*cm, 3*cm])
        visitas_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f5f5f5')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ]))
        elements.append(visitas_table)
    else:
        elements.append(Paragraph("Nenhuma visita encontrada no período selecionado.", styles['Normal']))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"relatorio_visitas_{user.nome_completo.replace(' ', '_')}_{mes_nome}_{ano}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype='application/pdf')


def _gerar_relatorio_excel(visitas, user, mes_nome, ano, visitas_por_status, visitas_por_projeto, total_visitas):
    """Gera relatório Excel de visitas mensais"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import send_file
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório Visitas"
    
    # Estilos
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=16)
    section_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws['A1'] = "Relatório Mensal de Visitas"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f"{mes_nome} de {ano} - {user.nome_completo}"
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:D2')
    
    # Resumo
    row = 4
    ws[f'A{row}'] = "Resumo do Período"
    ws[f'A{row}'].font = section_font
    row += 1
    
    ws[f'A{row}'] = "Métrica"
    ws[f'B{row}'] = "Valor"
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'B{row}'].fill = header_fill
    row += 1
    
    ws[f'A{row}'] = "Total de Visitas"
    ws[f'B{row}'] = total_visitas
    row += 1
    
    for status, count in visitas_por_status.items():
        ws[f'A{row}'] = f"Visitas {status}"
        ws[f'B{row}'] = count
        row += 1
    
    # Visitas por Projeto
    row += 2
    ws[f'A{row}'] = "Visitas por Projeto"
    ws[f'A{row}'].font = section_font
    row += 1
    
    ws[f'A{row}'] = "Projeto"
    ws[f'B{row}'] = "Quantidade"
    ws[f'A{row}'].font = header_font
    ws[f'B{row}'].font = header_font
    ws[f'A{row}'].fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    ws[f'B{row}'].fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    row += 1
    
    for projeto, count in sorted(visitas_por_projeto.items(), key=lambda x: x[1], reverse=True):
        ws[f'A{row}'] = projeto
        ws[f'B{row}'] = count
        row += 1
    
    # Detalhamento das Visitas
    row += 2
    ws[f'A{row}'] = "Detalhamento das Visitas"
    ws[f'A{row}'].font = section_font
    row += 1
    
    headers = ['Data', 'Nº Visita', 'Projeto', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
        cell.border = border
    row += 1
    
    for visita in visitas:
        ws[f'A{row}'] = visita.data_inicio.strftime('%d/%m/%Y %H:%M')
        ws[f'B{row}'] = visita.numero or f'V{visita.id}'
        ws[f'C{row}'] = visita.projeto_nome or 'Sem projeto'
        ws[f'D{row}'] = visita.status or 'Sem status'
        
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = border
        row += 1
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 15
    
    # Salvar em buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"relatorio_visitas_{user.nome_completo.replace(' ', '_')}_{mes_nome}_{ano}.xlsx"
    return send_file(buffer, as_attachment=True, download_name=filename, 
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')